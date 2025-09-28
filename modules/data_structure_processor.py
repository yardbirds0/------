#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据结构处理器
根据表格结构.md中的算法规范，实现层级关系计算和数据结构处理
"""

import sys
import os
import re
import json
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.data_models import WorkbookManager, WorksheetInfo, TargetItem, SourceItem, SheetType


class DataStructureProcessor:
    """数据结构处理器 - 实现表格结构.md中的层级关系算法"""

    def __init__(self, workbook_manager: WorkbookManager):
        self.workbook_manager = workbook_manager
        self.extracted_data = {
            'metadata': {},
            'targets': [],
            'sources': []
        }

    def _safe_get_sheet_name(self, sheet_item):
        """安全获取工作表名称（鲁棒性处理）

        Args:
            sheet_item: 可能是字符串、WorksheetInfo对象或其他类型

        Returns:
            str: 工作表名称
        """
        if isinstance(sheet_item, str):
            return sheet_item
        elif hasattr(sheet_item, 'name'):
            return str(sheet_item.name)
        else:
            # 兜底处理：转换为字符串
            return str(sheet_item)

    def process_all_data(self) -> Dict[str, Any]:
        """处理所有工作表数据，返回结构化的extracted_data"""

        # 设置元数据
        self.extracted_data['metadata'] = {
            'file_name': self.workbook_manager.file_name,
            'processed_at': self._get_current_time(),
            'total_sheets': len(self.workbook_manager.flash_report_sheets) + len(self.workbook_manager.data_source_sheets)
        }

        # 处理快报表（targets）
        for sheet in self.workbook_manager.flash_report_sheets:
            self._process_flash_report_sheet(sheet)

        # 处理数据来源表（sources）
        for sheet in self.workbook_manager.data_source_sheets:
            self._process_data_source_sheet(sheet)

        # 执行核心算法：生成唯一ID和层级关系计算
        self._generate_unique_ids()
        self._calculate_hierarchical_relationships()
        self._clean_source_names()

        return self.extracted_data

    def _process_flash_report_sheet(self, sheet_info):
        """处理快报表工作表（鲁棒性增强）"""
        try:
            import openpyxl

            # 安全获取工作表名称
            sheet_name = self._safe_get_sheet_name(sheet_info)

            # 加载工作簿和工作表
            workbook = openpyxl.load_workbook(self.workbook_manager.file_path, data_only=True)
            worksheet = workbook[sheet_name]

            row_num = 1
            for row in worksheet.iter_rows(values_only=False):
                cell_value = None
                index_indent = 0

                # 找到第一个非空单元格
                for col_idx, cell in enumerate(row):
                    if cell.value is not None and str(cell.value).strip():
                        cell_value = str(cell.value).strip()
                        # 估算缩进值：假设每个空列相当于缩进2
                        index_indent = col_idx * 2
                        break

                if cell_value:
                    # 创建目标项
                    target_item = {
                        'id': f'Target_sheet_{row_num}',  # 临时ID，后续会被替换
                        'sheet_name': sheet_name,
                        'row': row_num,
                        'item_name': cell_value,
                        'level': index_indent,  # 原始缩进值
                        'original_value': cell_value
                    }

                    self.extracted_data['targets'].append(target_item)

                row_num += 1

        except Exception as e:
            # 安全获取工作表名称用于错误报告
            sheet_name = self._safe_get_sheet_name(sheet_info)
            print(f"处理快报表 {sheet_name} 时发生错误: {e}")

    def _process_data_source_sheet(self, sheet_info):
        """处理数据来源表工作表（鲁棒性增强）"""
        try:
            import openpyxl

            # 安全获取工作表名称
            sheet_name = self._safe_get_sheet_name(sheet_info)

            # 加载工作簿和工作表
            workbook = openpyxl.load_workbook(self.workbook_manager.file_path, data_only=True)
            worksheet = workbook[sheet_name]

            row_num = 1
            for row in worksheet.iter_rows(values_only=False):
                cell_value = None

                # 找到第一个非空单元格
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        cell_value = str(cell.value).strip()
                        break

                if cell_value:
                    # 创建来源项
                    source_item = {
                        'id': f'Source_sheet_{row_num}',  # 临时ID，后续会被替换
                        'sheet_name': sheet_name,
                        'row': row_num,
                        'item_name': cell_value,
                        'original_value': cell_value
                    }

                    self.extracted_data['sources'].append(source_item)

                row_num += 1

        except Exception as e:
            # 安全获取工作表名称用于错误报告
            sheet_name = self._safe_get_sheet_name(sheet_info)
            print(f"处理数据来源表 {sheet_name} 时发生错误: {e}")

    def _generate_unique_ids(self):
        """步骤2：生成全局唯一ID"""

        # 为targets生成唯一ID
        for item in self.extracted_data['targets']:
            unique_id = f"{item['sheet_name']}_{item['row']}"
            item['unique_id'] = unique_id
            # 保留原有的id作为备用
            item['original_id'] = item['id']
            item['id'] = unique_id

        # 为sources生成唯一ID
        for item in self.extracted_data['sources']:
            unique_id = f"{item['sheet_name']}_{item['row']}"
            item['unique_id'] = unique_id
            # 保留原有的id作为备用
            item['original_id'] = item['id']
            item['id'] = unique_id

    def _calculate_hierarchical_relationships(self):
        """步骤4：核心层级关系计算 - 基于栈算法"""

        # 初始化父级栈
        parent_stack = []

        # 按原始顺序遍历targets数组
        for current_item in self.extracted_data['targets']:
            current_level = current_item['level']

            # 寻找父级：弹出所有level >= current_level的项
            while parent_stack and parent_stack[-1]['level'] >= current_level:
                parent_stack.pop()

            # 确定父级并分配层级
            if parent_stack:
                # 有父级
                parent = parent_stack[-1]
                current_item['parent_id'] = parent['unique_id']
                current_item['hierarchical_level'] = parent['hierarchical_level'] + 1
            else:
                # 无父级，是顶级项目
                current_item['parent_id'] = None
                current_item['hierarchical_level'] = 1

            # 将当前项目入栈
            parent_stack.append(current_item)

    def _clean_source_names(self):
        """步骤5：清理数据源名称"""

        for source_item in self.extracted_data['sources']:
            original_name = source_item['item_name']
            cleaned_name = self._apply_cleaning_rules(original_name)
            source_item['cleaned_name'] = cleaned_name

    def _apply_cleaning_rules(self, item_name: str) -> str:
        """应用表格结构.md中规定的清洗规则"""

        cleaned = item_name

        # 去除数字编号和点
        cleaned = re.sub(r'^[一二三四五六七八九十]+、\s*', '', cleaned)
        cleaned = re.sub(r'^\d+\.\s*', '', cleaned)
        cleaned = re.sub(r'^\([一二三四五六七八九十]+\)\s*', '', cleaned)

        # 去除关键词前缀
        prefixes = ['其中：', '其中:', '其中', '加：', '加:', '加', '减：', '减:', '减']
        for prefix in prefixes:
            if cleaned.startswith(prefix):
                cleaned = cleaned[len(prefix):].strip()
                break

        # 去除特殊符号
        cleaned = re.sub(r'[△☆*]', '', cleaned)

        # 替换全角括号为半角
        cleaned = cleaned.replace('（', '(').replace('）', ')')

        # 替换下划线和中文空格为标准空格
        cleaned = cleaned.replace('_', ' ').replace('　', ' ')

        # 去除首尾空格
        cleaned = cleaned.strip()

        return cleaned

    def _get_current_time(self) -> str:
        """获取当前时间字符串"""
        from datetime import datetime
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def save_to_file(self, output_path: Optional[str] = None) -> str:
        """保存处理结果到JSON文件"""

        if output_path is None:
            # 生成默认文件名
            base_name = Path(self.workbook_manager.file_name).stem
            output_path = f"extracted_data_{base_name}.json"

        # 确保输出目录存在
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)

        # 保存JSON文件
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.extracted_data, f, ensure_ascii=False, indent=2)

        return output_path

    def get_hierarchy_tree(self) -> Dict[str, Any]:
        """获取层级树结构（用于调试和可视化）"""

        tree = {}

        # 按层级组织数据
        for item in self.extracted_data['targets']:
            level = item['hierarchical_level']
            if level not in tree:
                tree[level] = []

            tree[level].append({
                'id': item['unique_id'],
                'name': item['item_name'],
                'parent_id': item['parent_id'],
                'original_level': item['level']
            })

        return tree

    def validate_hierarchy(self) -> List[str]:
        """验证层级关系的正确性"""

        errors = []

        # 检查parent_id的有效性
        all_ids = {item['unique_id'] for item in self.extracted_data['targets']}

        for item in self.extracted_data['targets']:
            parent_id = item['parent_id']
            if parent_id is not None and parent_id not in all_ids:
                errors.append(f"项目 {item['unique_id']} 的父级ID {parent_id} 不存在")

        # 检查层级深度的连续性
        for item in self.extracted_data['targets']:
            if item['parent_id'] is not None:
                # 查找父级项目
                parent_item = next(
                    (p for p in self.extracted_data['targets'] if p['unique_id'] == item['parent_id']),
                    None
                )
                if parent_item:
                    expected_level = parent_item['hierarchical_level'] + 1
                    if item['hierarchical_level'] != expected_level:
                        errors.append(
                            f"项目 {item['unique_id']} 的层级 {item['hierarchical_level']} "
                            f"与期望层级 {expected_level} 不符"
                        )

        return errors


def process_workbook_data(workbook_manager: WorkbookManager) -> Dict[str, Any]:
    """工厂函数：处理工作簿数据"""
    processor = DataStructureProcessor(workbook_manager)
    return processor.process_all_data()


def main():
    """测试函数"""
    # 这里可以添加测试代码
    print("数据结构处理器初始化完成")


if __name__ == "__main__":
    main()