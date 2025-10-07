#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强的数据提取器 - 支持科目余额表专用识别和多列数据处理
"""

import sys
import os
import json
import openpyxl
import re
from typing import List, Dict, Optional, Any, Tuple
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.data_models import (
    WorkbookManager,
    TargetItem,
    SourceItem,
    SheetType,
    update_hierarchy_structure,
    TargetColumnEntry,
)
from modules.table_schema_analyzer import TableSchemaAnalyzer, TableType, TableSchema
from utils.column_detector import ColumnDetector

class DataExtractor:
    """增强的数据提取器"""

    def __init__(self, workbook_manager: WorkbookManager):
        """初始化数据提取器"""
        self.workbook_manager = workbook_manager
        self.workbook = None
        self.schema_analyzer = TableSchemaAnalyzer()
        self.column_detector = ColumnDetector()
        self.sheet_column_map: Dict[str, Dict[int, Dict[str, Any]]] = {}

        # 加载表格规则
        self.table_rules = self._load_table_rules()

    def _load_table_rules(self) -> Dict:
        """加载表格规则配置"""
        try:
            rules_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'table_schema_rules.json')
            with open(rules_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print("警告：无法找到表格规则文件，使用默认规则")
            return {"table_schemas": {}}

    def extract_all_data(self) -> bool:
        """提取所有数据"""
        try:
            print("开始提取表格数据...")

            # 加载Excel文件
            if not self._load_workbook():
                return False

            # 提取快报表目标项
            target_count = self._extract_flash_report_targets()
            print(f"提取到目标项: {target_count} 个")

            # 计算层级关系
            print("计算层级关系...")
            if update_hierarchy_structure(self.workbook_manager):
                print("+ 层级关系计算完成")
            else:
                print("X 层级关系计算失败")

            # 提取数据源项（使用增强逻辑）
            source_count = self._extract_data_source_items_enhanced()
            print(f"提取到来源项: {source_count} 个")

            return True

        except Exception as e:
            print(f"数据提取失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _load_workbook(self) -> bool:
        """加载Excel工作簿"""
        try:
            if not self.workbook_manager.file_path or not os.path.exists(self.workbook_manager.file_path):
                print("Excel文件不存在")
                return False

            self.workbook = openpyxl.load_workbook(self.workbook_manager.file_path, data_only=True)
            print(f"Excel文件加载成功: {self.workbook_manager.file_path}")
            return True

        except Exception as e:
            print(f"Excel文件加载失败: {e}")
            return False

    def _extract_flash_report_targets(self) -> int:
        """提取快报表目标项（保持原有逻辑）"""
        target_count = 0

        # 重置目标列元数据
        self.workbook_manager.target_sheet_columns = {}

        for sheet_item in self.workbook_manager.flash_report_sheets:
            sheet_name = self._get_sheet_name(sheet_item)
            print(f"\n提取快报表 '{sheet_name}' 的目标项...")

            if sheet_name not in self.workbook.sheetnames:
                print(f"  工作表 '{sheet_name}' 不存在")
                continue

            sheet = self.workbook[sheet_name]

            table_schema = self.schema_analyzer.analyze_table_schema(sheet)
            print(f"  表头起始行: {table_schema.header_start_row}, 列头行数: {table_schema.header_rows}")

            target_column_map = self._register_sheet_columns(sheet_name, table_schema, registry='target')

            sheet_targets = self._extract_targets_from_sheet(sheet, sheet_name, table_schema, target_column_map)
            target_count += len(sheet_targets)
            print(f"  提取到 {len(sheet_targets)} 个目标项")

            # 添加到工作簿管理器
            for target in sheet_targets:
                self.workbook_manager.target_items[target.id] = target

        return target_count

    def _extract_data_source_items_enhanced(self) -> int:
        """提取数据源项（增强版）"""
        source_count = 0

        # 重置列元数据
        self.workbook_manager.source_sheet_columns = {}
        self.sheet_column_map = {}

        for sheet_item in self.workbook_manager.data_source_sheets:
            sheet_name = self._get_sheet_name(sheet_item)
            print(f"\n提取数据源表 '{sheet_name}' 的来源项...")

            if sheet_name not in self.workbook.sheetnames:
                print(f"  工作表 '{sheet_name}' 不存在")
                continue

            sheet = self.workbook[sheet_name]

            # 分析表格模式
            print(f"  分析表格模式...")
            table_schema = self.schema_analyzer.analyze_table_schema(sheet)
            print(f"  识别为: {table_schema.table_type.value}")
            self._register_sheet_columns(sheet_name, table_schema)

            # 根据表格类型使用不同的提取策略
            if table_schema.table_type == TableType.TRIAL_BALANCE:
                sheet_sources = self._extract_trial_balance_sources(sheet, sheet_name, table_schema)
            else:
                sheet_sources = self._extract_general_sources(sheet, sheet_name, table_schema)

            source_count += len(sheet_sources)
            print(f"  提取到 {len(sheet_sources)} 个来源项")

            # 添加到工作簿管理器
            for source in sheet_sources:
                self.workbook_manager.source_items[source.id] = source

        return source_count

    def _extract_trial_balance_sources(self, sheet, sheet_name: str, schema: TableSchema) -> List[SourceItem]:
        """提取科目余额表来源项（专用逻辑）"""
        sources = []

        print(f"    使用科目余额表专用提取逻辑")
        print(f"    数据开始行: {schema.data_start_row}")
        print(f"    发现 {len(schema.data_columns)} 个数据列")

        # 获取科目余额表的列结构
        trial_balance_structure = self.column_detector.get_trial_balance_structure(
            [col for col in schema.data_columns]
        )

        # 扫描所有数据行（支持大型科目余额表）
        max_row = sheet.max_row or 5000  # 科目余额表通常更大

        for row_num in range(schema.data_start_row, max_row + 1):
            # 提取科目信息
            account_info = self._extract_account_info(sheet, row_num, schema)

            if not account_info:
                continue

            account_code = account_info['code']
            account_name = account_info['name']

            # 确定层级
            hierarchy_level = self._calculate_account_level(account_code)

            # 提取所有数据列的值
            data_columns: Dict[str, Any] = {}
            column_details: Dict[str, Dict[str, Any]] = {}
            main_value: Optional[float] = None
            column_map = self.sheet_column_map.get(sheet_name, {})

            for col_info in schema.data_columns:
                meta = column_map.get(col_info.column_index)
                if not meta:
                    continue

                cell = sheet.cell(row=row_num, column=col_info.column_index)
                value: Optional[Any] = None

                # 先尝试提取数值
                if meta.get("is_data_column") and self._is_data_cell(cell):
                    value = self._extract_cell_value(cell)

                # 如果没有数值，尝试提取文本（确保文本列也能被提取）
                if value is None or value == "":
                    text_value = self._extract_text_value(cell)
                    if text_value and text_value.strip():
                        value = text_value

                if value is None or value == "":
                    continue

                display_name = meta.get("display_name") or self._generate_column_key(col_info, sheet_name)
                data_columns[display_name] = value
                column_details[display_name] = meta

                if meta.get("is_data_column") and main_value is None and isinstance(value, (int, float)):
                    main_value = float(value)

            # 只有找到有效数据才创建来源项
            if data_columns:
                source = self._create_enhanced_source_item(
                    sheet_name=sheet_name,
                    account_name=account_name,
                    account_code=account_code,
                    row_num=row_num,
                    hierarchy_level=hierarchy_level,
                    data_columns=data_columns,
                    main_value=main_value,
                    table_type="trial_balance"
                )
                source.column_info.update(column_details)
                sources.append(source)

        print(f"    科目余额表提取完成，共 {len(sources)} 个项目")
        return sources

    def _extract_general_sources(self, sheet, sheet_name: str, schema: TableSchema) -> List[SourceItem]:
        """提取通用表格来源项"""
        sources = []

        print(f"    使用通用表格提取逻辑")

        # 扫描所有数据行（移除行数限制）
        max_row = sheet.max_row or 2000  # 增加默认上限

        for row_num in range(schema.data_start_row, max_row + 1):
            # 提取项目名称
            item_name = None
            for name_col in schema.name_columns:
                cell = sheet.cell(row=row_num, column=name_col)
                if cell.value and str(cell.value).rstrip():  # 只删除尾部空白，保留前导缩进
                    item_name = str(cell.value).rstrip()  # 保留前导缩进
                    break

            if not item_name:
                continue

            # 提取所有数据列
            data_columns: Dict[str, Any] = {}
            column_details: Dict[str, Dict[str, Any]] = {}
            main_value: Optional[float] = None
            column_map = self.sheet_column_map.get(sheet_name, {})

            for col_info in schema.data_columns:
                meta = column_map.get(col_info.column_index)
                if not meta:
                    continue

                cell = sheet.cell(row=row_num, column=col_info.column_index)
                value: Optional[Any] = None

                # 先尝试提取数值
                if meta.get("is_data_column") and self._is_data_cell(cell):
                    value = self._extract_cell_value(cell)

                # 如果没有数值，尝试提取文本（确保文本列也能被提取）
                if value is None or value == "":
                    text_value = self._extract_text_value(cell)
                    if text_value and text_value.strip():
                        value = text_value

                if value is None or value == "":
                    continue

                display_name = meta.get("display_name") or self._generate_column_key(col_info, sheet_name)
                data_columns[display_name] = value
                column_details[display_name] = meta

                if meta.get("is_data_column") and main_value is None and isinstance(value, (int, float)):
                    main_value = float(value)

            if data_columns:
                source = self._create_enhanced_source_item(
                    sheet_name=sheet_name,
                    account_name=item_name,
                    account_code="",
                    row_num=row_num,
                    hierarchy_level=0,
                    data_columns=data_columns,
                    main_value=main_value,
                    table_type=schema.table_type.value
                )
                source.column_info.update(column_details)
                sources.append(source)

        return sources

    def _extract_account_info(self, sheet, row_num: int, schema: TableSchema) -> Optional[Dict[str, str]]:
        """提取科目信息"""
        account_code = ""
        account_name = ""

        # 从编码列提取科目代码
        for code_col in schema.code_columns:
            cell = sheet.cell(row=row_num, column=code_col)
            if cell.value:
                code_text = str(cell.value).strip()
                # 优化科目代码识别模式
                if re.match(r'^\d{3,12}(\.\d+)*$', code_text):  # 支持3-12位代码，可带小数点分隔
                    account_code = code_text
                    break

        # 从名称列提取科目名称
        for name_col in schema.name_columns:
            cell = sheet.cell(row=row_num, column=name_col)
            if cell.value:
                name_text = str(cell.value).rstrip()  # 保留前导缩进
                if self._is_account_name(name_text):
                    account_name = name_text
                    break

        if account_name or account_code:
            return {
                'code': account_code,
                'name': account_name if account_name else f"科目{account_code}"
            }

        return None

    def _calculate_account_level(self, account_code: str) -> int:
        """计算科目层级（优化版）"""
        if not account_code:
            return 0

        # 支持小数点分隔的科目代码（如1001.01.001）
        if '.' in account_code:
            parts = account_code.split('.')
            return len(parts)  # 按分隔段数确定层级

        # 根据科目代码长度确定层级（传统方式）
        code_length = len(account_code)
        if code_length >= 12:
            return 4
        elif code_length >= 9:
            return 3
        elif code_length >= 6:
            return 2
        elif code_length >= 3:
            return 1
        else:
            return 0

    def _create_enhanced_source_item(self, sheet_name: str, account_name: str, account_code: str,
                                     row_num: int, hierarchy_level: int, data_columns: Dict[str, Any],
                                     main_value: Any, table_type: str) -> SourceItem:
        """创建增强的来源项"""

        # 生成唯一ID
        source_id = f"{sheet_name}_{account_code}_{row_num}" if account_code else f"{sheet_name}_{row_num}"

        # 创建基本的SourceItem
        source = SourceItem(
            id=source_id,
            sheet_name=sheet_name,
            name=account_name,  # ⭐ 保留原始值（包括空格），用于UI显示
            cell_address=f"A{row_num}",  # 暂时使用A列
            row=row_num,
            column="A",
            value=main_value,
            account_code=account_code,
            hierarchy_level=hierarchy_level,
            table_type=table_type,
            data_columns=data_columns
        )

        # ⭐ 关键：将data_columns复制到values字典，用于三段式引用的验证和计算
        source.values = dict(data_columns)  # 创建副本避免共享引用

        # 设置层级信息
        if account_code:
            parent_code = self._get_parent_account_code(account_code)
            source.set_hierarchy_info(account_code, hierarchy_level, parent_code)

        return source

    def _register_sheet_columns(
        self,
        sheet_name: str,
        schema: TableSchema,
        registry: str = 'source'
    ) -> Dict[int, Dict[str, Any]]:
        """注册工作表的列元数据供后续展示与配置使用"""
        metadata_list: List[Dict[str, Any]] = []
        column_map: Dict[int, Dict[str, Any]] = {}

        used_keys: Dict[str, int] = {}
        used_display: Dict[str, int] = {}

        for info in schema.data_columns:
            base_key = info.normalized_key or f"col_{info.column_letter}"
            key = base_key
            if key in used_keys:
                used_keys[key] += 1
                key = f"{base_key}_{used_keys[base_key]}"
            else:
                used_keys[key] = 1

            base_display = info.display_name or info.header_text or f"列{info.column_letter}"
            display = base_display
            if display in used_display:
                used_display[display] += 1
                display = f"{base_display} ({used_display[base_display]})"
            else:
                used_display[display] = 1

            meta_entry = {
                "key": key,
                "display_name": display,
                "column_index": info.column_index,
                "column_letter": info.column_letter,
                "primary_header": info.primary_header,
                "secondary_header": info.secondary_header,
                "header_text": info.header_text,
                "data_type": info.data_type,
                "is_numeric": info.is_numeric,
                "is_data_column": info.is_data_column,
                "is_placeholder": info.is_placeholder,
                "primary_col_span": info.primary_col_span,
                "primary_row_span": info.primary_row_span,
                "primary_is_group_start": info.primary_is_group_start,
                "primary_start_column": info.primary_start_column or info.column_index,
                "secondary_col_span": info.secondary_col_span,
                "secondary_row_span": info.secondary_row_span,
                "secondary_is_group_start": info.secondary_is_group_start,
                "secondary_start_column": info.secondary_start_column or info.column_index,
                "header_row_count": schema.header_rows
            }

            metadata_list.append(meta_entry)
            column_map[info.column_index] = meta_entry

        if registry == 'target':
            self.workbook_manager.target_sheet_columns[sheet_name] = metadata_list
            self.workbook_manager.target_sheet_header_rows[sheet_name] = schema.header_rows
        else:
            self.workbook_manager.source_sheet_columns[sheet_name] = metadata_list
            self.workbook_manager.source_sheet_header_rows[sheet_name] = schema.header_rows
            self.sheet_column_map[sheet_name] = column_map

        return column_map

    def _extract_text_value(self, cell) -> Optional[Any]:
        """提取文本或通用单元格内容"""
        if cell.value is None:
            return None

        if isinstance(cell.value, str):
            text = cell.value.strip()
            return text if text else None

        return cell.value

    def _generate_column_key(self, col_info, sheet_name: str = "") -> str:
        """生成清晰的数据列键名（使用动态列头信息）"""
        meta = self.sheet_column_map.get(sheet_name, {}).get(col_info.column_index)
        if meta:
            return meta.get("display_name") or meta.get("key") or meta.get("header_text")

        if getattr(col_info, 'display_name', None):
            return col_info.display_name

        if getattr(col_info, 'normalized_key', None):
            return col_info.normalized_key

        return f"列{col_info.column_letter}"

    def _get_parent_account_code(self, account_code: str) -> str:
        """获取父级科目代码"""
        if len(account_code) <= 4:
            return ""

        # 根据层级规则返回父级代码
        if len(account_code) >= 6:
            return account_code[:4]  # 返回一级科目
        elif len(account_code) >= 8:
            return account_code[:6]  # 返回二级科目
        elif len(account_code) >= 10:
            return account_code[:8]  # 返回三级科目

        return ""

    def _extract_targets_from_sheet(
        self,
        sheet,
        sheet_name: str,
        schema: TableSchema,
        column_map: Dict[int, Dict[str, Any]]
    ) -> List[TargetItem]:
        """从快报表中提取目标项（增强版）"""
        targets: List[TargetItem] = []
        max_row = sheet.max_row or (schema.data_start_row + 500)

        name_columns = schema.name_columns or [1]
        data_start = schema.data_start_row or (schema.header_start_row + schema.header_rows)

        for row_num in range(data_start, max_row + 1):
            item_text = None
            for name_col in name_columns:
                if name_col <= 0 or name_col > sheet.max_column:
                    continue
                cell = sheet.cell(row=row_num, column=name_col)
                if cell.value and str(cell.value).rstrip():
                    item_text = str(cell.value).rstrip()
                    break

            if not item_text:
                continue

            # 跳过明显描述行
            if any(keyword in item_text for keyword in ['项目', '金额', '单位', '期间']) and row_num <= schema.data_start_row:
                continue

            item_info = self._analyze_target_item_text(item_text)
            if not item_info:
                continue

            target = TargetItem(
                id=f"{sheet_name}_{row_num}",
                name=item_info['clean_name'],
                original_text=item_text,
                sheet_name=sheet_name,
                row=row_num,
                level=item_info['level'],
                hierarchical_level=item_info['level'],
                hierarchical_number=item_info['numbering'],
                display_index=item_info['numbering'],  # 设置原始编号
            )

            column_entries: Dict[str, TargetColumnEntry] = {}
            default_cell = ""

            for col_index, meta in column_map.items():
                column_letter = meta.get("column_letter")
                cell = sheet.cell(row=row_num, column=col_index)
                value = cell.value
                cell_address = f"{column_letter}{row_num}" if column_letter else ""

                entry = TargetColumnEntry(
                    key=meta.get("key"),
                    display_name=meta.get("display_name"),
                    column_index=col_index,
                    column_letter=column_letter,
                    is_numeric=meta.get("is_numeric", False),
                    data_type=meta.get("data_type", "unknown"),
                    header_text=meta.get("header_text", ""),
                    is_data_column=meta.get("is_data_column", False),
                    cell_address=cell_address,
                    source_value=value,
                    is_placeholder=meta.get("is_placeholder", False)
                )
                if entry.key:
                    column_entries[entry.key] = entry
                    if not default_cell and entry.is_data_column and cell_address:
                        default_cell = cell_address

            target.columns = column_entries
            if default_cell:
                target.target_cell_address = default_cell
            elif column_entries:
                first_entry = next(iter(column_entries.values()))
                if first_entry.cell_address:
                    target.target_cell_address = first_entry.cell_address

            targets.append(target)

        return targets

    def _analyze_target_item_text(self, text: str) -> Optional[Dict]:
        """
        分析目标项文本，识别层级关系

        优先级：
        1. 文字表述关系（"其中"、"减"、"加"等关键词）
        2. 缩进关系（前导空格数量）
        """
        if not text or len(text.strip()) < 2:
            return None

        # 1. 检测前导空格（缩进）
        indent_count = len(text) - len(text.lstrip())

        # 2. 检测层级关键词
        # 一级关键词：表示顶级项目
        level_1_keywords = ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']
        level_1_patterns = [
            r'^[一二三四五六七八九十]+、',  # 中文编号
            r'^[（(]?[一二三四五六七八九十]+[)）]',  # 括号中文编号
        ]

        # 二级关键词：表示子项
        level_2_keywords = ['其中：', '其中', '包括：', '包括', '含：', '含']

        # 三级关键词：表示更细的子项
        level_3_keywords = ['减：', '减', '加：', '加', '明细：', '明细', '细项：', '细项']

        stripped_text = text.strip()

        # 3. 检测编号模式
        numbering = ''
        clean_name = stripped_text

        numbering_patterns = [
            r'^(\d+)\.\s*(.+)',  # "1. xxx"
            r'^(\d+)\s+(.+)',    # "1 xxx"
            r'^(\d+)、\s*(.+)',  # "1、xxx"
            r'^\((\d+)\)\s*(.+)',  # "(1) xxx"
        ]

        for pattern in numbering_patterns:
            match = re.match(pattern, stripped_text)
            if match:
                numbering = match.group(1)
                clean_name = match.group(2).rstrip()
                break

        # 4. 根据关键词和缩进计算层级
        # 优先使用文字表述关系判断
        calculated_level = indent_count  # 默认使用缩进值

        # 检查一级关键词
        is_level_1 = False
        for keyword in level_1_keywords:
            if stripped_text.startswith(keyword):
                is_level_1 = True
                calculated_level = 0  # 一级项目缩进为0
                break

        if not is_level_1:
            for pattern in level_1_patterns:
                if re.match(pattern, stripped_text):
                    is_level_1 = True
                    calculated_level = 0
                    break

        # 检查二级关键词（只有在非一级的情况下）
        if not is_level_1:
            for keyword in level_2_keywords:
                if keyword in stripped_text:
                    # 如果有缩进，使用缩进值；否则设置为较小的缩进
                    if indent_count == 0:
                        calculated_level = 2  # 默认二级缩进
                    break

            # 检查三级关键词
            for keyword in level_3_keywords:
                if keyword in stripped_text:
                    # 如果有缩进，使用缩进值；否则设置为更大的缩进
                    if indent_count == 0:
                        calculated_level = 4  # 默认三级缩进
                    break

        return {
            'numbering': numbering,
            'clean_name': clean_name,
            'level': calculated_level  # 使用计算出的层级值
        }

    def _is_account_name(self, text: str) -> bool:
        """判断是否是科目名称（增强版）"""
        if not text or len(text) < 2:
            return False

        # 排除模式（更全面）
        exclude_patterns = [
            r'^日期[:：]', r'^期间[:：]', r'^单位[:：]',
            r'^科目代码$', r'^科目名称$', r'^期初$', r'^期末$',
            r'^借方$', r'^贷方$', r'^合计$', r'^小计$',
            r'^年初$', r'^本期$', r'^余额$', r'^发生额$',
            r'^\d+$',  # 纯数字
            r'^[\d\.,\s\-\(\)]+$'  # 纯数字格式
        ]

        for pattern in exclude_patterns:
            if re.match(pattern, text, re.IGNORECASE):
                return False

        # 包含中文字符且长度合适
        if re.search(r'[\u4e00-\u9fff]', text) and 2 <= len(text) <= 50:
            return True

        return False

    def _is_data_cell(self, cell) -> bool:
        """判断是否是数据单元格（增强版）"""
        if cell.value is None:
            return False

        # 检查是否是数值
        if isinstance(cell.value, (int, float)):
            return True  # 允许0值

        # 检查是否是数值字符串
        if isinstance(cell.value, str):
            try:
                # 处理各种数值格式
                cleaned = cell.value.replace(',', '').replace(' ', '')
                # 处理负数括号格式
                if cleaned.startswith('(') and cleaned.endswith(')'):
                    cleaned = '-' + cleaned[1:-1]
                float(cleaned)
                return True
            except ValueError:
                return False

        return False

    def _extract_cell_value(self, cell) -> Optional[float]:
        """提取单元格数值"""
        if cell.value is None:
            return None

        if isinstance(cell.value, (int, float)):
            return float(cell.value)

        if isinstance(cell.value, str):
            try:
                cleaned = cell.value.replace(',', '').replace(' ', '')
                if cleaned.startswith('(') and cleaned.endswith(')'):
                    cleaned = '-' + cleaned[1:-1]
                return float(cleaned)
            except ValueError:
                return None

        return None

    def _get_sheet_name(self, sheet_item) -> str:
        """获取工作表名称"""
        if isinstance(sheet_item, str):
            return sheet_item
        elif hasattr(sheet_item, 'name'):
            return str(sheet_item.name)
        else:
            return str(sheet_item)

if __name__ == "__main__":
    # 测试用例
    print("增强数据提取器模块")
