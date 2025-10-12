#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件管理模块
负责Excel文件的加载、工作表分类、配置管理等功能
"""

import os
import json
import re
from pathlib import Path
import openpyxl
from typing import List, Dict, Tuple, Optional, Any
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def get_application_directory() -> Path:
    """
    获取应用程序目录
    - 打包成exe后：返回exe所在目录
    - 开发环境：返回项目根目录

    Returns:
        Path: 应用程序目录
    """
    if getattr(sys, 'frozen', False):
        # 打包成exe后，sys.executable是exe文件路径
        return Path(sys.executable).parent
    else:
        # 开发环境，使用__file__的父目录的父目录（项目根目录）
        return Path(__file__).resolve().parents[1]

from models.data_models import (
    WorkbookManager, WorksheetInfo, SheetType,
    TargetItem, SourceItem, MappingFormula, FormulaStatus
)

SENSITIVE_HEADER_KEYS = {
    "authorization",
    "api-key",
    "x-api-key",
    "api_key",
    "bearer",
    "token",
    "secret",
}

_FILENAME_SANITIZE_PATTERN = re.compile(r"[^0-9A-Za-z\u4e00-\u9fff._-]+")

class FileManager:
    """文件管理器类"""

    def __init__(self, classification_callback=None):
        """初始化文件管理器

        Args:
            classification_callback: 工作表分类确认回调函数 (已弃用，现在使用拖拽界面)
        """
        # 使用应用程序目录（打包后是exe所在目录，开发时是项目根目录）
        self.base_dir = get_application_directory()
        self.formula_dir = self.base_dir / "Fomular"
        self.formula_backup_dir = self.formula_dir / "Backup"  # 改为大写Backup
        self.request_history_dir = self.base_dir / "requesthistory"

        self.workbook_manager: Optional[WorkbookManager] = None
        self.current_workbook = None
        self.config_file = "workbook_config.json"
        # 不再使用回调，改为直接自动分类
        # self.classification_callback = classification_callback

    @staticmethod
    def _sanitize_filename(name: str) -> str:
        """将名称转换为安全的文件名。"""
        if not name:
            return "sheet"
        sanitized = _FILENAME_SANITIZE_PATTERN.sub("_", str(name))
        sanitized = sanitized.strip("._-")
        return sanitized or "sheet"

    def sanitize_filename(self, name: str) -> str:
        """对外暴露的文件名规范化方法。"""
        return self._sanitize_filename(name)

    @staticmethod
    def _mask_sensitive_value(value: Any) -> str:
        """对敏感信息进行掩码处理。"""
        text = str(value or "")
        if not text:
            return ""
        if len(text) <= 4:
            return "*" * len(text)
        return f"{text[:4]}***{text[-4:]}"

    def _mask_headers(self, headers: Dict[str, Any]) -> Dict[str, Any]:
        """返回掩码后的请求头。"""
        masked: Dict[str, Any] = {}
        for key, value in (headers or {}).items():
            try:
                if isinstance(value, (dict, list)):
                    masked_value = value
                else:
                    text = str(value)
                    if key.lower() in SENSITIVE_HEADER_KEYS:
                        masked_value = self._mask_sensitive_value(text)
                    else:
                        masked_value = text
            except Exception:
                masked_value = "***"
            masked[key] = masked_value
        return masked

    @staticmethod
    def _ensure_directory(path: Path) -> None:
        path.mkdir(parents=True, exist_ok=True)

    def save_analysis_request_history(
        self,
        *,
        sheet_name: str,
        payload: Dict[str, Any],
        prompt_text: str,
        headers: Dict[str, Any],
        endpoint: Optional[str] = None,
        model: Optional[str] = None,
        warnings: Optional[List[str]] = None,
    ) -> Optional[Path]:
        """将分析预发送请求记录到 requesthistory 目录。"""
        try:
            self._ensure_directory(self.request_history_dir)
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            filename = f"{self._sanitize_filename(sheet_name)}-request-{timestamp}.json"
            file_path = self.request_history_dir / filename

            record = {
                "sheet_name": sheet_name,
                "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                "prompt_preview": prompt_text,
                "payload": payload,
                "headers": self._mask_headers(headers),
                "endpoint": endpoint,
                "model": model,
                "warnings": list(warnings or []),
            }

            tmp_path = file_path.with_name(file_path.name + ".tmp")
            tmp_path.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
            tmp_path.replace(file_path)
            return file_path
        except Exception as exc:  # pragma: no cover - 防御性
            print(f"保存分析请求记录失败: {exc}")
            return None

    def export_formula_snapshot(
        self,
        *,
        sheet_name: str,
        entries: List[Dict[str, Any]],
        destination: Optional[Path] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Optional[Path]:
        """将映射公式导出为 JSON 文件，并在默认目录下进行备份。"""
        if not entries:
            return None

        try:
            if destination is not None:
                destination = Path(destination)
                self._ensure_directory(destination.parent)
            else:
                self._ensure_directory(self.formula_dir)
                destination = self.formula_dir / f"{self._sanitize_filename(sheet_name)}.json"
                if destination.exists():
                    self._ensure_directory(self.formula_backup_dir)
                    backup_name = (
                        f"{self._sanitize_filename(sheet_name)}-"
                        f"{datetime.now().strftime('%Y%m%d-%H%M%S')}-backup.json"
                    )
                    backup_path = self.formula_backup_dir / backup_name
                    try:
                        destination.replace(backup_path)
                    except Exception:
                        # 如果移动失败，尝试复制内容后删除
                        backup_path.write_text(
                            destination.read_text(encoding="utf-8"),
                            encoding="utf-8",
                        )
                        destination.unlink(missing_ok=True)

            snapshot = {
                "version": 1,
                "sheet_name": sheet_name,
                "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                "entries": entries,
                "metadata": metadata or {},
            }

            tmp_path = destination.with_name(destination.name + ".tmp")
            tmp_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
            tmp_path.replace(destination)
            return destination
        except Exception as exc:  # pragma: no cover - 防御性
            print(f"保存公式映射失败: {exc}")
            return None

    def import_formula_snapshot(
        self,
        *,
        sheet_name: str,
        file_path: Path,
    ) -> Dict[str, Any]:
        """从 JSON 文件加载公式映射，并进行基础校验。"""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(str(path))

        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            raise ValueError("导入的 JSON 结构无效。")

        snapshot_sheet = data.get("sheet_name")
        if snapshot_sheet and sheet_name and snapshot_sheet != sheet_name:
            raise ValueError(
                f"JSON 文件属于工作表 '{snapshot_sheet}'，与当前 '{sheet_name}' 不匹配。"
            )

        entries = data.get("entries")
        if not isinstance(entries, list) or not entries:
            mappings = data.get("mappings")
            if isinstance(mappings, dict) and mappings:
                converted_entries: List[Dict[str, Any]] = []
                for target_name, column_map in mappings.items():
                    if not isinstance(column_map, dict):
                        continue
                    for column_name, mapping_info in column_map.items():
                        formula: Optional[str] = None
                        confidence = None
                        reasoning = ""

                        if isinstance(mapping_info, dict):
                            formula = mapping_info.get("formula")
                            confidence = mapping_info.get("confidence")
                            reasoning = mapping_info.get("reasoning") or ""
                        elif isinstance(mapping_info, str):
                            formula = mapping_info

                        if not formula:
                            continue

                        converted_entries.append(
                            {
                                "target_name": target_name,
                                "column_key": column_name,
                                "column_display": column_name,
                                "formula": formula,
                                "confidence": confidence,
                                "reasoning": reasoning,
                            }
                        )

                if converted_entries:
                    entries = converted_entries

        if not isinstance(entries, list) or not entries:
            raise ValueError("导入的 JSON 未包含映射条目。")

        sanitized_entries: List[Dict[str, Any]] = []
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            target_name = entry.get("target_name") or entry.get("target")
            column_key = entry.get("column_key") or entry.get("key")
            formula = entry.get("formula")
            if not target_name or not column_key or not formula:
                continue
            column_display = (
                entry.get("column_display")
                or entry.get("column_name")
                or column_key
            )
            try:
                confidence = entry.get("confidence")
                confidence_value = float(confidence) if confidence is not None else None
            except (TypeError, ValueError):
                confidence_value = None

            sanitized_entries.append(
                {
                    "target_name": str(target_name),
                    "column_key": str(column_key),
                    "column_display": str(column_display),
                    "formula": str(formula),
                    "confidence": confidence_value,
                    "reasoning": entry.get("reasoning") or "",
                }
            )

        if not sanitized_entries:
            raise ValueError("导入的 JSON 未包含有效的映射条目。")

        return {
            "sheet_name": snapshot_sheet or sheet_name,
            "entries": sanitized_entries,
            "metadata": data.get("metadata", {}),
            "generated_at": data.get("generated_at"),
        }

    def load_excel_files(self, file_paths: List[str]) -> Tuple[bool, str]:
        """
        加载Excel文件（支持单文件和多文件模式）

        Args:
            file_paths: Excel文件路径列表

        Returns:
            Tuple[bool, str]: (是否成功, 错误信息)
        """
        try:
            if not file_paths:
                return False, "未选择任何文件"

            # 验证所有文件
            for file_path in file_paths:
                if not os.path.exists(file_path):
                    return False, f"文件不存在: {file_path}"
                if not file_path.lower().endswith(('.xlsx', '.xls')):
                    return False, f"不支持的文件格式: {file_path}"

            # 判断单文件还是多文件模式
            is_multi_file = len(file_paths) > 1

            if is_multi_file:
                # 多文件模式
                print(f"正在加载 {len(file_paths)} 个Excel文件...")
                return self._load_multiple_files(file_paths)
            else:
                # 单文件模式（原有逻辑）
                file_path = file_paths[0]
                print(f"正在加载Excel文件: {file_path}")

                # 加载工作簿
                self.current_workbook = openpyxl.load_workbook(file_path, data_only=True)

                # 创建工作簿管理器
                self.workbook_manager = WorkbookManager(file_path=file_path)
                self.workbook_manager.is_multi_file_mode = False

                # 分析所有工作表
                self._analyze_all_sheets()

                # 自动分类工作表
                self._auto_classify_sheets()

                print(f"成功加载工作簿，包含 {len(self.current_workbook.sheetnames)} 个工作表")
                return True, "文件加载成功"

        except Exception as e:
            error_msg = f"加载文件失败: {str(e)}"
            print(error_msg)
            return False, error_msg

    def _load_multiple_files(self, file_paths: List[str]) -> Tuple[bool, str]:
        """
        加载多个Excel文件并合并

        Args:
            file_paths: 文件路径列表

        Returns:
            Tuple[bool, str]: (是否成功, 消息)
        """
        try:
            # 创建工作簿管理器（多文件模式）
            self.workbook_manager = WorkbookManager(
                file_path=f"merged_{len(file_paths)}_files",  # 虚拟路径
                is_multi_file_mode=True,
                source_files=file_paths.copy()
            )

            all_sheets = []
            sheet_file_map = {}

            # 遍历所有文件，收集所有sheet
            for file_idx, file_path in enumerate(file_paths, 1):
                print(f"  [{file_idx}/{len(file_paths)}] 加载: {Path(file_path).name}")

                wb = openpyxl.load_workbook(file_path, data_only=True)

                for sheet_name in wb.sheetnames:
                    # 处理sheet名称冲突
                    original_name = sheet_name
                    unique_name = sheet_name
                    suffix = 1

                    while unique_name in sheet_file_map:
                        unique_name = f"{original_name}_{suffix}"
                        suffix += 1

                    if unique_name != original_name:
                        print(f"    ⚠️ Sheet名称冲突: '{original_name}' → '{unique_name}'")

                    all_sheets.append(unique_name)
                    sheet_file_map[unique_name] = file_path

                    # 添加到工作簿管理器
                    sheet_info = self.workbook_manager.add_worksheet(unique_name, SheetType.DATA_SOURCE)

                    # 获取sheet的基本信息
                    sheet = wb[original_name]
                    sheet_info.max_row = sheet.max_row or 0
                    sheet_info.max_column = sheet.max_column or 0

                wb.close()

            # 保存sheet到文件的映射
            self.workbook_manager.sheet_file_mapping = sheet_file_map

            print(f"\n合并完成: 共 {len(all_sheets)} 个sheet来自 {len(file_paths)} 个文件")

            # 自动分类所有sheet
            self._auto_classify_sheets_multi_file()

            return True, f"成功加载 {len(file_paths)} 个文件，共 {len(all_sheets)} 个工作表"

        except Exception as e:
            error_msg = f"加载多个文件失败: {str(e)}"
            print(error_msg)
            return False, error_msg

    def _auto_classify_sheets_multi_file(self) -> None:
        """自动分类工作表（多文件模式）"""
        if not self.workbook_manager:
            return

        print("\n=== 自动分类工作表（多文件模式）===")

        # 清空之前的分类
        self.workbook_manager.flash_report_sheets.clear()
        self.workbook_manager.data_source_sheets.clear()

        for sheet_name, sheet_info in self.workbook_manager.worksheets.items():
            # 自动分类
            if self._is_flash_report_sheet(sheet_name):
                sheet_info.sheet_type = SheetType.FLASH_REPORT
                self.workbook_manager.flash_report_sheets.append(sheet_name)
                print(f"  ✅ {sheet_name} → 快报表")
            else:
                sheet_info.sheet_type = SheetType.DATA_SOURCE
                self.workbook_manager.data_source_sheets.append(sheet_name)
                print(f"  ✅ {sheet_name} → 数据来源表")

        print(f"\n快报表数量: {len(self.workbook_manager.flash_report_sheets)}")
        print(f"数据来源表数量: {len(self.workbook_manager.data_source_sheets)}")

    def _analyze_all_sheets(self) -> None:
        """分析所有工作表的基本信息"""
        if not self.current_workbook or not self.workbook_manager:
            return

        for sheet_name in self.current_workbook.sheetnames:
            sheet = self.current_workbook[sheet_name]

            # 统计单元格信息
            total_cells = sheet.max_row * sheet.max_column if sheet.max_row and sheet.max_column else 0
            non_empty_cells = 0
            numeric_cells = 0
            text_cells = 0

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        non_empty_cells += 1
                        if isinstance(cell.value, (int, float)):
                            numeric_cells += 1
                        else:
                            text_cells += 1

            # 计算填充率
            fill_rate = round(non_empty_cells / total_cells * 100, 2) if total_cells > 0 else 0

            # 添加到工作簿管理器
            sheet_info = self.workbook_manager.add_worksheet(sheet_name, SheetType.DATA_SOURCE)

            # 更新额外的元数据
            sheet_info.max_row = sheet.max_row or 0
            sheet_info.max_column = sheet.max_column or 0
            sheet_info.has_merged_cells = len(list(sheet.merged_cells)) > 0
            if sheet.max_column and sheet.max_row:
                try:
                    col_letter = chr(64 + min(sheet.max_column, 26))  # 限制在A-Z范围内
                    sheet_info.data_range = f"A1:{col_letter}{sheet.max_row}"
                except:
                    sheet_info.data_range = "A1:Z1000"  # 默认范围
            else:
                sheet_info.data_range = "A1:A1"

            print(f"分析工作表 '{sheet_name}': {sheet.max_row}x{sheet.max_column}, 填充率: {fill_rate}%")

    def _auto_classify_sheets(self) -> None:
        """自动分类工作表"""
        if not self.workbook_manager:
            return

        print("\n=== 自动分类工作表 ===")

        # 清空之前的分类
        self.workbook_manager.flash_report_sheets.clear()
        self.workbook_manager.data_source_sheets.clear()

        for sheet_name, sheet_info in self.workbook_manager.worksheets.items():
            # 获取用户确认的分类
            classification = self._get_user_sheet_classification(sheet_name)

            # 如果用户选择"全部使用系统建议"，则自动分类所有剩余工作表
            if classification == "auto_all":
                print("用户选择全部使用系统建议，自动分类所有工作表...")
                for remaining_sheet, remaining_info in self.workbook_manager.worksheets.items():
                    auto_type = self._is_flash_report_sheet(remaining_sheet)
                    if auto_type:
                        remaining_info.sheet_type = SheetType.FLASH_REPORT
                        if remaining_sheet not in self.workbook_manager.flash_report_sheets:
                            self.workbook_manager.flash_report_sheets.append(remaining_sheet)
                    else:
                        remaining_info.sheet_type = SheetType.DATA_SOURCE
                        if remaining_sheet not in self.workbook_manager.data_source_sheets:
                            self.workbook_manager.data_source_sheets.append(remaining_sheet)
                    print(f"工作表 '{remaining_sheet}' -> {remaining_info.sheet_type.value}")
                break

            elif classification == "flash_report":
                sheet_info.sheet_type = SheetType.FLASH_REPORT
                if sheet_name not in self.workbook_manager.flash_report_sheets:
                    self.workbook_manager.flash_report_sheets.append(sheet_name)
            elif classification == "data_source":
                sheet_info.sheet_type = SheetType.DATA_SOURCE
                if sheet_name not in self.workbook_manager.data_source_sheets:
                    self.workbook_manager.data_source_sheets.append(sheet_name)
            else:
                # 用户选择跳过
                print(f"工作表 '{sheet_name}' 已跳过识别")

            if classification not in ["skip", "auto_all"]:
                print(f"工作表 '{sheet_name}' -> {sheet_info.sheet_type.value}")

        print(f"快报表数量: {len(self.workbook_manager.flash_report_sheets)}")
        print(f"数据来源表数量: {len(self.workbook_manager.data_source_sheets)}")

        # 更新时间戳
        self.workbook_manager.updated_at = datetime.now().isoformat()

    def _get_user_sheet_classification(self, sheet_name: str) -> str:
        """获取工作表分类（现在直接使用自动分类，不再弹出确认对话框）

        Args:
            sheet_name: 工作表名称

        Returns:
            str: 'flash_report', 'data_source'
        """
        try:
            # 直接进行自动分析，不再要求用户确认
            auto_classification = "快报表" if self._is_flash_report_sheet(sheet_name) else "数据来源表"

            # 根据自动分类返回结果
            if auto_classification == "快报表":
                print(f"✅ 工作表 '{sheet_name}' 自动分类为: 快报表")
                return "flash_report"
            else:
                print(f"✅ 工作表 '{sheet_name}' 自动分类为: 数据来源表")
                return "data_source"

        except Exception as e:
            print(f"分类时出错: {e}")
            # 出错时默认为数据来源表
            return "data_source"

    def _is_flash_report_sheet(self, sheet_name: str) -> bool:
        """
        判断工作表是否为快报表
        只以"快报"两个字为基准进行判断

        Args:
            sheet_name: 工作表名称

        Returns:
            bool: 是否为快报表
        """
        # 只检查是否包含"快报"两个字
        return "快报" in sheet_name

    def adjust_classification_manual(self, adjustments: Dict[str, str]) -> bool:
        """
        手动调整工作表分类

        Args:
            adjustments: 调整映射 {sheet_name: new_type}
                        new_type: 'flash_report' 或 'data_source'

        Returns:
            bool: 调整是否成功
        """
        try:
            if not self.workbook_manager:
                print("错误: 尚未加载工作簿")
                return False

            for sheet_name, new_type in adjustments.items():
                if sheet_name not in self.workbook_manager.worksheets:
                    print(f"警告: 工作表 '{sheet_name}' 不存在")
                    continue

                sheet_info = self.workbook_manager.worksheets[sheet_name]
                old_type = sheet_info.type

                # 更新类型
                if new_type == 'flash_report':
                    sheet_info.type = SheetType.FLASH_REPORT
                    # 从数据来源列表中移除，添加到快报表列表
                    if sheet_name in self.workbook_manager.data_sources:
                        self.workbook_manager.data_sources.remove(sheet_name)
                    if sheet_name not in self.workbook_manager.flash_reports:
                        self.workbook_manager.flash_reports.append(sheet_name)

                elif new_type == 'data_source':
                    sheet_info.type = SheetType.DATA_SOURCE
                    # 从快报表列表中移除，添加到数据来源列表
                    if sheet_name in self.workbook_manager.flash_reports:
                        self.workbook_manager.flash_reports.remove(sheet_name)
                    if sheet_name not in self.workbook_manager.data_sources:
                        self.workbook_manager.data_sources.append(sheet_name)

                print(f"工作表 '{sheet_name}' 分类已调整: {old_type.value} -> {new_type}")

            # 更新时间戳
            self.workbook_manager.updated_at = datetime.now().isoformat()

            print("手动分类调整完成")
            return True

        except Exception as e:
            print(f"调整分类失败: {str(e)}")
            return False

    def save_configuration(self, config_path: str = None) -> bool:
        """
        保存配置信息

        Args:
            config_path: 配置文件路径，默认使用 self.config_file

        Returns:
            bool: 保存是否成功
        """
        try:
            if not self.workbook_manager:
                print("错误: 没有可保存的工作簿配置")
                return False

            if config_path is None:
                config_path = self.config_file

            config_data = {
                'file_info': {
                    'file_path': self.workbook_manager.file_path,
                    'file_name': self.workbook_manager.file_name,
                    'saved_at': datetime.now().isoformat()
                },
                'classification': {
                    'flash_reports': self.workbook_manager.flash_reports,
                    'data_sources': self.workbook_manager.data_sources
                },
                'worksheets_info': {
                    name: {
                        'type': info.type.value,
                        'max_row': info.max_row,
                        'max_column': info.max_column,
                        'fill_rate': info.fill_rate
                    }
                    for name, info in self.workbook_manager.worksheets.items()
                }
            }

            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, ensure_ascii=False, indent=2)

            print(f"配置已保存到: {config_path}")
            return True

        except Exception as e:
            print(f"保存配置失败: {str(e)}")
            return False

    def load_configuration(self, config_path: str = None) -> bool:
        """
        加载配置信息

        Args:
            config_path: 配置文件路径

        Returns:
            bool: 加载是否成功
        """
        try:
            if config_path is None:
                config_path = self.config_file

            if not os.path.exists(config_path):
                print(f"配置文件不存在: {config_path}")
                return False

            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)

            # 验证配置文件格式
            if 'file_info' not in config_data or 'classification' not in config_data:
                print("配置文件格式不正确")
                return False

            file_path = config_data['file_info']['file_path']

            # 重新加载文件
            success, error_msg = self.load_excel_files([file_path])
            if not success:
                print(f"无法重新加载文件: {error_msg}")
                return False

            # 应用保存的分类
            flash_reports = config_data['classification']['flash_reports']
            data_sources = config_data['classification']['data_sources']

            # 构建调整映射
            adjustments = {}
            for sheet_name in flash_reports:
                adjustments[sheet_name] = 'flash_report'
            for sheet_name in data_sources:
                adjustments[sheet_name] = 'data_source'

            # 应用调整
            self.adjust_classification_manual(adjustments)

            print(f"配置已从 {config_path} 加载")
            return True

        except Exception as e:
            print(f"加载配置失败: {str(e)}")
            return False

    def get_workbook_summary(self) -> Dict[str, Any]:
        """
        获取工作簿摘要信息

        Returns:
            Dict[str, Any]: 摘要信息
        """
        if not self.workbook_manager:
            return {}

        return self.workbook_manager.export_summary()

    def get_worksheets_by_type(self, sheet_type: SheetType) -> List[str]:
        """
        获取指定类型的工作表列表

        Args:
            sheet_type: 工作表类型

        Returns:
            List[str]: 工作表名称列表
        """
        if not self.workbook_manager:
            return []

        if sheet_type == SheetType.FLASH_REPORT:
            return self.workbook_manager.flash_reports.copy()
        else:
            return self.workbook_manager.data_sources.copy()

    def get_workbook_manager(self) -> Optional[WorkbookManager]:
        """获取工作簿管理器实例"""
        return self.workbook_manager

    def _get_mapping_save_path(self, workbook_manager: WorkbookManager) -> Optional[str]:
        """
        根据工作簿生成映射公式保存路径
        保存到程序目录/Fomular/Backup/{文件名}.mapping.json

        Args:
            workbook_manager: 工作簿管理器

        Returns:
            Optional[str]: 保存路径，如果失败返回None
        """
        if not workbook_manager or not workbook_manager.file_path:
            return None

        # 获取Excel文件名（不含路径）
        excel_file_name = Path(workbook_manager.file_path).stem

        # 保存到程序目录/Fomular/Backup/
        self._ensure_directory(self.formula_backup_dir)
        mapping_file = self.formula_backup_dir / f"{excel_file_name}.mapping.json"

        return str(mapping_file)

    def save_mapping_formulas(self, workbook_manager: WorkbookManager) -> bool:
        """将映射公式保存到本地JSON文件"""
        try:
            save_path = self._get_mapping_save_path(workbook_manager)
            if not save_path:
                return False

            mapping_formulas = []
            for target_id, column_map in workbook_manager.mapping_formulas.items():
                for column_key, formula in column_map.items():
                    if not formula or not isinstance(formula, MappingFormula):
                        continue
                    mapping_formulas.append({
                        "target_id": target_id,
                        "column_key": column_key,
                        "column_name": formula.column_name,
                        "formula": formula.formula,
                        "constant_value": formula.constant_value,
                        "status": formula.status.value if hasattr(formula, 'status') else FormulaStatus.USER_MODIFIED.value,
                        "ai_confidence": getattr(formula, "ai_confidence", 0.0),
                        "ai_reasoning": getattr(formula, "ai_reasoning", ""),
                        "notes": getattr(formula, 'notes', "")
                    })

            data = {
                "excel_path": workbook_manager.file_path,
                "saved_at": datetime.now().isoformat(),
                "mapping_formulas": mapping_formulas,
                "column_configs": workbook_manager.column_configs  # ✅ 保存列配置
            }

            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return True
        except Exception as e:
            print(f"自动保存映射公式失败: {e}")
            return False

    def load_saved_formulas(self, workbook_manager: WorkbookManager) -> int:
        """加载已保存的映射公式"""
        try:
            load_path = self._get_mapping_save_path(workbook_manager)
            if not load_path or not os.path.exists(load_path):
                return 0

            with open(load_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # ✅ 加载列配置
            column_configs = data.get("column_configs", {})
            workbook_manager.column_configs = column_configs

            formulas = data.get("mapping_formulas", [])
            applied_count = 0

            for entry in formulas:
                target_id = entry.get("target_id")
                if not target_id or target_id not in workbook_manager.target_items:
                    continue

                column_key = entry.get("column_key", "__default__")
                column_name = entry.get("column_name", "")
                formula_text = entry.get("formula", "")
                constant_value = entry.get("constant_value")
                status_value = entry.get("status", FormulaStatus.USER_MODIFIED.value)
                notes = entry.get("notes", "")
                ai_confidence = entry.get("ai_confidence")
                ai_reasoning = entry.get("ai_reasoning", "")

                try:
                    status = FormulaStatus(status_value)
                except ValueError:
                    status = FormulaStatus.USER_MODIFIED

                mapping = workbook_manager.ensure_mapping(target_id, column_key, column_name)
                mapping.update_formula(formula_text, status=status, column_name=column_name)
                mapping.constant_value = constant_value
                mapping.notes = notes
                if ai_confidence is not None:
                    try:
                        mapping.ai_confidence = max(0.0, min(1.0, float(ai_confidence)))
                    except (TypeError, ValueError):
                        mapping.ai_confidence = 0.0
                mapping.ai_reasoning = ai_reasoning

                applied_count += 1

            return applied_count
        except Exception as e:
            print(f"加载映射公式失败: {e}")
            return 0


class FileManagerUI:
    """文件管理器用户界面"""

    def __init__(self, parent=None):
        """初始化用户界面"""
        self.file_manager = FileManager()
        self.parent = parent
        self.window = None

        # UI组件
        self.flash_reports_listbox = None
        self.data_sources_listbox = None
        self.summary_text = None

    def create_ui(self) -> tk.Toplevel:
        """
        创建文件管理器界面

        Returns:
            tk.Toplevel: 窗口对象
        """
        # 创建窗口
        self.window = tk.Toplevel(self.parent) if self.parent else tk.Tk()
        self.window.title("文件管理器 - Excel文件加载与分类")
        self.window.geometry("800x600")

        # 创建主框架
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 配置网格权重
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)

        # 文件操作区
        file_frame = ttk.LabelFrame(main_frame, text="文件操作", padding="5")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))

        ttk.Button(file_frame, text="选择Excel文件", command=self._select_file).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(file_frame, text="保存配置", command=self._save_config).grid(row=0, column=1, padx=(0, 5))
        ttk.Button(file_frame, text="加载配置", command=self._load_config).grid(row=0, column=2, padx=(0, 5))

        # 工作表分类区
        classification_frame = ttk.LabelFrame(main_frame, text="工作表分类", padding="5")
        classification_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        classification_frame.columnconfigure(0, weight=1)
        classification_frame.columnconfigure(1, weight=1)

        # 快报表列表
        flash_frame = ttk.Frame(classification_frame)
        flash_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))

        ttk.Label(flash_frame, text="快报表").grid(row=0, column=0, sticky=tk.W)
        self.flash_reports_listbox = tk.Listbox(flash_frame, height=8)
        self.flash_reports_listbox.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        flash_scrollbar = ttk.Scrollbar(flash_frame, orient="vertical")
        flash_scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.flash_reports_listbox.config(yscrollcommand=flash_scrollbar.set)
        flash_scrollbar.config(command=self.flash_reports_listbox.yview)

        flash_frame.columnconfigure(0, weight=1)
        flash_frame.rowconfigure(1, weight=1)

        # 数据来源表列表
        data_frame = ttk.Frame(classification_frame)
        data_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))

        ttk.Label(data_frame, text="数据来源表").grid(row=0, column=0, sticky=tk.W)
        self.data_sources_listbox = tk.Listbox(data_frame, height=8)
        self.data_sources_listbox.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        data_scrollbar = ttk.Scrollbar(data_frame, orient="vertical")
        data_scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.data_sources_listbox.config(yscrollcommand=data_scrollbar.set)
        data_scrollbar.config(command=self.data_sources_listbox.yview)

        data_frame.columnconfigure(0, weight=1)
        data_frame.rowconfigure(1, weight=1)

        # 调整按钮区
        button_frame = ttk.Frame(classification_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=(10, 0))

        ttk.Button(button_frame, text="← 移至快报表", command=self._move_to_flash).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(button_frame, text="移至数据源 →", command=self._move_to_data).grid(row=0, column=1, padx=(5, 0))

        # 摘要信息区
        summary_frame = ttk.LabelFrame(main_frame, text="工作簿摘要", padding="5")
        summary_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        summary_frame.columnconfigure(0, weight=1)
        summary_frame.rowconfigure(0, weight=1)

        self.summary_text = tk.Text(summary_frame, height=10, wrap=tk.WORD)
        self.summary_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        summary_scrollbar = ttk.Scrollbar(summary_frame, orient="vertical")
        summary_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.summary_text.config(yscrollcommand=summary_scrollbar.set)
        summary_scrollbar.config(command=self.summary_text.yview)

        return self.window

    def _select_file(self):
        """选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if file_path:
            success, message = self.file_manager.load_excel_files([file_path])
            if success:
                self._update_display()
                messagebox.showinfo("成功", message)
            else:
                messagebox.showerror("错误", message)

    def _update_display(self):
        """更新显示内容"""
        # 清空列表
        self.flash_reports_listbox.delete(0, tk.END)
        self.data_sources_listbox.delete(0, tk.END)

        # 更新工作表列表
        flash_reports = self.file_manager.get_worksheets_by_type(SheetType.FLASH_REPORT)
        data_sources = self.file_manager.get_worksheets_by_type(SheetType.DATA_SOURCE)

        for sheet_name in flash_reports:
            self.flash_reports_listbox.insert(tk.END, sheet_name)

        for sheet_name in data_sources:
            self.data_sources_listbox.insert(tk.END, sheet_name)

        # 更新摘要
        summary = self.file_manager.get_workbook_summary()
        self._display_summary(summary)

    def _display_summary(self, summary: Dict[str, Any]):
        """显示摘要信息"""
        self.summary_text.delete(1.0, tk.END)

        summary_text = f"""
文件信息:
  路径: {summary.get('file_info', {}).get('path', 'N/A')}
  文件名: {summary.get('file_info', {}).get('name', 'N/A')}
  分析时间: {summary.get('file_info', {}).get('analyzed_at', 'N/A')}

工作表统计:
  总数: {summary.get('worksheets', {}).get('total', 0)}
  快报表: {summary.get('worksheets', {}).get('flash_reports', 0)}
  数据来源表: {summary.get('worksheets', {}).get('data_sources', 0)}

数据项统计:
  目标项: {summary.get('data_items', {}).get('target_items', 0)}
  空目标项: {summary.get('data_items', {}).get('empty_targets', 0)}
  来源项: {summary.get('data_items', {}).get('source_items', 0)}
  已映射项: {summary.get('data_items', {}).get('mapped_items', 0)}

快报表列表:
{chr(10).join(f"  - {sheet}" for sheet in summary.get('flash_reports_list', []))}

数据来源表列表:
{chr(10).join(f"  - {sheet}" for sheet in summary.get('data_sources_list', []))}
"""

        self.summary_text.insert(tk.END, summary_text.strip())

    def _move_to_flash(self):
        """移动选中的表到快报表"""
        selection = self.data_sources_listbox.curselection()
        if selection:
            sheet_name = self.data_sources_listbox.get(selection[0])
            adjustments = {sheet_name: 'flash_report'}
            if self.file_manager.adjust_classification_manual(adjustments):
                self._update_display()

    def _move_to_data(self):
        """移动选中的表到数据来源表"""
        selection = self.flash_reports_listbox.curselection()
        if selection:
            sheet_name = self.flash_reports_listbox.get(selection[0])
            adjustments = {sheet_name: 'data_source'}
            if self.file_manager.adjust_classification_manual(adjustments):
                self._update_display()

    def _save_config(self):
        """保存配置"""
        if self.file_manager.save_configuration():
            messagebox.showinfo("成功", "配置已保存")
        else:
            messagebox.showerror("错误", "保存配置失败")

    def _load_config(self):
        """加载配置"""
        config_path = filedialog.askopenfilename(
            title="选择配置文件",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if config_path:
            if self.file_manager.load_configuration(config_path):
                self._update_display()
                messagebox.showinfo("成功", "配置已加载")
            else:
                messagebox.showerror("错误", "加载配置失败")


def main():
    """主函数，用于测试文件管理模块"""
    print("文件管理模块测试")
    print("="*50)

    # 创建文件管理器实例
    file_manager = FileManager()

    # 测试加载Excel文件
    test_file = "（科电）国资委、财政快报模板-纯净版 的副本.xlsx"
    success, message = file_manager.load_excel_files([test_file])

    if success:
        print(f"✅ {message}")

        # 显示工作簿摘要
        summary = file_manager.get_workbook_summary()
        print("\n📊 工作簿摘要:")
        for key, value in summary.items():
            print(f"  {key}: {value}")

        # 测试手动分类调整
        print("\n🔄 测试分类调整:")
        adjustments = {
            "Sheet1": "flash_report"  # 将Sheet1改为快报表
        }
        if file_manager.adjust_classification_manual(adjustments):
            print("✅ 分类调整成功")

        # 测试保存配置
        if file_manager.save_configuration():
            print("✅ 配置保存成功")

        print("\n📋 最终分类结果:")
        print(f"快报表: {file_manager.get_worksheets_by_type(SheetType.FLASH_REPORT)}")
        print(f"数据来源表: {file_manager.get_worksheets_by_type(SheetType.DATA_SOURCE)}")

    else:
        print(f"❌ {message}")

    print("\n" + "="*50)
    print("文件管理模块测试完成")


if __name__ == "__main__":
    main()
