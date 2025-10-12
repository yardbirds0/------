#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 公式导出模块
负责将转换后的公式写入Excel文件
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import numbers, Alignment
from openpyxl.utils import get_column_letter
import os
import shutil
import re
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field
from datetime import datetime
import json

from models.data_models import (
    WorkbookManager, TargetItem, MappingFormula, FormulaStatus
)
from modules.formula_converter import (
    FormulaConverter, ExcelReference, ConversionError, SecurityValidator
)


@dataclass
class ExportOptions:
    """导出配置选项"""
    include_metadata_sheet: bool = True
    preserve_values_on_error: bool = True
    auto_validate: bool = False  # 默认关闭自动验证(需要Excel计算)
    error_handling_mode: str = "preserve"  # "preserve" | "skip" | "fail"
    use_absolute_path: bool = True
    add_formula_comments: bool = False  # 是否添加公式注释


@dataclass
class ExportResult:
    """导出结果报告"""
    success: bool
    output_file_path: str
    total_formulas: int
    converted_formulas: int
    failed_conversions: List[ConversionError]
    execution_time: float
    validation_report: Optional[Dict[str, Any]] = None

    def get_summary_message(self) -> str:
        """生成用户友好的总结信息"""
        if self.success:
            success_rate = (self.converted_formulas / self.total_formulas * 100) if self.total_formulas > 0 else 0
            msg = (
                f"导出成功!\n"
                f"文件路径: {self.output_file_path}\n"
                f"总公式数: {self.total_formulas}\n"
                f"成功转换: {self.converted_formulas} ({success_rate:.1f}%)\n"
                f"失败转换: {len(self.failed_conversions)}\n"
                f"耗时: {self.execution_time:.2f}秒"
            )
            if self.validation_report:
                msg += f"\n\n验证结果:\n"
                msg += f"- 语法检查: {'通过' if self.validation_report.get('syntax_valid', False) else '失败'}\n"
                msg += f"- 循环引用: {'无' if not self.validation_report.get('circular_references', []) else '检测到'}"
            return msg
        else:
            return (
                f"导出失败!\n"
                f"总公式数: {self.total_formulas}\n"
                f"失败转换: {len(self.failed_conversions)}\n"
                f"请检查错误日志获取详细信息"
            )


class ValidationEngine:
    """公式验证引擎"""

    def __init__(self):
        """初始化验证引擎"""
        self.security_validator = SecurityValidator()

    def validate_exported_file(
        self,
        exported_file_path: str,
        original_target_items: List[TargetItem],
        converter: FormulaConverter
    ) -> Dict[str, Any]:
        """
        验证导出文件

        Args:
            exported_file_path: 导出的文件路径
            original_target_items: 原始目标项列表
            converter: 公式转换器 (用于获取循环引用检测器)

        Returns:
            验证报告字典
        """
        report = {
            "syntax_valid": True,
            "syntax_errors": [],
            "circular_references": [],
            "sample_calculation_valid": True,
            "calculation_mismatches": [],
            "overall_valid": True
        }

        try:
            # 1. 验证文件是否存在和可读
            if not os.path.exists(exported_file_path):
                report["overall_valid"] = False
                report["syntax_errors"].append("导出文件不存在")
                return report

            # 2. 打开工作簿进行验证
            wb = openpyxl.load_workbook(exported_file_path, data_only=False)

            # 3. 语法验证
            syntax_errors = self._check_formula_syntax(wb)
            if syntax_errors:
                report["syntax_valid"] = False
                report["syntax_errors"] = syntax_errors
                report["overall_valid"] = False

            # 4. 检查循环引用
            circular_refs = converter.circular_detector.validate_all()
            if circular_refs:
                report["circular_references"] = [" -> ".join(cycle) for cycle in circular_refs]
                report["overall_valid"] = False

            wb.close()

        except Exception as e:
            report["overall_valid"] = False
            report["syntax_errors"].append(f"验证过程中发生错误: {str(e)}")

        return report

    def _check_formula_syntax(self, workbook: Workbook) -> List[str]:
        """
        检查公式语法

        Args:
            workbook: 工作簿对象

        Returns:
            语法错误列表
        """
        errors = []

        try:
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            # 检查基本语法
                            formula = cell.value

                            # 检查括号匹配
                            open_count = formula.count('(')
                            close_count = formula.count(')')
                            if open_count != close_count:
                                errors.append(
                                    f"{sheet_name}!{cell.coordinate}: 括号不匹配"
                                )

                            # 检查公式长度
                            if len(formula) > 8192:
                                errors.append(
                                    f"{sheet_name}!{cell.coordinate}: 公式长度超过限制"
                                )

        except Exception as e:
            errors.append(f"语法检查过程中发生错误: {str(e)}")

        return errors

    def _sample_calculation_check(
        self,
        workbook: Workbook,
        original_items: List[TargetItem],
        sample_ratio: float = 0.1
    ) -> List[str]:
        """
        抽样计算验证

        Args:
            workbook: 工作簿对象
            original_items: 原始目标项列表
            sample_ratio: 抽样比例

        Returns:
            计算不匹配列表
        """
        mismatches = []

        # 简化实现: 不进行实际计算比较
        # 实际应用中需要Excel计算引擎或者手动实现公式计算
        # 这里只是框架

        return mismatches


class ExcelFormulaWriter:
    """Excel公式写入器"""

    def __init__(self):
        """初始化写入器"""
        self.last_export_result: Optional[ExportResult] = None
        self.formula_converter: Optional[FormulaConverter] = None
        self.security_validator = SecurityValidator()
        self.validation_engine = ValidationEngine()

    def export_with_formulas(
        self,
        workbook_manager: WorkbookManager,
        target_sheet_name: str,
        output_path: str,
        options: ExportOptions = ExportOptions()
    ) -> ExportResult:
        """
        导出包含公式的Excel文件

        Args:
            workbook_manager: 工作簿管理器
            target_sheet_name: 目标工作表名称
            output_path: 输出文件��径
            options: 导出选项

        Returns:
            ExportResult: 导出结果
        """
        start_time = datetime.now()

        # 初始化转换器
        self.formula_converter = FormulaConverter(workbook_manager)

        # 统计信息
        total_formulas = 0
        converted_formulas = 0
        failed_conversions: List[ConversionError] = []

        try:
            # 0. 验证输出路径安全性
            is_valid, error_msg = self.security_validator.validate_path(output_path)
            if not is_valid:
                raise ValueError(f"输出路径验证失败: {error_msg}")

            # 1. 加载源工作簿
            source_wb = openpyxl.load_workbook(workbook_manager.file_path, data_only=False)

            # 2. 克隆工作簿结构
            print(f"正在克隆工作簿结构...")
            output_wb = self._clone_workbook_structure(source_wb)

            # 3. 获取目标工作表
            if target_sheet_name not in output_wb.sheetnames:
                raise ValueError(f"目标工作表 '{target_sheet_name}' 不存在")

            target_ws = output_wb[target_sheet_name]

            # 4. 收集目标工作表的所有公式
            formulas_to_process = {}
            for target_id, target_item in workbook_manager.target_items.items():
                if target_item.sheet_name != target_sheet_name:
                    continue

                # 获取映射公式
                mapping = workbook_manager.mapping_formulas.get(target_id)
                if mapping:
                    formulas_to_process[target_id] = mapping

            # 正确计算公式总数（嵌套字典结构）
            total_formulas = sum(len(column_map) if isinstance(column_map, dict) else 1
                               for column_map in formulas_to_process.values())
            print(f"找到 {total_formulas} 个公式需要转换...")

            # 5. 批量转换公式
            conversion_results = self.formula_converter.batch_convert(
                formulas_to_process,
                use_absolute_path=options.use_absolute_path
            )

            # 6. 写入公式到Excel
            print(f"正在写入公式到Excel...")
            for result_key, (excel_formula, references, errors) in conversion_results.items():
                # 解析 result_key
                if '#' in result_key:
                    target_id, column_key = result_key.split('#', 1)
                else:
                    target_id = result_key
                    column_key = "__default__"

                target_item = workbook_manager.target_items.get(target_id)
                if not target_item:
                    continue

                # 获取对应的mapping_formula
                mapping_dict = workbook_manager.mapping_formulas.get(target_id)
                if isinstance(mapping_dict, dict):
                    mapping_formula = mapping_dict.get(column_key)
                else:
                    mapping_formula = mapping_dict

                if not mapping_formula:
                    continue

                # 确定正确的单元格地址
                cell_address = None
                if column_key != "__default__" and hasattr(target_item, 'columns') and target_item.columns:
                    # 多列模式：从columns字典中获取列地址
                    column_entry = target_item.columns.get(column_key)
                    if column_entry and hasattr(column_entry, 'cell_address'):
                        cell_address = column_entry.cell_address

                # 后备方案：使用target_cell_address
                if not cell_address:
                    cell_address = target_item.target_cell_address

                # 如果仍然没有地址，跳过
                if not cell_address:
                    print(f"警告: 目标项 {target_item.name} 列 {column_key} 没有单元格地址，跳过")
                    continue

                # 处理转换错误
                if errors:
                    failed_conversions.extend(errors)
                    if options.error_handling_mode == "preserve":
                        # 使用计算好的值
                        value = mapping_formula.calculation_result
                        if value is not None:
                            self._write_value_to_cell(
                                target_ws,
                                cell_address,
                                value,
                                add_comment=f"原公式转换失败: {errors[0].error_message}"
                            )
                    elif options.error_handling_mode == "skip":
                        # 跳过该单元格
                        continue
                    else:  # fail
                        # 停止导出
                        raise Exception(f"公式转换失败: {errors[0].error_message}")
                else:
                    # 成功转换,写入公式
                    if excel_formula:
                        self._write_formula_to_cell(
                            target_ws,
                            cell_address,
                            excel_formula,
                            fallback_value=mapping_formula.calculation_result,
                            add_comment=options.add_formula_comments
                        )
                        converted_formulas += 1

            # 7. 添加元数据表 (可选)
            if options.include_metadata_sheet:
                self._add_metadata_sheet(
                    output_wb,
                    workbook_manager,
                    target_sheet_name,
                    total_formulas,
                    converted_formulas,
                    failed_conversions
                )

            # 8. 设置Excel为自动计算模式
            output_wb.calculation.calcMode = 'auto'

            # 9. 保存文件
            print(f"正在保存文件到: {output_path}")
            output_wb.save(output_path)
            output_wb.close()
            source_wb.close()

            # 10. 生成失败报告文件 (如果有失败的转换)
            if failed_conversions:
                self._generate_failure_report(
                    output_path,
                    workbook_manager,
                    target_sheet_name,
                    total_formulas,
                    converted_formulas,
                    failed_conversions
                )

            # 11. 自动验证 (如果启用)
            validation_report = None
            if options.auto_validate:
                print("正在验证导出的文件...")
                target_items_list = [item for item in workbook_manager.target_items.values()
                                   if item.sheet_name == target_sheet_name]
                validation_report = self.validation_engine.validate_exported_file(
                    output_path,
                    target_items_list,
                    self.formula_converter
                )
                if not validation_report.get("overall_valid", False):
                    print(f"⚠️  验证发现问题: {validation_report}")

            # 12. 计算耗时
            end_time = datetime.now()
            execution_time = (end_time - start_time).total_seconds()

            # 13. 创建结果报告
            result = ExportResult(
                success=True,
                output_file_path=output_path,
                total_formulas=total_formulas,
                converted_formulas=converted_formulas,
                failed_conversions=failed_conversions,
                execution_time=execution_time,
                validation_report=validation_report
            )

            self.last_export_result = result
            print(result.get_summary_message())
            return result

        except Exception as e:
            end_time = datetime.now()
            execution_time = (end_time - start_time).total_seconds()

            result = ExportResult(
                success=False,
                output_file_path=output_path,
                total_formulas=total_formulas,
                converted_formulas=converted_formulas,
                failed_conversions=failed_conversions,
                execution_time=execution_time
            )

            print(f"导出失败: {str(e)}")
            import traceback
            traceback.print_exc()

            self.last_export_result = result
            return result

    def _clone_workbook_structure(self, source_wb: Workbook) -> Workbook:
        """
        克隆工作簿结构

        Args:
            source_wb: 源工作簿

        Returns:
            新工作簿
        """
        # 创建新工作簿
        output_wb = Workbook()

        # 删除默认工作表
        if 'Sheet' in output_wb.sheetnames:
            output_wb.remove(output_wb['Sheet'])

        # 复制所有工作表
        for sheet_name in source_wb.sheetnames:
            source_ws = source_wb[sheet_name]
            target_ws = output_wb.create_sheet(sheet_name)

            # 复制单元格值和格式
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws[cell.coordinate]

                    # 复制值
                    if cell.value is not None:
                        target_cell.value = cell.value

                    # 复制格式
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy()
                        target_cell.alignment = cell.alignment.copy()

            # 复制列宽
            for col in source_ws.column_dimensions:
                if col in source_ws.column_dimensions:
                    target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

            # 复制行高
            for row in source_ws.row_dimensions:
                if row in source_ws.row_dimensions:
                    target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

            # 复制合并单元格
            for merged_cell in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merged_cell))

        return output_wb

    def _create_merged_workbook(self, workbook_manager: WorkbookManager) -> Workbook:
        """
        创建合并的工作簿（多文件模式）

        从多个源文件中复制所有sheet到一个新工作簿

        Args:
            workbook_manager: 工作簿管理器（包含sheet_file_mapping）

        Returns:
            合并后的工作簿
        """
        # 创建新工作簿
        output_wb = Workbook()

        # 删除默认工作表
        if 'Sheet' in output_wb.sheetnames:
            output_wb.remove(output_wb['Sheet'])

        # 遍历所有需要的sheet
        all_sheets = list(workbook_manager.flash_report_sheets) + list(workbook_manager.data_source_sheets)

        for sheet_idx, sheet_name in enumerate(all_sheets, 1):
            # 获取该sheet来自哪个文件
            source_file = workbook_manager.sheet_file_mapping.get(sheet_name)
            if not source_file:
                print(f"  ⚠️ 未找到sheet '{sheet_name}' 的源文件，跳过")
                continue

            try:
                print(f"  [{sheet_idx}/{len(all_sheets)}] 复制sheet: {sheet_name}")

                # 加载源文件
                source_wb = openpyxl.load_workbook(source_file, data_only=False)

                # 找到源sheet（处理重命名情况）
                source_sheet_name = sheet_name
                if sheet_name not in source_wb.sheetnames:
                    # 可能被重命名了（如"利润表_1" → "利润表"）
                    base_name = re.sub(r'_\d+$', '', sheet_name)
                    if base_name in source_wb.sheetnames:
                        source_sheet_name = base_name
                    elif len(source_wb.sheetnames) == 1:
                        # 单sheet文件，直接使用第一个
                        source_sheet_name = source_wb.sheetnames[0]
                    else:
                        print(f"    ⚠️ 在文件中未找到sheet，跳过")
                        source_wb.close()
                        continue

                source_ws = source_wb[source_sheet_name]

                # 在输出工作簿中创建新sheet
                target_ws = output_wb.create_sheet(sheet_name)

                # 复制单元格值和格式
                for row in source_ws.iter_rows():
                    for cell in row:
                        target_cell = target_ws[cell.coordinate]

                        # 复制值
                        if cell.value is not None:
                            target_cell.value = cell.value

                        # 复制格式
                        if cell.has_style:
                            try:
                                target_cell.font = cell.font.copy()
                                target_cell.border = cell.border.copy()
                                target_cell.fill = cell.fill.copy()
                                target_cell.number_format = cell.number_format
                                target_cell.protection = cell.protection.copy()
                                target_cell.alignment = cell.alignment.copy()
                            except:
                                pass  # 忽略格式复制错误

                # 复制列宽
                for col in source_ws.column_dimensions:
                    if col in source_ws.column_dimensions:
                        target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

                # 复制行高
                for row in source_ws.row_dimensions:
                    if row in source_ws.row_dimensions:
                        target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

                # 复制合并单元格
                for merged_cell in source_ws.merged_cells.ranges:
                    target_ws.merge_cells(str(merged_cell))

                source_wb.close()

            except Exception as e:
                print(f"    ❌ 复制sheet失败: {e}")
                continue

        print(f"合并完成: 共复制 {len(output_wb.sheetnames)} 个sheet")

        return output_wb

    def _write_formula_to_cell(
        self,
        worksheet,
        cell_address: str,
        formula_string: str,
        fallback_value: Any = None,
        add_comment: bool = False
    ) -> None:
        """
        写入公式到单元格

        Args:
            worksheet: 工作表对象
            cell_address: 单元格地址
            formula_string: 公式字符串
            fallback_value: 后备值
            add_comment: 是否添加注释
        """
        if not cell_address:
            return

        cell = worksheet[cell_address]

        # 写入公式
        cell.value = formula_string

        # 保留格式
        if not cell.number_format or cell.number_format == 'General':
            cell.number_format = '0.00'

        if not cell.alignment:
            cell.alignment = Alignment(horizontal='right', vertical='center')

        # 添加注释 (可选)
        if add_comment and fallback_value is not None:
            from openpyxl.comments import Comment
            comment_text = f"原始值: {fallback_value}\n由AI映射工具生成"
            cell.comment = Comment(comment_text, "Formula Exporter")

    def _write_value_to_cell(
        self,
        worksheet,
        cell_address: str,
        value: Any,
        add_comment: str = ""
    ) -> None:
        """
        写入值到单元格

        Args:
            worksheet: 工作表对象
            cell_address: 单元格地址
            value: 值
            add_comment: 注释文本
        """
        if not cell_address:
            return

        cell = worksheet[cell_address]

        # 写入值
        cell.value = value

        # 保留格式
        if not cell.number_format or cell.number_format == 'General':
            cell.number_format = '0.00'

        if not cell.alignment:
            cell.alignment = Alignment(horizontal='right', vertical='center')

        # 添加注释 (如果有)
        if add_comment:
            from openpyxl.comments import Comment
            cell.comment = Comment(add_comment, "Formula Exporter")

    def _add_metadata_sheet(
        self,
        workbook: Workbook,
        workbook_manager: WorkbookManager,
        target_sheet_name: str,
        total_formulas: int,
        converted_formulas: int,
        failed_conversions: List[ConversionError]
    ) -> None:
        """
        添加元数据工作表

        Args:
            workbook: 工作簿
            workbook_manager: 工作簿管理器
            target_sheet_name: 目标工作表名称
            total_formulas: 总公式数
            converted_formulas: 成功转换的公式数
            failed_conversions: 失败的转换
        """
        # 创建元数据工作表
        if "_Export_Metadata" in workbook.sheetnames:
            ws = workbook["_Export_Metadata"]
        else:
            ws = workbook.create_sheet("_Export_Metadata")

        # 写入基本信息
        ws['A1'] = "公式导出元数据"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)

        ws['A3'] = "导出时间:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws['A4'] = "源文件:"
        ws['B4'] = workbook_manager.file_path

        ws['A5'] = "目标工作表:"
        ws['B5'] = target_sheet_name

        ws['A7'] = "统计信息"
        ws['A7'].font = openpyxl.styles.Font(bold=True)

        ws['A8'] = "总公式数:"
        ws['B8'] = total_formulas

        ws['A9'] = "成功转换:"
        ws['B9'] = converted_formulas

        ws['A10'] = "失败转换:"
        ws['B10'] = len(failed_conversions)

        success_rate = (converted_formulas / total_formulas * 100) if total_formulas > 0 else 0
        ws['A11'] = "成功率:"
        ws['B11'] = f"{success_rate:.1f}%"

        # 写入错误列表
        if failed_conversions:
            ws['A13'] = "转换失败的公式"
            ws['A13'].font = openpyxl.styles.Font(bold=True)

            ws['A14'] = "目标项"
            ws['B14'] = "错误类型"
            ws['C14'] = "错误信息"

            row = 15
            for error in failed_conversions[:100]:  # 最多显示100个错误
                target_name = error.target_item.name if error.target_item else "未知"
                ws[f'A{row}'] = target_name
                ws[f'B{row}'] = error.error_type
                ws[f'C{row}'] = error.error_message
                row += 1

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 60

    def _generate_failure_report(
        self,
        output_path: str,
        workbook_manager: WorkbookManager,
        target_sheet_name: str,
        total_formulas: int,
        converted_formulas: int,
        failed_conversions: List[ConversionError]
    ) -> None:
        """
        生成失败报告txt文件

        Args:
            output_path: Excel输出文件路径
            workbook_manager: 工作簿管理器
            target_sheet_name: 目标工作表名称
            total_formulas: 总公式数
            converted_formulas: 成功转换的公式数
            failed_conversions: 失败的转换列表
        """
        # 生成报告文件名（与Excel文件同名，扩展名为txt）
        import os
        base_path = os.path.splitext(output_path)[0]
        report_path = f"{base_path}_导出失败报告.txt"

        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                # 写入标题
                f.write("="*80 + "\n")
                f.write("Excel公式导出失败报告\n")
                f.write("="*80 + "\n\n")

                # 写入导出信息
                f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"源文件: {workbook_manager.file_path}\n")
                f.write(f"目标工作表: {target_sheet_name}\n")
                f.write(f"导出文件: {output_path}\n\n")

                # 写入统计信息
                f.write("-"*80 + "\n")
                f.write("统计信息\n")
                f.write("-"*80 + "\n")
                success_rate = (converted_formulas / total_formulas * 100) if total_formulas > 0 else 0
                f.write(f"总公式数: {total_formulas}\n")
                f.write(f"成功转换: {converted_formulas} ({success_rate:.1f}%)\n")
                f.write(f"失败转换: {len(failed_conversions)} ({100-success_rate:.1f}%)\n\n")

                # 写入失败原因说明
                f.write("-"*80 + "\n")
                f.write("失败原因说明\n")
                f.write("-"*80 + "\n")
                f.write("以下单元格的公式转换失败，已使用计算值代替：\n\n")

                # 按错误类型分组统计
                error_types = {}
                for error in failed_conversions:
                    error_type = error.error_type
                    if error_type not in error_types:
                        error_types[error_type] = []
                    error_types[error_type].append(error)

                # 写入错误类型统计
                f.write("错误类型统计：\n")
                for error_type, errors in error_types.items():
                    f.write(f"  - {error_type}: {len(errors)} 个\n")
                f.write("\n")

                # 写入详细的失败列表
                f.write("="*80 + "\n")
                f.write("详细失败列表\n")
                f.write("="*80 + "\n\n")

                for i, error in enumerate(failed_conversions, 1):
                    # 获取目标项信息
                    target_item = error.target_item
                    if target_item:
                        target_name = target_item.name
                        target_sheet = target_item.sheet_name
                        target_cell = target_item.target_cell_address or "未知"

                        # 尝试获取列信息
                        column_info = ""
                        if hasattr(target_item, 'columns') and target_item.columns:
                            # 多列模式，尝试找到对应的列
                            for col_key, col_entry in target_item.columns.items():
                                if hasattr(col_entry, 'cell_address'):
                                    column_info = f" ({col_key})"
                                    target_cell = col_entry.cell_address
                                    break
                    else:
                        target_name = "未知"
                        target_sheet = "未知"
                        target_cell = "未知"
                        column_info = ""

                    # 写入失败信息
                    f.write(f"{i}. 目标项: {target_name}{column_info}\n")
                    f.write(f"   工作表: {target_sheet}\n")
                    f.write(f"   单元格: {target_cell}\n")
                    f.write(f"   错误类型: {error.error_type}\n")
                    f.write(f"   错误信息: {error.error_message}\n")
                    f.write(f"   原始公式: {error.internal_formula}\n")
                    f.write(f"   处理方式: {error.fallback_action}\n")

                    # 如果使用了计算值，显示计算值
                    if error.fallback_action == "used_value" and target_item:
                        # 尝试从workbook_manager获取计算结果
                        mapping_formula = None
                        for target_id, mapping_dict in workbook_manager.mapping_formulas.items():
                            if workbook_manager.target_items.get(target_id) == target_item:
                                if isinstance(mapping_dict, dict):
                                    # 多列模式，找到第一个匹配的
                                    for col_key, formula in mapping_dict.items():
                                        if formula.formula == error.internal_formula:
                                            mapping_formula = formula
                                            break
                                else:
                                    mapping_formula = mapping_dict
                                break

                        if mapping_formula and mapping_formula.calculation_result is not None:
                            f.write(f"   使用值: {mapping_formula.calculation_result}\n")

                    f.write("\n")

                # 写入结尾说明
                f.write("="*80 + "\n")
                f.write("说明\n")
                f.write("="*80 + "\n")
                f.write("1. 公式转换失败的原因通常包括：\n")
                f.write("   - cell_not_found: 未找到引用的源数据项或单元格\n")
                f.write("   - syntax_error: 公式语法错误\n")
                f.write("   - reference_error: 引用解析错误\n")
                f.write("   - security_error: 安全验证失败（可能包含危险字符）\n")
                f.write("   - cell_bounds_error: 单元格地址超出Excel限制\n\n")
                f.write("2. 失败的单元格已使用计算值填充，确保数据完整性。\n\n")
                f.write("3. 建议检查源数据工作表，确保所有引用的项目都存在。\n\n")
                f.write("4. 如需重新生成公式，请修正源数据后重新导出。\n")

            print(f"失败报告已生成: {report_path}")

        except Exception as e:
            print(f"生成失败报告时出错: {str(e)}")
            import traceback
            traceback.print_exc()

    def get_last_export_result(self) -> Optional[ExportResult]:
        """获取最后一次导出结果"""
        return self.last_export_result

    def export_all_flash_reports_with_formulas(
        self,
        workbook_manager: WorkbookManager,
        output_path: str,
        options: ExportOptions = ExportOptions()
    ) -> ExportResult:
        """
        导出所有待写入表的公式到Excel文件

        Args:
            workbook_manager: 工作簿管理器
            output_path: 输出文件路径
            options: 导出选项

        Returns:
            ExportResult: 导出结果
        """
        start_time = datetime.now()

        # 初始化转换器
        self.formula_converter = FormulaConverter(workbook_manager)

        # 统计信息（所有表的总计）
        total_formulas = 0
        converted_formulas = 0
        failed_conversions: List[ConversionError] = []

        try:
            # 0. 验证输出路径安全性
            is_valid, error_msg = self.security_validator.validate_path(output_path)
            if not is_valid:
                raise ValueError(f"输出路径验证失败: {error_msg}")

            # 1. 判断模式并加载/合并工作簿
            if workbook_manager.is_multi_file_mode:
                # 多文件模式：合并多个源文件
                print(f"多文件模式: 合并 {len(workbook_manager.source_files)} 个源文件...")
                output_wb = self._create_merged_workbook(workbook_manager)
            else:
                # 单文件模式：克隆单个源文件
                source_wb = openpyxl.load_workbook(workbook_manager.file_path, data_only=False)
                print(f"正在克隆工作簿结构...")
                output_wb = self._clone_workbook_structure(source_wb)
                source_wb.close()

            # 2. 获取所有待写入表
            flash_report_sheets = workbook_manager.flash_report_sheets or []
            if not flash_report_sheets:
                raise ValueError("没有待写入表（快报表），无法导出")

            print(f"找到 {len(flash_report_sheets)} 个待写入表: {flash_report_sheets}")

            # 4. 遍历每个待写入表，处理其公式
            for sheet_name in flash_report_sheets:
                if sheet_name not in output_wb.sheetnames:
                    print(f"警告: 工作表 '{sheet_name}' 不存在，跳过")
                    continue

                target_ws = output_wb[sheet_name]
                print(f"\n处理工作表: {sheet_name}")

                # 4.1 收集该工作表的所有公式
                formulas_to_process = {}
                for target_id, target_item in workbook_manager.target_items.items():
                    if target_item.sheet_name != sheet_name:
                        continue

                    # 获取映射公式
                    mapping = workbook_manager.mapping_formulas.get(target_id)
                    if mapping:
                        formulas_to_process[target_id] = mapping

                # 计算该表的公式总数
                sheet_formulas_count = sum(len(column_map) if isinstance(column_map, dict) else 1
                                          for column_map in formulas_to_process.values())
                total_formulas += sheet_formulas_count
                print(f"  找到 {sheet_formulas_count} 个公式需要转换...")

                if sheet_formulas_count == 0:
                    print(f"  该表没有公式，跳过")
                    continue

                # 4.2 批量转换公式
                conversion_results = self.formula_converter.batch_convert(
                    formulas_to_process,
                    use_absolute_path=options.use_absolute_path
                )

                # 4.3 写入公式到Excel
                print(f"  正在写入公式到Excel...")
                sheet_converted_count = 0
                for result_key, (excel_formula, references, errors) in conversion_results.items():
                    # 解析 result_key
                    if '#' in result_key:
                        target_id, column_key = result_key.split('#', 1)
                    else:
                        target_id = result_key
                        column_key = "__default__"

                    target_item = workbook_manager.target_items.get(target_id)
                    if not target_item:
                        continue

                    # 获取对应的mapping_formula
                    mapping_dict = workbook_manager.mapping_formulas.get(target_id)
                    if isinstance(mapping_dict, dict):
                        mapping_formula = mapping_dict.get(column_key)
                    else:
                        mapping_formula = mapping_dict

                    if not mapping_formula:
                        continue

                    # 确定正确的单元格地址
                    cell_address = None
                    if column_key != "__default__" and hasattr(target_item, 'columns') and target_item.columns:
                        # 多列模式：从columns字典中获取列地址
                        column_entry = target_item.columns.get(column_key)
                        if column_entry and hasattr(column_entry, 'cell_address'):
                            cell_address = column_entry.cell_address

                    # 后备方案：使用target_cell_address
                    if not cell_address:
                        cell_address = target_item.target_cell_address

                    # 如果仍然没有地址，跳过
                    if not cell_address:
                        print(f"  警告: 目标项 {target_item.name} 列 {column_key} 没有单元格地址，跳过")
                        continue

                    # 处理转换错误
                    if errors:
                        failed_conversions.extend(errors)
                        if options.error_handling_mode == "preserve":
                            # 使用计算好的值
                            value = mapping_formula.calculation_result
                            if value is not None:
                                self._write_value_to_cell(
                                    target_ws,
                                    cell_address,
                                    value,
                                    add_comment=f"原公式转换失败: {errors[0].error_message}"
                                )
                        elif options.error_handling_mode == "skip":
                            # 跳过该单元格
                            continue
                        else:  # fail
                            # 停止导出
                            raise Exception(f"公式转换失败: {errors[0].error_message}")
                    else:
                        # 成功转换,写入公式
                        if excel_formula:
                            self._write_formula_to_cell(
                                target_ws,
                                cell_address,
                                excel_formula,
                                fallback_value=mapping_formula.calculation_result,
                                add_comment=options.add_formula_comments
                            )
                            converted_formulas += 1
                            sheet_converted_count += 1

                print(f"  该表成功转换 {sheet_converted_count} 个公式")

            # 5. 设置Excel为自动计算模式
            output_wb.calculation.calcMode = 'auto'

            # 6. 保存文件
            print(f"\n正在保存文件到: {output_path}")
            output_wb.save(output_path)
            output_wb.close()

            # 7. 生成失败报告文件 (如果有失败的转换)
            if failed_conversions:
                self._generate_all_sheets_failure_report(
                    output_path,
                    workbook_manager,
                    flash_report_sheets,
                    total_formulas,
                    converted_formulas,
                    failed_conversions
                )

            # 8. 计算耗时
            end_time = datetime.now()
            execution_time = (end_time - start_time).total_seconds()

            # 9. 创建结果报告
            result = ExportResult(
                success=True,
                output_file_path=output_path,
                total_formulas=total_formulas,
                converted_formulas=converted_formulas,
                failed_conversions=failed_conversions,
                execution_time=execution_time,
                validation_report=None
            )

            self.last_export_result = result
            print(f"\n总计: 成功转换 {converted_formulas}/{total_formulas} 个公式")
            return result

        except Exception as e:
            end_time = datetime.now()
            execution_time = (end_time - start_time).total_seconds()

            result = ExportResult(
                success=False,
                output_file_path=output_path,
                total_formulas=total_formulas,
                converted_formulas=converted_formulas,
                failed_conversions=failed_conversions,
                execution_time=execution_time
            )

            print(f"导出失败: {str(e)}")
            import traceback
            traceback.print_exc()

            self.last_export_result = result
            return result

    def _generate_all_sheets_failure_report(
        self,
        output_path: str,
        workbook_manager: WorkbookManager,
        sheet_names: list,
        total_formulas: int,
        converted_formulas: int,
        failed_conversions: List[ConversionError]
    ) -> None:
        """
        生成所有表的失败报告txt文件

        Args:
            output_path: Excel输出文件路径
            workbook_manager: 工作簿管理器
            sheet_names: 导出的工作表名称列表
            total_formulas: 总公式数
            converted_formulas: 成功转换的公式数
            failed_conversions: 失败的转换列表
        """
        # 生成报告文件名（与Excel文件同名，扩展名为txt）
        import os
        base_path = os.path.splitext(output_path)[0]
        report_path = f"{base_path}_导出失败报告.txt"

        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                # 写入标题
                f.write("="*80 + "\n")
                f.write("Excel公式导出失败报告（所有待写入表）\n")
                f.write("="*80 + "\n\n")

                # 写入导出信息
                f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"源文件: {workbook_manager.file_path}\n")
                f.write(f"导出的待写入表: {', '.join(sheet_names)}\n")
                f.write(f"导出文件: {output_path}\n\n")

                # 写入统计信息
                f.write("-"*80 + "\n")
                f.write("统计信息\n")
                f.write("-"*80 + "\n")
                success_rate = (converted_formulas / total_formulas * 100) if total_formulas > 0 else 0
                f.write(f"总公式数: {total_formulas}\n")
                f.write(f"成功转换: {converted_formulas} ({success_rate:.1f}%)\n")
                f.write(f"失败转换: {len(failed_conversions)} ({100-success_rate:.1f}%)\n\n")

                # 按工作表分组统计
                f.write("按工作表分组统计：\n")
                sheet_stats = {}
                for error in failed_conversions:
                    if error.target_item:
                        sheet = error.target_item.sheet_name
                        sheet_stats[sheet] = sheet_stats.get(sheet, 0) + 1

                for sheet_name in sheet_names:
                    error_count = sheet_stats.get(sheet_name, 0)
                    f.write(f"  • {sheet_name}: {error_count} 个失败\n")
                f.write("\n")

                # 写入失败原因说明
                f.write("-"*80 + "\n")
                f.write("失败原因说明\n")
                f.write("-"*80 + "\n")
                f.write("以下单元格的公式转换失败，已使用计算值代替：\n\n")

                # 按错误类型分组统计
                error_types = {}
                for error in failed_conversions:
                    error_type = error.error_type
                    if error_type not in error_types:
                        error_types[error_type] = []
                    error_types[error_type].append(error)

                # 写入错误类型统计
                f.write("错误类型统计：\n")
                for error_type, errors in error_types.items():
                    f.write(f"  - {error_type}: {len(errors)} 个\n")
                f.write("\n")

                # 写入详细的失败列表
                f.write("="*80 + "\n")
                f.write("详细失败列表\n")
                f.write("="*80 + "\n\n")

                for i, error in enumerate(failed_conversions, 1):
                    # 获取目标项信息
                    target_item = error.target_item
                    if target_item:
                        target_name = target_item.name
                        target_sheet = target_item.sheet_name
                        target_cell = target_item.target_cell_address or "未知"

                        # 尝试获取列信息
                        column_info = ""
                        if hasattr(target_item, 'columns') and target_item.columns:
                            # 多列模式，尝试找到对应的列
                            for col_key, col_entry in target_item.columns.items():
                                if hasattr(col_entry, 'cell_address'):
                                    column_info = f" ({col_key})"
                                    target_cell = col_entry.cell_address
                                    break
                    else:
                        target_name = "未知"
                        target_sheet = "未知"
                        target_cell = "未知"
                        column_info = ""

                    # 写入失败信息
                    f.write(f"{i}. 工作表: {target_sheet}\n")
                    f.write(f"   目标项: {target_name}{column_info}\n")
                    f.write(f"   单元格: {target_cell}\n")
                    f.write(f"   错误类型: {error.error_type}\n")
                    f.write(f"   错误信息: {error.error_message}\n")
                    f.write(f"   原始公式: {error.internal_formula}\n")
                    f.write(f"   处理方式: {error.fallback_action}\n")

                    # 如果使用了计算值，显示计算值
                    if error.fallback_action == "used_value" and target_item:
                        # 尝试从workbook_manager获取计算结果
                        mapping_formula = None
                        for target_id, mapping_dict in workbook_manager.mapping_formulas.items():
                            if workbook_manager.target_items.get(target_id) == target_item:
                                if isinstance(mapping_dict, dict):
                                    # 多列模式，找到第一个匹配的
                                    for col_key, formula in mapping_dict.items():
                                        if formula.formula == error.internal_formula:
                                            mapping_formula = formula
                                            break
                                else:
                                    mapping_formula = mapping_dict
                                break

                        if mapping_formula and mapping_formula.calculation_result is not None:
                            f.write(f"   使用值: {mapping_formula.calculation_result}\n")

                    f.write("\n")

                # 写入结尾说明
                f.write("="*80 + "\n")
                f.write("说明\n")
                f.write("="*80 + "\n")
                f.write("1. 公式转换失败的原因通常包括：\n")
                f.write("   - cell_not_found: 未找到引用的源数据项或单元格\n")
                f.write("   - syntax_error: 公式语法错误\n")
                f.write("   - reference_error: 引用解析错误\n")
                f.write("   - security_error: 安全验证失败（可能包含危险字符）\n")
                f.write("   - cell_bounds_error: 单元格地址超出Excel限制\n\n")
                f.write("2. 失败的单元格已使用计算值填充，确保数据完整性。\n\n")
                f.write("3. 建议检查源数据工作表，确保所有引用的项目都存在。\n\n")
                f.write("4. 如需重新生成公式，请修正源数据后重新导出。\n")

            print(f"失败报告已生成: {report_path}")

        except Exception as e:
            print(f"生成失败报告时出错: {str(e)}")
            import traceback
            traceback.print_exc()
