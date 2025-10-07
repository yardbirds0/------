#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
表格模式分析器 - 智能识别财务表格类型和列头结构
"""

import re
import unicodedata
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
from enum import Enum
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

class TableType(Enum):
    """表格类型枚举"""
    BALANCE_SHEET = "资产负债表"
    INCOME_STATEMENT = "利润表"
    CASH_FLOW = "现金流量表"
    TRIAL_BALANCE = "科目余额表"
    UNKNOWN = "未知表格"

@dataclass
class ColumnInfo:
    """列信息"""
    column_index: int  # 列索引
    column_letter: str  # 列字母
    primary_header: str  # 一级列头
    secondary_header: str = ""  # 二级列头
    data_type: str = "unknown"  # 数据类型：debit, credit, amount, text
    is_numeric: bool = False  # 是否为数值列
    normalized_key: str = ""  # 规范化键名（用于数据字典）
    display_name: str = ""  # 展示名称
    header_text: str = ""  # 完整列头文本
    is_data_column: bool = False  # 是否判定为数据列
    is_placeholder: bool = False  # 是否为占位列头
    primary_col_span: int = 1  # 一级列头横向跨度
    primary_row_span: int = 1  # 一级列头纵向跨度
    primary_is_group_start: bool = True  # 是否为一级列头合并块起始列
    primary_start_column: int = 0  # 一级列头合并块起始列索引
    primary_merge_key: Optional[Tuple[int, int, int, int]] = None  # 一级列头合并标识
    secondary_col_span: int = 1  # 二级列头横向跨度
    secondary_row_span: int = 1  # 二级列头纵向跨度
    secondary_is_group_start: bool = True  # 是否为二级列头合并块起始列
    secondary_start_column: int = 0  # 二级列头合并块起始列索引
    secondary_merge_key: Optional[Tuple[int, int, int, int]] = None  # 二级列头合并标识

@dataclass
class TableSchema:
    """表格模式"""
    table_type: TableType
    header_rows: int  # 列头行数
    header_start_row: int  # 列头起始行
    data_start_row: int  # 数据开始行
    name_columns: List[int]  # 项目名称列
    code_columns: List[int]  # 编码列
    data_columns: List[ColumnInfo]  # 数据列信息
    has_hierarchy: bool = False  # 是否有层级结构

class TableSchemaAnalyzer:
    """表格模式分析器"""

    def __init__(self):
        # 财务表格关键词模式
        self.table_type_patterns = {
            TableType.BALANCE_SHEET: [
                r'资产负债表', r'资产.*负债', r'balance.*sheet',
                r'资产总计', r'负债总计', r'所有者权益'
            ],
            TableType.INCOME_STATEMENT: [
                r'利润表', r'损益表', r'income.*statement', r'profit.*loss',
                r'营业收入', r'营业利润', r'净利润'
            ],
            TableType.CASH_FLOW: [
                r'现金流量表', r'cash.*flow',
                r'经营活动', r'投资活动', r'筹资活动'
            ],
            TableType.TRIAL_BALANCE: [
                r'科目余额表', r'trial.*balance', r'余额表',
                r'科目代码', r'科目名称', r'期初余额', r'期末余额'
            ]
        }

        # 列头关键词模式
        self.column_patterns = {
            'debit': [r'借方', r'debit', r'dr'],
            'credit': [r'贷方', r'credit', r'cr'],
            'beginning': [r'期初', r'年初', r'beginning', r'opening'],
            'ending': [r'期末', r'年末', r'ending', r'closing'],
            'current': [r'本期', r'本年', r'current', r'this.*year'],
            'previous': [r'上期', r'上年', r'去年', r'previous', r'last.*year'],
            'amount': [r'金额', r'数额', r'amount', r'value'],
            'balance': [r'余额', r'balance']
        }

        # 列头识别关键词
        self.header_keywords = [
            '项目', '科目', '指标', '名称', '内容', '行次', '行号', '序号',
            '金 额', '金额', '本期', '本月', '本年', '累计', '上期', '上年', '期末',
            '期初', '余额', '借方', '贷方', '备注', '数量', '单价', '本季', '科目代码'
        ]

        self.header_meta_keywords = [
            '编制单位', '填报单位', '金额单位', '单位：', '单位:', '单位（', '单位(',
            '制表', '审核', '复核', '主管', '负责人', '联系电话', '联系人', '报出',
            '填报', '年月日', '年 月 日', '日期', '财务负责人', '公司', '集团', '部门'
        ]

        self.header_description_patterns = [
            r'编制单位', r'填报单位', r'单位[:：]', r'金额单位', r'日期', r'年\s*月',
            r'表$', r'^备注', r'^说明', r'公司', r'集团', r'部门', r'负责人', r'审核',
            r'复核', r'主管', r'报出', r'联系电话', r'联系人', r'制表', r'填报',
            r'财务负责人', r'^\s*单位\s*$', r'^\s*联系方式\s*$'
        ]

    def _clean_header_value(self, value: Any) -> str:
        """清洗列头文本，移除多余空格与全角字符"""
        if value is None:
            return ""

        if isinstance(value, str):
            text = value
        else:
            text = str(value)

        text = unicodedata.normalize("NFKC", text)
        text = text.replace('\u3000', ' ').replace('\xa0', ' ')
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def _is_placeholder_text(self, text: str) -> bool:
        """判定列头文本是否为空或占位符"""
        if not text:
            return True

        stripped = text.strip()
        if not stripped:
            return True

        # 去除常见占位符字符后若为空则视为占位
        normalized = re.sub(r'[\-_=~/\\·\.\s—–]+', '', stripped)
        if not normalized:
            return True

        # 单一重复字符
        if len(set(stripped)) == 1 and stripped[0] in '-_=~•·.。*#/\\|':
            return True

        return False

    def _is_noise_cell_text(self, text: str) -> bool:
        """判定单元格文本是否缺乏有效信息"""
        if not text:
            return True

        stripped = text.strip()
        if not stripped:
            return True

        if re.fullmatch(r'[\-_=~/\\·\.·\s—–\|]+', stripped):
            return True

        # 缺少字母或汉字且长度极短
        if len(stripped) <= 2 and not re.search(r'[A-Za-z0-9\u4e00-\u9fff]', stripped):
            return True

        return False

    def _contains_meta_keyword(self, text: str) -> bool:
        """检测文本是否包含说明性关键词"""
        if not text:
            return False

        lower = text.lower()
        if any(keyword.lower() in lower for keyword in self.header_meta_keywords):
            return True

        return any(re.search(pattern, text) for pattern in self.header_description_patterns)

    def _looks_like_data_row(self, sheet: Worksheet, row: int, header_start_row: int) -> bool:
        """
        判断指定行是否像数据行而不是列头

        检查指标：
        1. 第一列是否包含数字开头的项目名（如"1."、"一、"）
        2. 单元格内容是否过长（超过20字符）
        3. 是否包含明显的数据行特征
        """
        max_columns = min(10, sheet.max_column or 0)

        # 检查前几列的内容
        for col in range(1, max_columns + 1):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            if value is None:
                continue

            text = self._clean_header_value(value)
            if not text:
                continue

            # 检查1：是否为数字开头的项目名称（如"1.营业总收入"、"一、资产类"）
            if re.match(r'^[\d一二三四五六七八九十]+[.、）\)]', text):
                return True

            # 检查2：内容是否过长（列头通常较短）
            if len(text) > 25:
                return True

            # 检查3：包含"总计"、"合计"等汇总词（通常在数据行）
            if re.search(r'总计|合计|小计|总额|合计数', text):
                return True

        return False

    def analyze_table_schema(self, sheet: Worksheet) -> TableSchema:
        """分析表格模式"""
        # 识别表格类型
        table_type = self._identify_table_type(sheet)

        # 分析列头结构
        header_info = self._analyze_headers(sheet)

        # 识别数据列
        data_columns = self._identify_data_columns(sheet, header_info)

        # 识别名称和编码列
        name_columns, code_columns = self._identify_name_code_columns(sheet, table_type, header_info)

        # 确定数据开始行
        data_start_row = self._find_data_start_row(sheet, header_info)

        return TableSchema(
            table_type=table_type,
            header_rows=header_info['header_rows'],
            header_start_row=header_info['header_start_row'],
            data_start_row=data_start_row,
            name_columns=name_columns,
            code_columns=code_columns,
            data_columns=data_columns,
            has_hierarchy=(table_type == TableType.TRIAL_BALANCE or len(code_columns) > 0)
        )

    def _identify_table_type(self, sheet: Worksheet) -> TableType:
        """识别表格类型"""
        # 检查工作表名称
        sheet_name = sheet.title.lower()

        # 收集前10行的所有文本
        all_text = []
        for row in range(1, min(11, sheet.max_row + 1)):
            for col in range(1, min(11, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    all_text.append(str(cell.value).lower())

        combined_text = ' '.join([sheet_name] + all_text)

        # 匹配表格类型
        for table_type, patterns in self.table_type_patterns.items():
            for pattern in patterns:
                if re.search(pattern, combined_text, re.IGNORECASE):
                    return table_type

        return TableType.UNKNOWN

    def _analyze_headers(self, sheet: Worksheet) -> Dict[str, Any]:
        """分析列头结构，自动跳过前置说明行"""
        max_row = sheet.max_row or 0
        if max_row == 0:
            return {
                'header_rows': 1,
                'header_start_row': 1,
                'has_merged_cells': False
            }

        max_scan_rows = min(12, max_row)
        row_metrics: List[Dict[str, Any]] = []

        for row in range(1, max_scan_rows + 1):
            metrics = self._evaluate_header_row(sheet, row)
            row_metrics.append(metrics)

        header_start_row = 1
        header_rows = 1

        header_candidates = [m for m in row_metrics if m.get('is_header')]
        selected_metrics: Optional[Dict[str, Any]] = None

        if header_candidates:
            selected_metrics = min(
                header_candidates,
                key=lambda m: m.get('row', 1)
            )
        else:
            fallback_candidates = [m for m in row_metrics if not m.get('is_description')]
            if fallback_candidates:
                selected_metrics = max(
                    fallback_candidates,
                    key=lambda m: (m.get('header_score', 0), -m.get('row', 0))
                )

        if not selected_metrics and row_metrics:
            selected_metrics = row_metrics[0]

        if selected_metrics:
            header_start_row = selected_metrics.get('row', 1)
            header_rows = 1

            expected_row = header_start_row + 1
            for follow in row_metrics:
                if follow['row'] < expected_row:
                    continue
                if follow.get('is_description'):
                    continue

                if follow['row'] != expected_row:
                    break

                # 🔧 增强鲁棒性：如果紧跟列头的行只有很少的非空单元格（<=2个），
                # 很可能是数据行的开始，不应该被计入列头
                if follow.get('non_empty', 0) <= 2:
                    break

                # 🔧 新增：检查是否为数据行开始
                if self._looks_like_data_row(sheet, expected_row, header_start_row):
                    break

                if follow.get('is_header_continuation') or follow.get('is_header'):
                    header_rows += 1
                    expected_row += 1
                else:
                    break

        return {
            'header_rows': header_rows,
            'header_start_row': header_start_row,
            'has_merged_cells': self._check_merged_cells(sheet, header_start_row, header_rows)
        }

    def _evaluate_header_row(self, sheet: Worksheet, row: int) -> Dict[str, Any]:
        max_columns = min(20, sheet.max_column or 0)
        non_empty = 0
        numeric_cells = 0
        keyword_hits = 0
        description_hits = 0
        meta_hits = 0
        noise_cells = 0
        short_text_count = 0
        unique_texts: set = set()

        for col in range(1, max_columns + 1):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            if value is None:
                continue

            text = self._clean_header_value(value)
            if not text:
                continue

            non_empty += 1
            unique_texts.add(text)

            lower = text.lower()
            if any(keyword.lower() in lower for keyword in self.header_keywords):
                keyword_hits += 1

            if any(re.search(pattern, text) for pattern in self.header_description_patterns):
                description_hits += 1

            if self._contains_meta_keyword(text):
                meta_hits += 1

            if self._is_noise_cell_text(text):
                noise_cells += 1

            if isinstance(value, (int, float)) or self._is_numeric_string(value):
                numeric_cells += 1
            elif self._is_numeric_string(text):
                numeric_cells += 1

            if len(text) <= 8:
                short_text_count += 1

        numeric_ratio = (numeric_cells / non_empty) if non_empty else 0.0
        short_ratio = (short_text_count / non_empty) if non_empty else 0.0
        has_keywords = keyword_hits > 0
        meta_ratio = (meta_hits / non_empty) if non_empty else 0.0
        noise_ratio = (noise_cells / non_empty) if non_empty else 0.0
        is_description = (
            non_empty == 0
            or (meta_hits >= 1 and meta_ratio >= 0.5)
            or (meta_hits >= 2 and keyword_hits == 0)
            or (description_hits > 0 and keyword_hits == 0 and non_empty <= 4)
            or noise_ratio >= 0.6
        )
        is_data_row = non_empty > 0 and numeric_ratio >= 0.5 and not has_keywords

        is_header = (
            non_empty >= 2
            and not is_description
            and not is_data_row
            and (
                has_keywords
                or (short_ratio >= 0.6 and len(unique_texts) >= 2)
            )
        )

        # 紧跟在表头之后的描述性行（如“单位：元”）不算在列头行内
        is_header_continuation = (
            not is_header
            and not is_description
            and not is_data_row
            and non_empty >= 1
            and (short_ratio >= 0.5 or keyword_hits >= 1)
        )

        return {
            'row': row,
            'non_empty': non_empty,
            'is_header': is_header,
            'is_header_continuation': is_header_continuation,
            'numeric_ratio': numeric_ratio,
            'keyword_hits': keyword_hits,
            'description_hits': description_hits,
            'meta_hits': meta_hits,
            'noise_ratio': noise_ratio,
            'is_description': is_description,
            'header_score': keyword_hits * 2 + non_empty - meta_hits * 1.5 - (5 if is_description else 0)
        }

    def _build_header_merge_lookup(
        self,
        sheet: Worksheet,
        header_start: int,
        header_rows: int
    ) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
        """构建列头区域合并单元格查找表"""
        lookup: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
        header_end = header_start + header_rows - 1

        for merged in sheet.merged_cells.ranges:
            if merged.max_row < header_start or merged.min_row > header_end:
                continue

            min_row, max_row = merged.min_row, merged.max_row
            min_col, max_col = merged.min_col, merged.max_col

            for row in range(min_row, max_row + 1):
                if row < header_start or row > header_end:
                    continue
                for col in range(min_col, max_col + 1):
                    lookup[(row, col)] = (min_row, max_row, min_col, max_col)

        return lookup

    def _get_header_cell_info(
        self,
        sheet: Worksheet,
        row: int,
        col: int,
        header_start: int,
        header_rows: int,
        merged_lookup: Dict[Tuple[int, int], Tuple[int, int, int, int]]
    ) -> Dict[str, Any]:
        """获取列头单元格的文本及合并信息"""
        merge = merged_lookup.get((row, col))
        header_end = header_start + header_rows - 1

        if merge:
            min_row, max_row, min_col, max_col = merge
            base_cell = sheet.cell(row=min_row, column=min_col)
            text = self._clean_header_value(base_cell.value)
            effective_min_row = max(min_row, header_start)
            effective_max_row = min(max_row, header_end)
            row_span = effective_max_row - effective_min_row + 1
            col_span = max_col - min_col + 1
            is_group_start = (col == min_col and row == effective_min_row)
            start_col = min_col
            start_row = effective_min_row
        else:
            cell = sheet.cell(row=row, column=col)
            text = self._clean_header_value(cell.value)
            row_span = 1
            col_span = 1
            is_group_start = True
            start_col = col
            start_row = row
            min_row = row
            max_row = row
            min_col = col
            max_col = col

        if self._is_placeholder_text(text):
            text = ""

        return {
            'text': text,
            'row_span': row_span,
            'col_span': col_span,
            'is_group_start': is_group_start,
            'start_col': start_col,
            'start_row': start_row,
            'merge_key': (min_row, max_row, min_col, max_col) if merge else None
        }

    def _identify_data_columns(self, sheet: Worksheet, header_info: Dict) -> List[ColumnInfo]:
        """识别数据列"""
        data_columns: List[ColumnInfo] = []
        header_rows = header_info.get('header_rows', 1)
        header_start = header_info.get('header_start_row', 1)
        merged_lookup = self._build_header_merge_lookup(sheet, header_start, header_rows) if header_rows >= 1 else {}

        max_columns = sheet.max_column or 0
        for col in range(1, min(16, max_columns + 1)):
            col_letter = openpyxl.utils.get_column_letter(col)

            primary_info = {
                'text': '',
                'row_span': 1,
                'col_span': 1,
                'is_group_start': True,
                'start_col': col,
                'start_row': header_start,
                'merge_key': None
            }
            secondary_info = {
                'text': '',
                'row_span': 0,
                'col_span': 1,
                'is_group_start': True,
                'start_col': col,
                'start_row': header_start + 1,
                'merge_key': None
            }

            if header_rows >= 1:
                primary_info = self._get_header_cell_info(
                    sheet,
                    header_start,
                    col,
                    header_start,
                    header_rows,
                    merged_lookup
                )

            primary_header = primary_info.get('text', '')

            # 先获取二级列头信息，以便完整判断是否需要保留该列
            if header_rows >= 2:
                if primary_info.get('row_span', 1) < header_rows:
                    second_row = header_start + primary_info.get('row_span', 1)
                    secondary_info = self._get_header_cell_info(
                        sheet,
                        second_row,
                        col,
                        header_start,
                        header_rows,
                        merged_lookup
                    )
                else:
                    secondary_info['row_span'] = max(0, header_rows - primary_info.get('row_span', 1))
                    secondary_info['start_row'] = primary_info.get('start_row', header_start) + primary_info.get('row_span', 1)
                secondary_header = secondary_info.get('text', '')
            else:
                secondary_header = ""

            if self._is_placeholder_text(primary_header):
                primary_header = ""
            if self._is_placeholder_text(secondary_header):
                secondary_header = ""

            # 🔧 修复：跳过横跨多列的合并单元格的非起始列，避免重复识别列头
            # 但如果该列有明确的二级列头文本（如"借方"、"贷方"），则保留，无论是否有数据
            if not primary_info.get('is_group_start', True):
                # 如果有明确的二级列头文本，保留该列
                if secondary_header:
                    pass  # 保留有二级列头的列
                else:
                    # 检查是否有独立的数值数据
                    data_start_guess = header_start + max(header_rows, 1)
                    is_numeric_independent = self._check_column_numeric(sheet, col, data_start_guess)
                    if not is_numeric_independent:
                        # 如果不是合并单元格起始列、没有二级列头、且没有独立数值数据，跳过
                        continue

            is_placeholder = not primary_header and not secondary_header

            data_start_guess = header_start + max(header_rows, 1)
            is_numeric = self._check_column_numeric(sheet, col, data_start_guess)

            if is_numeric or primary_header or secondary_header:
                data_type = self._identify_data_type(primary_header, secondary_header)

                normalized_key, display_name = self._build_column_identifiers(
                    primary_header,
                    secondary_header,
                    col_letter
                )

                column_info = ColumnInfo(
                    column_index=col,
                    column_letter=col_letter,
                    primary_header=primary_header,
                    secondary_header=secondary_header,
                    data_type=data_type,
                    is_numeric=is_numeric,
                    normalized_key=normalized_key,
                    display_name=display_name,
                    header_text=self._compose_header_text(primary_header, secondary_header, col_letter),
                    is_data_column=self._determine_is_data_column(is_numeric, data_type, primary_header, secondary_header),
                    is_placeholder=is_placeholder,
                    primary_col_span=primary_info.get('col_span', 1),
                    primary_row_span=primary_info.get('row_span', 1),
                    primary_is_group_start=primary_info.get('is_group_start', True),
                    primary_start_column=primary_info.get('start_col', col),
                    primary_merge_key=primary_info.get('merge_key'),
                    secondary_col_span=secondary_info.get('col_span', 1),
                    secondary_row_span=secondary_info.get('row_span', 0),
                    secondary_is_group_start=secondary_info.get('is_group_start', True),
                    secondary_start_column=secondary_info.get('start_col', col),
                    secondary_merge_key=secondary_info.get('merge_key')
                )
                data_columns.append(column_info)

        return data_columns

    def _build_column_identifiers(self, primary: str, secondary: str, column_letter: str) -> Tuple[str, str]:
        """生成列的规范化键名与展示名称"""
        candidates = []
        primary_clean = primary.strip() if primary else ""
        secondary_clean = secondary.strip() if secondary else ""

        if primary_clean and secondary_clean:
            candidates.append(f"{primary_clean}_{secondary_clean}")
        if primary_clean:
            candidates.append(primary_clean)
        if secondary_clean:
            candidates.append(secondary_clean)

        for candidate in candidates:
            normalized = self._sanitize_key(candidate)
            if normalized:
                return normalized, candidate

        # 回退使用列字母
        fallback_key = f"col_{column_letter}"
        return fallback_key, f"列{column_letter}"

    def _sanitize_key(self, text: str) -> str:
        """将文本转换为规范化的键名"""
        if not text:
            return ""

        cleaned = re.sub(r'\s+', '_', text.strip())
        cleaned = re.sub(r'[^0-9A-Za-z_\u4e00-\u9fff]+', '_', cleaned)
        cleaned = re.sub(r'_+', '_', cleaned)
        cleaned = cleaned.strip('_')
        return cleaned.lower()

    def _compose_header_text(self, primary: str, secondary: str, column_letter: str) -> str:
        """组合完整列头文本"""
        parts = [part.strip() for part in [primary, secondary] if part and part.strip()]
        if parts:
            return " / ".join(parts)
        return f"列{column_letter}"

    def _determine_is_data_column(self, is_numeric: bool, data_type: str, primary: str, secondary: str) -> bool:
        """根据内容判断是否为数据列"""
        if is_numeric:
            return True

        combined = (primary or "") + " " + (secondary or "")
        combined_lower = combined.lower()

        keyword_patterns = [
            r'金额', r'余额', r'数值', r'合计', r'本期', r'上期', r'年初', r'期末',
            r'借方', r'贷方', r'累计', r'发生'
        ]

        if any(re.search(pattern, combined_lower) for pattern in keyword_patterns):
            return True

        return data_type in {"debit", "credit", "amount", "balance", "current", "previous", "ending", "beginning"}

    def _identify_name_code_columns(
        self,
        sheet: Worksheet,
        table_type: TableType,
        header_info: Dict[str, Any]
    ) -> Tuple[List[int], List[int]]:
        """识别名称列和编码列"""
        name_columns = []
        code_columns = []

        header_start = header_info.get('header_start_row', 1)
        header_rows = header_info.get('header_rows', 1)
        data_start_row = header_start + header_rows

        # 检查前5列
        for col in range(1, min(6, sheet.max_column + 1)):
            # 获取列头
            header_text = ""
            for row in range(header_start, min(header_start + header_rows, sheet.max_row + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    header_text += str(cell.value).lower()

            # 检查几行数据来判断列的性质
            sample_values = []
            for row in range(data_start_row, min(data_start_row + 6, sheet.max_row + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    sample_values.append(str(cell.value).strip())

            if not sample_values:
                continue

            # 判断是否为编码列
            if self._is_code_column(header_text, sample_values):
                code_columns.append(col)
            # 判断是否为名称列
            elif self._is_name_column(header_text, sample_values, table_type):
                name_columns.append(col)

        return name_columns, code_columns

    def _identify_data_type(self, primary: str, secondary: str) -> str:
        """识别数据类型"""
        combined = (primary + " " + secondary).lower()

        # 检查各种类型
        for data_type, patterns in self.column_patterns.items():
            for pattern in patterns:
                if re.search(pattern, combined, re.IGNORECASE):
                    return data_type

        return "amount"

    def _is_code_column(self, header_text: str, sample_values: List[str]) -> bool:
        """判断是否为编码列"""
        # 检查列头
        if re.search(r'代码|code|编号|number', header_text, re.IGNORECASE):
            return True

        # 检查数据特征
        if not sample_values:
            return False

        numeric_count = 0
        for value in sample_values[:5]:  # 检查前5个值
            # 科目代码通常是纯数字或以数字开头
            if re.match(r'^\d{4,8}', value):
                numeric_count += 1

        return numeric_count >= len(sample_values) * 0.7

    def _is_name_column(self, header_text: str, sample_values: List[str], table_type: TableType) -> bool:
        """判断是否为名称列"""
        # 检查列头
        name_patterns = [r'名称|项目|科目|name|item|account']
        for pattern in name_patterns:
            if re.search(pattern, header_text, re.IGNORECASE):
                return True

        # 检查数据特征
        if not sample_values:
            return False

        chinese_count = 0
        for value in sample_values[:5]:
            # 包含中文且不是纯数字
            if re.search(r'[\u4e00-\u9fff]', value) and not re.match(r'^\d+\.?\d*$', value):
                chinese_count += 1

        return chinese_count >= len(sample_values) * 0.7

    def _check_column_numeric(self, sheet: Worksheet, col: int, start_row: int) -> bool:
        """检查列是否为数值列"""
        numeric_count = 0
        total_count = 0

        for row in range(start_row, min(start_row + 10, sheet.max_row + 1)):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                total_count += 1
                if isinstance(cell.value, (int, float)) or self._is_numeric_string(cell.value):
                    numeric_count += 1

        return total_count > 0 and (numeric_count / total_count) >= 0.5

    def _is_numeric_string(self, value) -> bool:
        """判断是否为数值字符串"""
        if not isinstance(value, str):
            return False

        try:
            # 去除常见的格式符号
            cleaned = value.replace(',', '').replace(' ', '').replace('(', '-').replace(')', '')
            float(cleaned)
            return True
        except ValueError:
            return False

    def _check_merged_cells(self, sheet: Worksheet, header_start: int, header_rows: int) -> bool:
        """检查列头区域是否存在合并单元格"""
        header_end = header_start + header_rows - 1
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.max_row < header_start or merged_range.min_row > header_end:
                continue
            return True
        return False

    def _find_data_start_row(self, sheet: Worksheet, header_info: Dict[str, Any]) -> int:
        """寻找数据开始行"""
        header_start = header_info.get('header_start_row', 1)
        header_rows = header_info.get('header_rows', 1)
        candidate_start = header_start + header_rows
        max_row = sheet.max_row or candidate_start

        for row in range(candidate_start, min(candidate_start + 8, max_row + 1)):
            row_metrics = self._evaluate_header_row(sheet, row)
            if row_metrics.get('non_empty', 0) == 0:
                continue
            if row_metrics.get('is_description'):
                continue

            has_name = False
            has_number = False

            for col in range(1, min(8, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value is None:
                    continue
                value_str = str(value).strip()
                if not value_str:
                    continue

                if re.search(r'[\u4e00-\u9fff]', value_str) and not self._is_numeric_string(value_str):
                    has_name = True
                if isinstance(value, (int, float)) or self._is_numeric_string(value):
                    has_number = True

            if has_name or has_number:
                return row

        return candidate_start

if __name__ == "__main__":
    # 测试代码
    print("表格模式分析器模块")
