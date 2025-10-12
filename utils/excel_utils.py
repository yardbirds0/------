#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel处理工具函数 - PySide6版本
支持新的公式格式: [工作表名称]A1
"""

import openpyxl
from typing import List, Dict, Tuple, Any, Optional, Union
from datetime import datetime
import re
import shutil
import os


# ==================== 公式引用格式常量 ====================

# 新格式示例: [利润表]A12
SIMPLE_REFERENCE_PATTERN = re.compile(
    r"\[(?P<sheet>[^\]]+)\](?P<cell>\$?[A-Z]+\$?\d+)")


def safe_get_cell_value(sheet, row: int, column: Union[int, str]) -> Any:
    """
    安全获取单元格值

    Args:
        sheet: openpyxl工作表对象
        row: 行号（1开始）
        column: 列号（整数）或列字母（字符串）

    Returns:
        Any: 单元格值，如果出错则返回None
    """
    try:
        if isinstance(column, str):
            cell = sheet[f"{column}{row}"]
        else:
            cell = sheet.cell(row=row, column=column)
        return cell.value
    except:
        return None


def get_merged_cell_value(sheet, row: int, column: int) -> Any:
    """
    获取合并单元格的值

    Args:
        sheet: openpyxl工作表对象
        row: 行号
        column: 列号

    Returns:
        Any: 单元格值
    """
    cell = sheet.cell(row=row, column=column)

    # 检查是否为合并单元格
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 获取合并区域的左上角单元格值
            top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left_cell.value

    return cell.value


def convert_column_letter_to_number(column_letter: str) -> int:
    """
    将列字母转换为列号

    Args:
        column_letter: 列字母，如'A', 'B', 'AA'

    Returns:
        int: 列号（1开始）
    """
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result


def convert_column_number_to_letter(column_number: int) -> str:
    """
    将列号转换为列字母

    Args:
        column_number: 列号（1开始）

    Returns:
        str: 列字母
    """
    result = ""
    while column_number > 0:
        column_number -= 1
        result = chr(column_number % 26 + ord('A')) + result
        column_number //= 26
    return result


def parse_cell_address(cell_address: str) -> Tuple[int, str]:
    """
    解析单元格地址

    Args:
        cell_address: 单元格地址，如'A1', 'BC25'

    Returns:
        Tuple[int, str]: (行号, 列字母)
    """
    match = re.match(r'^([A-Z]+)(\d+)$', cell_address.upper())
    if match:
        column_letter = match.group(1)
        row_number = int(match.group(2))
        return row_number, column_letter
    else:
        raise ValueError(f"无效的单元格地址: {cell_address}")


def format_cell_address(row: int, column: Union[int, str]) -> str:
    """格式化单元格地址"""
    if isinstance(column, int):
        column_letter = convert_column_number_to_letter(column)
    else:
        column_letter = column.upper()
    return f"{column_letter}{row}"


def find_header_row(sheet, keywords: List[str], max_search_rows: int = 20) -> Optional[int]:
    """在指定行数内查找包含指定关键词的表头行"""
    for row_num in range(1, min(max_search_rows + 1, sheet.max_row + 1)):
        row_text = ""
        for col_num in range(1, min(sheet.max_column + 1, 10)):
            cell_value = safe_get_cell_value(sheet, row_num, col_num)
            if cell_value:
                row_text += str(cell_value).lower()

        if any(keyword.lower() in row_text for keyword in keywords):
            return row_num

    return None


def extract_number_from_text(text: str) -> Optional[Union[int, float]]:
    """从文本中提取数字"""
    if not isinstance(text, str):
        return None

    cleaned_text = re.sub(r'[^\d\.\-]', '', text)
    if not cleaned_text:
        return None

    try:
        return float(cleaned_text) if '.' in cleaned_text else int(cleaned_text)
    except ValueError:
        return None


def is_numeric_cell(cell_value: Any) -> bool:
    """判断单元格值是否为数值"""
    if cell_value is None:
        return False
    if isinstance(cell_value, (int, float)):
        return True
    if isinstance(cell_value, str):
        return extract_number_from_text(cell_value) is not None
    return False


def get_column_values(
    sheet,
    column: Union[int, str],
    start_row: int = 1,
    end_row: Optional[int] = None
) -> List[Any]:
    """获取指定范围内某列的所有值"""
    if end_row is None:
        end_row = sheet.max_row

    values: List[Any] = []
    for row_num in range(start_row, end_row + 1):
        values.append(safe_get_cell_value(sheet, row_num, column))
    return values


def find_empty_cells(
    sheet,
    column: Union[int, str],
    start_row: int = 1,
    end_row: Optional[int] = None
) -> List[int]:
    """查找某列中的空单元格行号"""
    if end_row is None:
        end_row = sheet.max_row

    empty_rows: List[int] = []
    for row_num in range(start_row, end_row + 1):
        value = safe_get_cell_value(sheet, row_num, column)
        if value is None or (isinstance(value, str) and not value.strip()):
            empty_rows.append(row_num)
    return empty_rows


def calculate_fill_rate(sheet) -> float:
    """计算工作表的填充率（百分比）"""
    if not sheet.max_row or not sheet.max_column:
        return 0.0

    total_cells = sheet.max_row * sheet.max_column
    non_empty = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                non_empty += 1

    return round(non_empty / total_cells * 100, 2) if total_cells else 0.0

def validate_formula_syntax_v2(formula: str) -> Tuple[bool, Optional[str]]:
    """
    验证新版公式语法: [工作表名]A1（兼容旧格式）

    Args:
        formula: 公式字符串

    Returns:
        Tuple[bool, Optional[str]]: (是否有效, 错误信息)
    """
    if not formula or not formula.strip():
        return False, "公式不能为空"

    try:
        # 验证公式中的引用格式
        references = parse_formula_references_v2(formula)
        if not references:
            return False, "公式中未发现有效的引用"

        # 检查数学表达式语法
        # 将引用替换为数字进行语法检查
        test_formula = formula
        for ref_data in references:
            test_formula = test_formula.replace(ref_data['full_reference'], '1')

        # 验证数学表达式
        try:
            # 只允许基本数学运算符
            allowed_chars = set('0123456789+-*/()., ')
            if not all(c in allowed_chars for c in test_formula):
                return False, "公式包含不支持的字符"

            # 尝试编译表达式（不执行）
            compile(test_formula, '<string>', 'eval')
            return True, None

        except SyntaxError as e:
            return False, f"数学表达式语法错误: {str(e)}"

    except Exception as e:
        return False, f"公式验证失败: {str(e)}"


def build_formula_reference_simple(sheet_name: str, cell_address: str) -> str:
    """构建新格式的公式引用字符串 [工作表]单元格"""
    if not sheet_name:
        sheet_name = ""
    sheet_part = str(sheet_name).strip()
    cell_part = str(cell_address or "").strip().upper()
    return f"[{sheet_part}]{cell_part}"


def parse_formula_references_simple(formula: str) -> List[Dict[str, str]]:
    """解析新格式公式引用

    Returns 字典列表，包含 sheet_name、cell_address、full_reference
    """
    if not formula:
        return []

    references: List[Dict[str, str]] = []
    for match in SIMPLE_REFERENCE_PATTERN.finditer(formula):
        sheet_name = match.group("sheet").strip()
        cell_address = match.group("cell").strip().upper()
        references.append({
            "sheet_name": sheet_name,
            "item_name": None,
            "column_key": None,
            "cell_address": cell_address,
            "full_reference": match.group(0)
        })

    return references


def convert_formula_to_new_format(formula: str) -> str:
    """将旧格式公式转换为新格式，如果已是新格式则原样返回"""
    if not formula:
        return formula

    converted_formula = formula
    # 优先解析 v3（包含列键）
    references = parse_formula_references_v3(converted_formula)
    if not references:
        references = parse_formula_references_v2(converted_formula)

    if not references:
        return converted_formula

    for ref in references:
        full_ref = ref.get('full_reference')
        sheet_name = ref.get('sheet_name')
        cell_address = ref.get('cell_address')

        if not full_ref or not sheet_name or not cell_address:
            continue

        new_ref = build_formula_reference_simple(sheet_name, cell_address)
        converted_formula = converted_formula.replace(full_ref, new_ref)

    return converted_formula


def parse_formula_references_v2(formula: str) -> List[Dict[str, str]]:
    """
    解析公式中的引用。

    优先识别新格式 [工作表]A1，若不存在则回退到旧格式
    [工作表:"项目"](单元格地址) 以及 [工作表:"项目:列"](单元格地址)。

    Args:
        formula: 公式字符串

    Returns:
        List[Dict[str, str]]: 引用信息列表
        格式: [{'sheet_name': str, 'item_name': str, 'column_key': str, 'cell_address': str, 'full_reference': str}]
    """
    simple_refs = parse_formula_references_simple(formula)
    if simple_refs:
        return simple_refs

    references = []

    # 正则表达式匹配带列名的格式: [工作表名:"项目名:列名"](单元格地址)
    column_pattern = r'\[([^\]]+):"([^"]+):([^"]+)"\]\(([A-Z]+\d+)\)'

    for match in re.finditer(column_pattern, formula):
        sheet_name = match.group(1).strip()
        item_name = match.group(2).strip()
        column_key = match.group(3).strip()
        cell_address = match.group(4).strip()
        full_reference = match.group(0)

        references.append({
            'sheet_name': sheet_name,
            'item_name': item_name,
            'column_key': column_key,
            'cell_address': cell_address,
            'full_reference': full_reference
        })

    # 正则表达式匹配标准格式: [工作表名:"项目名"](单元格地址)
    # 排除已经匹配到列名格式的引用
    standard_pattern = r'\[([^\]]+):"([^"]+)"\]\(([A-Z]+\d+)\)'

    for match in re.finditer(standard_pattern, formula):
        # 检查这个匹配是否已经被列名格式匹配了
        if ':' not in match.group(2):  # 项目名中没有冒号，说明不是列名格式
            sheet_name = match.group(1).strip()
            item_name = match.group(2).strip()
            cell_address = match.group(3).strip()
            full_reference = match.group(0)

            references.append({
                'sheet_name': sheet_name,
                'item_name': item_name,
                'column_key': '',  # 标准格式没有列名
                'cell_address': cell_address,
                'full_reference': full_reference
            })

    return references


def parse_formula_references_v3(formula: str) -> List[Dict[str, str]]:
    """
    解析公式中的引用 - 支持多列数据格式

    支持格式:
    - [工作表:"项目名"](单元格) - 标准格式
    - [工作表:"项目名"|列键](单元格) - 多列格式

    Args:
        formula: 公式字符串

    Returns:
        list: 引用信息列表，每个元素包含 sheet_name, item_name, column_key(可选), cell_address
    """
    references = []

    # 匹配多列格式: [工作表:"项目名"|列键](单元格)
    multi_column_pattern = r'\[([^:]+):"([^"]+)"\|([^\]]+)\]\(([^)]+)\)'
    matches = re.finditer(multi_column_pattern, formula)

    for match in matches:
        sheet_name = match.group(1).strip()
        item_name = match.group(2).strip()
        column_key = match.group(3).strip()
        cell_address = match.group(4).strip()

        references.append({
            'sheet_name': sheet_name,
            'item_name': item_name,
            'column_key': column_key,
            'cell_address': cell_address,
            'full_reference': match.group(0)
        })

    # 匹配标准格式: [工作表:"项目名"](单元格)
    standard_pattern = r'\[([^:]+):"([^"]+)"\]\(([^)]+)\)'
    matches = re.finditer(standard_pattern, formula)

    for match in matches:
        # 确保不是多列格式的一部分
        if '|' not in match.group(0):
            sheet_name = match.group(1).strip()
            item_name = match.group(2).strip()
            cell_address = match.group(3).strip()

            references.append({
                'sheet_name': sheet_name,
                'item_name': item_name,
                'column_key': None,
                'cell_address': cell_address,
                'full_reference': match.group(0)
            })

    return references


def build_formula_reference_v2(sheet_name: str, item_name: str, cell_address: str, column_key: str = None) -> str:
    """
    兼容接口：返回新格式公式引用字符串 [工作表]单元格。

    保留原参数签名以兼容旧调用，内部仅使用工作表与单元格。
    """
    return build_formula_reference_simple(sheet_name, cell_address)


def build_formula_reference_v3(sheet_name: str, item_name: str, column_key: str = None, cell_address: str = None) -> str:
    """
    构建新版公式引用字符串（支持多列数据）

    Args:
        sheet_name: 工作表名
        item_name: 项目名
        column_key: 数据列键（如"年初余额_借方"，可选）
        cell_address: 单元格地址（可选，如果没有则使用默认）

    Returns:
        str: 公式引用字符串
    """
    if column_key:
        return f'[{sheet_name}:"{item_name}"|{column_key}]({cell_address or "AUTO"})'
    else:
        return f'[{sheet_name}:"{item_name}"]({cell_address or "AUTO"})'


# ==================== 三段式公式引用系统 (新版) ====================

def normalize_formula_punctuation(formula: str) -> str:
    """
    规范化公式中的标点符号，支持中文输入法

    将中文标点转换为英文标点：
    - 中文感叹号 ！ → 英文感叹号 !
    - 中文方括号 【】 → 英文方括号 []

    注意：不转换小括号（），因为：
    1. 三段式公式格式是[工作表]![项目]![列名]，不使用小括号作为语法符号
    2. Excel工作表名可能包含中文字符，转换括号会破坏工作表名匹配
    3. 小括号只用于数学运算，用户应直接输入英文括号

    Args:
        formula: 可能包含中文标点的公式字符串

    Returns:
        str: 标点规范化后的公式字符串
    """
    if not formula:
        return formula

    # 中文标点 → 英文标点（只转换公式语法符号）
    normalized = formula
    normalized = normalized.replace('！', '!')   # 中文感叹号
    normalized = normalized.replace('【', '[')   # 中文左方括号
    normalized = normalized.replace('】', ']')   # 中文右方括号
    # ⚠️ 不转换小括号（），避免破坏Excel工作表名

    return normalized

def parse_formula_references_three_segment(formula: str) -> List[Dict[str, str]]:
    """
    解析三段式公式引用: [工作表名]![项目名]![列名]

    这是最新的公式引用格式,使用感叹号分隔三个部分:
    - 工作表名: Excel中的sheet名称
    - 项目名: 数据项名称(如"营业收入")
    - 列名: 数据列名称(如"本期金额"、"本年累计")

    Args:
        formula: 公式字符串,如 "[利润表]![营业收入]![本年累计] + [利润表]![营业成本]![本年累计]"

    Returns:
        List[Dict]: 引用信息列表,每个元素包含:
            - sheet_name: 工作表名
            - item_name: 项目名
            - column_name: 列名
            - full_reference: 完整引用字符串
    """
    if not formula:
        return []

    # ⭐ 规范化中文标点，支持用户使用中文输入法
    formula = normalize_formula_punctuation(formula)

    references = []
    # 正则模式: [工作表名]![项目名]![列名]
    pattern = r'\[([^\]]+)\]!\[([^\]]+)\]!\[([^\]]+)\]'

    for match in re.finditer(pattern, formula):
        sheet_name = match.group(1).strip()
        item_name = match.group(2).strip()
        column_name = match.group(3).strip()

        references.append({
            'sheet_name': sheet_name,
            'item_name': item_name,
            'column_name': column_name,
            'column_key': None,  # 三段式不使用column_key
            'cell_address': None,  # 三段式不使用cell_address
            'full_reference': match.group(0)
        })

    return references


def build_formula_reference_three_segment(sheet_name: str, item_name: str, column_name: str) -> str:
    """
    构建三段式公式引用字符串: [工作表名]![项目名]![列名]

    Args:
        sheet_name: 工作表名
        item_name: 项目名
        column_name: 列名

    Returns:
        str: 三段式引用字符串,如 "[利润表]![营业收入]![本年累计]"
    """
    return f"[{sheet_name}]![{item_name}]![{column_name}]"


def validate_formula_syntax_three_segment(formula: str, workbook_manager=None) -> Tuple[bool, Optional[str]]:
    """
    验证三段式公式语法,并检查引用的有效性

    Args:
        formula: 公式字符串
        workbook_manager: 工作簿管理器(可选,用于验证引用存在性)

    Returns:
        Tuple[bool, Optional[str]]: (是否有效, 错误信息)
    """
    if not formula or not formula.strip():
        return False, "公式不能为空"

    # ⭐ 规范化中文标点
    formula = normalize_formula_punctuation(formula)

    try:
        # 解析三段式引用
        references = parse_formula_references_three_segment(formula)

        if not references:
            return False, "公式中未发现有效的三段式引用"

        # 如果提供了workbook_manager,验证引用存在性
        if workbook_manager:
            for ref in references:
                sheet_name = ref['sheet_name']
                item_name = ref['item_name']
                column_name = ref['column_name']

                # 检查工作表是否存在
                if sheet_name not in workbook_manager.worksheets:
                    return False, f"未找到工作表'{sheet_name}'"

                # 检查来源项是否存在（比对时strip去除空格，支持前缀匹配）
                source_item = None
                for source in workbook_manager.source_items.values():
                    # ⭐ 比对时对两边都strip()，但不改变原始值（保留用于UI显示）
                    source_name_stripped = source.name.strip() if isinstance(source.name, str) else source.name
                    item_name_stripped = item_name.strip() if isinstance(item_name, str) else item_name

                    if source.sheet_name == sheet_name and source_name_stripped == item_name_stripped:
                        source_item = source
                        break

                # 如果直接查找失败，尝试添加常见财务报表前缀
                if not source_item:
                    common_prefixes = [
                        "加：", "减：", "其中：", "其中:",
                        "*", "☆", "△", "▲", "√"
                    ]

                    for prefix in common_prefixes:
                        prefixed_name = prefix + item_name_stripped
                        for source in workbook_manager.source_items.values():
                            source_name_stripped = source.name.strip() if isinstance(source.name, str) else source.name
                            if source.sheet_name == sheet_name and source_name_stripped == prefixed_name:
                                source_item = source
                                break
                        if source_item:
                            break

                if not source_item:
                    return False, f"在工作表'{sheet_name}'中未找到项目'{item_name}'"

                # 检查列名是否存在（比对时strip去除空格）
                if hasattr(source_item, 'values') and isinstance(source_item.values, dict):
                    # ⭐ 对列名进行strip比对
                    column_found = False
                    for col_key in source_item.values.keys():
                        col_key_stripped = col_key.strip() if isinstance(col_key, str) else col_key
                        column_name_stripped = column_name.strip() if isinstance(column_name, str) else column_name
                        if col_key_stripped == column_name_stripped:
                            column_found = True
                            break

                    if not column_found:
                        available_columns = list(source_item.values.keys())
                        return False, f"项目'{item_name}'不包含列'{column_name}',可用列: {', '.join(available_columns)}"

        # 检查数学表达式语法
        test_formula = formula
        for ref in references:
            test_formula = test_formula.replace(ref['full_reference'], '1')

        # 验证数学表达式
        allowed_chars = set('0123456789+-*/()., ')
        if not all(c in allowed_chars for c in test_formula):
            return False, "公式包含不支持的字符"

        try:
            compile(test_formula, '<string>', 'eval')
            return True, None
        except SyntaxError as e:
            return False, f"数学表达式语法错误: {str(e)}"

    except Exception as e:
        return False, f"公式验证失败: {str(e)}"


def evaluate_formula_with_values_three_segment(formula: str, source_items: Dict[str, Any]) -> Tuple[bool, Union[float, str]]:
    """
    使用三段式引用计算公式

    Args:
        formula: 三段式公式字符串
        source_items: 来源项字典 {item_id: SourceItem}

    Returns:
        Tuple[bool, Union[float, str]]: (成功标志, 计算结果或错误信息)
    """
    if not formula or not formula.strip():
        return False, "公式为空"

    # ⭐ 规范化中文标点
    formula = normalize_formula_punctuation(formula)

    try:
        # 解析三段式引用
        references = parse_formula_references_three_segment(formula)

        if not references:
            # 可能是纯数学表达式
            try:
                result = eval(formula)
                return True, float(result) if isinstance(result, (int, float)) else result
            except:
                return False, "无法计算表达式"

        # 构建值映射表
        working_formula = formula

        def _coerce_numeric_value(value: Any) -> float:
            """将引用值转换为可计算的数值, 支持常见占位符视为0"""
            if value is None:
                return 0.0

            if isinstance(value, (int, float)):
                return float(value)

            value_str = str(value).strip()

            placeholder_tokens = {"", "-", "--", "—", "–", "－"}
            if value_str in placeholder_tokens:
                return 0.0

            cleaned = value_str.replace(",", "")
            if cleaned.startswith("(") and cleaned.endswith(")"):
                cleaned = f"-{cleaned[1:-1]}"

            try:
                return float(cleaned)
            except (ValueError, TypeError):
                raise ValueError(f"引用值'{value}'不是有效的数字")

        for ref in references:
            sheet_name = ref['sheet_name']
            item_name = ref['item_name']
            column_name = ref['column_name']

            # 查找对应的来源项（比对时strip去除空格，支持前缀匹配）
            ref_value = None
            source_item_found = None

            # 第一次查找：直接匹配
            for source in source_items.values():
                # ⭐ 比对时对两边都strip()，但不改变原始值
                source_name_stripped = source.name.strip() if isinstance(source.name, str) else source.name
                item_name_stripped = item_name.strip() if isinstance(item_name, str) else item_name

                if source.sheet_name == sheet_name and source_name_stripped == item_name_stripped:
                    source_item_found = source
                    break

            # 第二次查找：如果直接匹配失败，尝试添加常见财务报表前缀
            if not source_item_found:
                common_prefixes = [
                    "加：", "减：", "其中：", "其中:",
                    "*", "☆", "△", "▲", "√"
                ]

                item_name_stripped = item_name.strip() if isinstance(item_name, str) else item_name
                for prefix in common_prefixes:
                    prefixed_name = prefix + item_name_stripped
                    for source in source_items.values():
                        source_name_stripped = source.name.strip() if isinstance(source.name, str) else source.name
                        if source.sheet_name == sheet_name and source_name_stripped == prefixed_name:
                            source_item_found = source
                            break
                    if source_item_found:
                        break

            # 如果找到来源项，从values字典中获取对应列的值
            if source_item_found:
                if hasattr(source_item_found, 'values') and isinstance(source_item_found.values, dict):
                    # ⭐ 对列名进行strip比对查找
                    column_name_stripped = column_name.strip() if isinstance(column_name, str) else column_name
                    for col_key, col_value in source_item_found.values.items():
                        col_key_stripped = col_key.strip() if isinstance(col_key, str) else col_key
                        if col_key_stripped == column_name_stripped:
                            ref_value = col_value
                            break

            if ref_value is None:
                return False, f"无法找到引用值: {ref['full_reference']}"

            # 替换引用为实际值
            try:
                numeric_value = _coerce_numeric_value(ref_value)
                working_formula = working_formula.replace(ref['full_reference'], str(numeric_value))
            except (ValueError, TypeError) as e:
                return False, str(e)

        # 计算最终结果
        try:
            result = eval(working_formula)
            return True, float(result) if isinstance(result, (int, float)) else result
        except Exception as e:
            return False, f"计算错误: {str(e)}"

    except Exception as e:
        return False, f"公式计算失败: {str(e)}"


def parse_formula_smart(formula: str) -> List[Dict[str, str]]:
    """
    智能识别并解析公式格式

    自动检测公式格式并使用相应的解析器:
    - 三段式: [表]![项]![列]
    - v3格式: [表:"项"|列键](单元格)
    - v2格式: [表:"项"](单元格)
    - simple格式: [表]单元格

    Args:
        formula: 公式字符串

    Returns:
        List[Dict]: 统一格式的引用信息列表
    """
    if not formula:
        return []

    # 检测三段式格式 (包含 ]![ 模式)
    if ']![' in formula:
        return parse_formula_references_three_segment(formula)

    # 检测v3格式 (包含 :" 和 | 模式)
    elif ':"' in formula and '|' in formula:
        return parse_formula_references_v3(formula)

    # 检测v2格式 (包含 :" 模式)
    elif ':"' in formula:
        return parse_formula_references_v2(formula)

    # 默认使用simple格式
    else:
        return parse_formula_references_simple(formula)


def evaluate_formula_with_values_v3(formula: str, source_items: Dict[str, Any]) -> Tuple[bool, Union[float, str]]:
    """
    使用来源项直接计算公式 - 支持多列数据格式

    Args:
        formula: 公式字符串
        source_items: 来源项字典 {item_id: SourceItem}

    Returns:
        Tuple[bool, Union[float, str]]: (成功标志, 计算结果或错误信息)
    """
    if not formula or not formula.strip():
        return False, "公式为空"

    try:
        # 解析引用
        references = parse_formula_references_v3(formula)

        if not references:
            # 可能是纯数学表达式
            try:
                result = eval(formula)
                return True, float(result) if isinstance(result, (int, float)) else result
            except:
                return False, "无法计算表达式"

        # 替换引用为实际值
        working_formula = formula
        for ref in references:
            ref_value = _resolve_reference_value_v3(ref, source_items)
            if ref_value is None:
                return False, f"找不到引用: {ref['full_reference']}"

            # 替换引用为数值
            working_formula = working_formula.replace(ref['full_reference'], str(ref_value))

        # 计算结果
        try:
            result = eval(working_formula)
            return True, float(result) if isinstance(result, (int, float)) else result
        except Exception as e:
            return False, f"计算错误: {str(e)}"

    except Exception as e:
        return False, f"公式解析错误: {str(e)}"


def _resolve_reference_value_v3(reference: Dict[str, str], source_items: Dict[str, Any]) -> Union[float, None]:
    """
    解析引用值 - 支持多列数据

    Args:
        reference: 引用信息
        source_items: 来源项字典

    Returns:
        float: 引用的值，如果找不到返回None
    """
    sheet_name = reference.get('sheet_name')
    item_name = reference.get('item_name')
    column_key = reference.get('column_key')
    cell_address = reference.get('cell_address')

    # 查找匹配的来源项
    for item in source_items.values():
        if (hasattr(item, 'sheet_name') and item.sheet_name == sheet_name and
            (
                (item_name and hasattr(item, 'name') and item.name == item_name) or
                (not item_name and hasattr(item, 'cell_address') and item.cell_address == cell_address)
            )):

            if column_key:
                # 多列数据模式
                if hasattr(item, 'data_columns') and item.data_columns:
                    value = item.data_columns.get(column_key)
                    if value is not None:
                        try:
                            return float(value)
                        except (ValueError, TypeError):
                            return None
            else:
                # 标准模式 - 使用主要值
                if hasattr(item, 'value') and item.value is not None:
                    try:
                        return float(item.value)
                    except (ValueError, TypeError):
                        return None

    return None


def evaluate_formula_with_values_v2(formula: str, value_map: Dict[str, Any]) -> Tuple[bool, Union[float, str]]:
    """
    使用值映射表计算公式 - 新版格式

    Args:
        formula: 公式字符串
        value_map: 值映射表 {引用字符串: 值}

    Returns:
        Tuple[bool, Union[float, str]]: (成功标志, 计算结果或错误信息)
    """
    if not formula or not formula.strip():
        return False, "公式为空"

    try:
        # 解析引用
        references = parse_formula_references_v2(formula)

        if not references:
            # 可能是纯数学表达式
            try:
                result = eval(formula)
                return True, float(result) if isinstance(result, (int, float)) else result
            except:
                return False, "无法计算表达式"

        # 替换引用为实际值
        calculated_formula = formula

        for ref_data in references:
            full_ref = ref_data['full_reference']

            # 在值映射表中查找对应的值
            value = None

            # 方法1: 直接匹配完整引用
            if full_ref in value_map:
                value = value_map[full_ref]
            else:
                # 方法2: 通过工作表名和项目名匹配
                sheet_name = ref_data.get('sheet_name')
                item_name = ref_data.get('item_name')
                if sheet_name and item_name:
                    for map_key, map_value in value_map.items():
                        # 如果值映射表使用旧格式，需要转换匹配
                        if (sheet_name in map_key and item_name in map_key):
                            value = map_value
                            break

            if value is None:
                return False, f"未找到引用: {full_ref}"

            # 确保值是数字
            if not isinstance(value, (int, float)):
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    return False, f"引用值不是数字: {full_ref} = {value}"

            # 替换引用为值
            calculated_formula = calculated_formula.replace(full_ref, str(value))

        # 计算最终表达式
        try:
            result = eval(calculated_formula)
            return True, float(result)
        except Exception as e:
            return False, f"计算错误: {str(e)}"

    except Exception as e:
        return False, f"公式评估失败: {str(e)}"


def convert_old_formula_to_new(old_formula: str, source_items: Dict[str, Any]) -> str:
    """
    将旧格式公式转换为新格式
    旧格式: [工作表名]![项目名]
    新格式: [工作表名:"项目名"](单元格地址)

    Args:
        old_formula: 旧格式公式
        source_items: 来源项字典，用于查找单元格地址

    Returns:
        str: 新格式公式
    """
    if not old_formula:
        return ""

    new_formula = old_formula

    # 匹配旧格式: [工作表名]![项目名]
    old_pattern = r'\[([^\]]+)\]!\[([^\]]+)\]'

    def replace_reference(match):
        sheet_name = match.group(1)
        item_name = match.group(2)

        # 查找对应的单元格地址
        cell_address = "A1"  # 默认地址

        # 在source_items中查找匹配的项目
        for source_id, source_item in source_items.items():
            if (hasattr(source_item, 'sheet_name') and
                hasattr(source_item, 'name') and
                hasattr(source_item, 'cell_address')):
                if (source_item.sheet_name == sheet_name and
                    source_item.name == item_name):
                    cell_address = source_item.cell_address
                    break

        return build_formula_reference_v2(sheet_name, item_name, cell_address)

    new_formula = re.sub(old_pattern, replace_reference, new_formula)
    return new_formula


def write_values_to_excel(workbook_path: str, values: Dict[str, Dict[str, Any]]) -> bool:
    """
    将值写入Excel文件

    Args:
        workbook_path: Excel文件路径
        values: 值字典 {工作表名: {单元格地址: 值}}

    Returns:
        bool: 是否成功
    """
    try:
        workbook = openpyxl.load_workbook(workbook_path)

        for sheet_name, sheet_values in values.items():
            if sheet_name not in workbook.sheetnames:
                continue

            sheet = workbook[sheet_name]

            for cell_address, value in sheet_values.items():
                try:
                    sheet[cell_address] = value
                except Exception as e:
                    print(f"写入失败 {sheet_name}!{cell_address}: {e}")

        workbook.save(workbook_path)
        workbook.close()
        return True

    except Exception as e:
        print(f"写入Excel文件失败: {e}")
        return False


def backup_excel_file(file_path: str) -> str:
    """
    创建Excel文件备份

    Args:
        file_path: 原文件路径

    Returns:
        str: 备份文件路径
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_dir = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)
    name, ext = os.path.splitext(file_name)

    backup_name = f"{name}_backup_{timestamp}{ext}"
    backup_path = os.path.join(file_dir, backup_name)

    shutil.copy2(file_path, backup_path)
    return backup_path


def extract_sheet_data(sheet, start_row: int = 1, end_row: Optional[int] = None) -> List[List[Any]]:
    """
    提取工作表数据

    Args:
        sheet: openpyxl工作表对象
        start_row: 开始行号（1开始）
        end_row: 结束行号（可选）

    Returns:
        List[List[Any]]: 行数据列表
    """
    if end_row is None:
        end_row = sheet.max_row

    data = []
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        data.append(list(row))

    return data


def find_data_range(sheet) -> Tuple[int, int, int, int]:
    """
    查找工作表中的数据范围

    Args:
        sheet: openpyxl工作表对象

    Returns:
        Tuple[int, int, int, int]: (最小行, 最大行, 最小列, 最大列)
    """
    min_row = sheet.min_row or 1
    max_row = sheet.max_row or 1
    min_col = sheet.min_column or 1
    max_col = sheet.max_column or 1

    # 查找实际有数据的范围
    actual_max_row = min_row
    actual_max_col = min_col

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                actual_max_row = max(actual_max_row, row)
                actual_max_col = max(actual_max_col, col)

    return min_row, actual_max_row, min_col, actual_max_col


def clean_cell_text(text: Any) -> str:
    """
    清理单元格文本

    Args:
        text: 单元格值

    Returns:
        str: 清理后的文本
    """
    if text is None:
        return ""

    text_str = str(text).strip()

    # 移除常见的格式字符
    text_str = re.sub(r'\s+', ' ', text_str)  # 多个空白字符替换为单个空格
    text_str = text_str.replace('\n', ' ').replace('\r', ' ')

    return text_str


def is_numeric_value(value: Any) -> bool:
    """
    检查值是否为数字

    Args:
        value: 要检查的值

    Returns:
        bool: 是否为数字
    """
    if value is None:
        return False

    if isinstance(value, (int, float)):
        return True

    try:
        float(str(value))
        return True
    except (ValueError, TypeError):
        return False


def format_currency_value(value: Union[int, float], precision: int = 2) -> str:
    """
    格式化货币值

    Args:
        value: 数值
        precision: 小数位数

    Returns:
        str: 格式化后的字符串
    """
    if not isinstance(value, (int, float)):
        return str(value)

    return f"{value:,.{precision}f}"


# 兼容性函数 - 保持向后兼容
def validate_formula_syntax(formula: str) -> Tuple[bool, Optional[str]]:
    """向后兼容的公式验证函数"""
    return validate_formula_syntax_v2(formula)


def parse_formula_references(formula: str) -> List[Dict[str, str]]:
    """向后兼容的公式解析函数"""
    return parse_formula_references_v2(formula)


def build_formula_reference(sheet_name: str, item_name: str) -> str:
    """向后兼容的引用构建函数（使用默认单元格地址）"""
    return build_formula_reference_v2(sheet_name, item_name, "A1")


def evaluate_formula_with_values(formula: str, value_map: Dict[str, Any]) -> Tuple[bool, Union[float, str]]:
    """向后兼容的公式计算函数"""
    return evaluate_formula_with_values_v2(formula, value_map)
