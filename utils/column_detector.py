#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
列头检测器 - 智能识别和解析财务表格的列头结构
"""

import re
from typing import Dict, List, Optional, Tuple, Set
from dataclasses import dataclass
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

@dataclass
class ColumnHeader:
    """列头信息"""
    column_index: int
    column_letter: str
    merged_range: Optional[Tuple[int, int, int, int]] = None  # (min_row, max_row, min_col, max_col)
    level1_text: str = ""  # 一级列头文本
    level2_text: str = ""  # 二级列头文本
    full_text: str = ""   # 完整文本
    data_type: str = "unknown"  # 数据类型
    category: str = ""    # 数据分类

class ColumnDetector:
    """列头检测器"""

    def __init__(self):
        # 财务术语分类
        self.financial_terms = {
            'balance_types': {
                'beginning': [r'期初', r'年初', r'opening', r'beginning'],
                'ending': [r'期末', r'年末', r'closing', r'ending'],
                'current': [r'本期', r'本年', r'current', r'this.*period'],
                'previous': [r'上期', r'上年', r'去年', r'previous', r'last.*period']
            },
            'debit_credit': {
                'debit': [r'借方', r'debit', r'dr\.?$'],
                'credit': [r'贷方', r'credit', r'cr\.?$']
            },
            'amount_types': {
                'amount': [r'金额', r'数额', r'amount', r'value'],
                'balance': [r'余额', r'balance'],
                'total': [r'合计', r'总计', r'total', r'sum'],
                'subtotal': [r'小计', r'subtotal']
            },
            'sheet_specific': {
                'trial_balance': [
                    r'年初余额', r'期初余额', r'本期发生额', r'期末余额',
                    r'年初借方', r'年初贷方', r'期初借方', r'期初贷方',
                    r'本期借方', r'本期贷方', r'期末借方', r'期末贷方'
                ],
                'income_statement': [
                    r'本期金额', r'上期金额', r'本年累计', r'上年累计'
                ],
                'balance_sheet': [
                    r'期末余额', r'年初余额', r'期末数', r'年初数'
                ]
            }
        }

    def detect_columns(self, sheet: Worksheet, header_rows: int = 2) -> List[ColumnHeader]:
        """检测所有列头"""
        columns = []

        # 处理合并单元格
        merged_ranges = self._get_merged_ranges(sheet, header_rows)

        # 分析每一列
        for col in range(1, sheet.max_column + 1):
            column_header = self._analyze_column(sheet, col, header_rows, merged_ranges)
            if column_header:
                columns.append(column_header)

        return columns

    def _get_merged_ranges(self, sheet: Worksheet, header_rows: int) -> List[Tuple[int, int, int, int]]:
        """获取列头区域的合并单元格范围"""
        merged_ranges = []

        for merged_range in sheet.merged_cells.ranges:
            # 只关注列头区域的合并单元格
            if merged_range.min_row <= header_rows:
                merged_ranges.append((
                    merged_range.min_row,
                    merged_range.max_row,
                    merged_range.min_col,
                    merged_range.max_col
                ))

        return merged_ranges

    def _analyze_column(self, sheet: Worksheet, col: int, header_rows: int,
                       merged_ranges: List[Tuple[int, int, int, int]]) -> Optional[ColumnHeader]:
        """分析单个列的列头信息"""
        col_letter = openpyxl.utils.get_column_letter(col)

        # 收集列头文本
        level1_text = ""
        level2_text = ""
        merged_range = None

        # 检查是否在合并单元格中
        for merge_range in merged_ranges:
            min_row, max_row, min_col, max_col = merge_range
            if min_col <= col <= max_col:
                merged_range = merge_range
                # 从合并单元格的起始位置获取文本
                cell = sheet.cell(row=min_row, column=min_col)
                if cell.value:
                    if min_row == 1:
                        level1_text = str(cell.value).strip()
                    elif min_row == 2:
                        level2_text = str(cell.value).strip()
                break

        # 如果不在合并单元格中，分别获取各行的文本
        if not merged_range:
            if header_rows >= 1:
                cell1 = sheet.cell(row=1, column=col)
                if cell1.value:
                    level1_text = str(cell1.value).strip()

            if header_rows >= 2:
                cell2 = sheet.cell(row=2, column=col)
                if cell2.value:
                    level2_text = str(cell2.value).strip()

        # 如果没有任何文本，跳过
        if not level1_text and not level2_text:
            return None

        # 组合完整文本
        full_text = self._combine_header_text(level1_text, level2_text)

        # 识别数据类型和分类
        data_type, category = self._classify_column(full_text, level1_text, level2_text)

        return ColumnHeader(
            column_index=col,
            column_letter=col_letter,
            merged_range=merged_range,
            level1_text=level1_text,
            level2_text=level2_text,
            full_text=full_text,
            data_type=data_type,
            category=category
        )

    def _combine_header_text(self, level1: str, level2: str) -> str:
        """组合列头文本"""
        if level1 and level2:
            return f"{level1}-{level2}"
        elif level1:
            return level1
        elif level2:
            return level2
        else:
            return ""

    def _classify_column(self, full_text: str, level1: str, level2: str) -> Tuple[str, str]:
        """分类列头"""
        combined_text = (full_text + " " + level1 + " " + level2).lower()

        # 优先检查科目余额表特有的列头
        if self._match_patterns(combined_text, self.financial_terms['sheet_specific']['trial_balance']):
            return self._classify_trial_balance_column(combined_text)

        # 检查借方/贷方
        if self._match_patterns(combined_text, self.financial_terms['debit_credit']['debit']):
            return 'debit', self._get_balance_period(combined_text)
        elif self._match_patterns(combined_text, self.financial_terms['debit_credit']['credit']):
            return 'credit', self._get_balance_period(combined_text)

        # 检查余额类型
        balance_category = self._get_balance_period(combined_text)
        if balance_category:
            return 'balance', balance_category

        # 检查金额类型
        if self._match_patterns(combined_text, self.financial_terms['amount_types']['amount']):
            period = self._get_period_type(combined_text)
            return 'amount', period if period else 'current'

        return 'unknown', 'unknown'

    def _classify_trial_balance_column(self, text: str) -> Tuple[str, str]:
        """分类科目余额表列头"""
        # 年初余额
        if re.search(r'年初.*余额', text):
            if re.search(r'借方|debit', text):
                return 'debit', 'beginning_balance'
            elif re.search(r'贷方|credit', text):
                return 'credit', 'beginning_balance'
            else:
                return 'balance', 'beginning'

        # 期初余额
        elif re.search(r'期初.*余额', text):
            if re.search(r'借方|debit', text):
                return 'debit', 'opening_balance'
            elif re.search(r'贷方|credit', text):
                return 'credit', 'opening_balance'
            else:
                return 'balance', 'opening'

        # 本期发生额
        elif re.search(r'本期.*发生', text):
            if re.search(r'借方|debit', text):
                return 'debit', 'current_movement'
            elif re.search(r'贷方|credit', text):
                return 'credit', 'current_movement'
            else:
                return 'movement', 'current'

        # 期末余额
        elif re.search(r'期末.*余额', text):
            if re.search(r'借方|debit', text):
                return 'debit', 'ending_balance'
            elif re.search(r'贷方|credit', text):
                return 'credit', 'ending_balance'
            else:
                return 'balance', 'ending'

        return 'unknown', 'unknown'

    def _get_balance_period(self, text: str) -> str:
        """获取余额期间"""
        if self._match_patterns(text, self.financial_terms['balance_types']['beginning']):
            return 'beginning'
        elif self._match_patterns(text, self.financial_terms['balance_types']['ending']):
            return 'ending'
        elif self._match_patterns(text, self.financial_terms['balance_types']['current']):
            return 'current'
        elif self._match_patterns(text, self.financial_terms['balance_types']['previous']):
            return 'previous'

        return 'unknown'

    def _get_period_type(self, text: str) -> str:
        """获取期间类型"""
        return self._get_balance_period(text)

    def _match_patterns(self, text: str, patterns: List[str]) -> bool:
        """匹配模式列表"""
        for pattern in patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return True
        return False

    def get_data_column_mapping(self, columns: List[ColumnHeader]) -> Dict[str, List[ColumnHeader]]:
        """获取数据列映射"""
        mapping = {}

        for column in columns:
            if column.data_type != 'unknown':
                category_key = f"{column.data_type}_{column.category}"
                if category_key not in mapping:
                    mapping[category_key] = []
                mapping[category_key].append(column)

        return mapping

    def get_trial_balance_structure(self, columns) -> Dict[str, Dict[str, int]]:
        """获取科目余额表结构（增强版）"""
        structure = {
            'beginning_balance': {'debit': None, 'credit': None, 'total': None},
            'opening_balance': {'debit': None, 'credit': None, 'total': None},
            'current_movement': {'debit': None, 'credit': None, 'total': None},
            'ending_balance': {'debit': None, 'credit': None, 'total': None}
        }

        for column in columns:
            # 获取完整的列头文本（支持二级列头）
            header_text = self._get_full_header_text(column)

            # 解析列头结构
            balance_type, direction_type = self._parse_trial_balance_header(header_text)

            # 保存映射关系
            if balance_type and balance_type in structure and direction_type:
                structure[balance_type][direction_type] = column.column_index

        return structure

    def _get_full_header_text(self, column) -> str:
        """获取完整的列头文本"""
        if hasattr(column, 'primary_header'):
            primary = column.primary_header or ""
            secondary = getattr(column, 'secondary_header', '') or ""
            return f"{primary} {secondary}".strip()
        elif hasattr(column, 'full_text'):
            return column.full_text or ""
        elif hasattr(column, 'level1_text'):
            level1 = column.level1_text or ""
            level2 = getattr(column, 'level2_text', '') or ""
            return f"{level1} {level2}".strip()
        else:
            return str(column)

    def _parse_trial_balance_header(self, header_text: str) -> tuple[str, str]:
        """解析科目余额表列头（支持二级列头）"""
        if not header_text:
            return None, None

        # 规范化文本
        text = header_text.lower().strip()

        # 确定余额类型（一级列头信息）
        balance_type = None
        if re.search(r'期初|年初|opening|beginning', text):
            balance_type = 'beginning_balance'
        elif re.search(r'期末|年末|closing|ending', text):
            balance_type = 'ending_balance'
        elif re.search(r'本期|current|movement|发生', text):
            balance_type = 'current_movement'
        elif re.search(r'开始|起始', text):
            balance_type = 'opening_balance'

        # 确定借贷方向（二级列头信息）
        direction_type = None
        if re.search(r'借方|debit|dr\.?$', text):
            direction_type = 'debit'
        elif re.search(r'贷方|credit|cr\.?$', text):
            direction_type = 'credit'
        elif re.search(r'合计|总计|total|sum', text):
            direction_type = 'total'
        elif re.search(r'余额|balance', text) and not re.search(r'借|贷|debit|credit', text):
            direction_type = 'total'  # 如果只有"余额"没有借贷标识，视为合计

        return balance_type, direction_type

if __name__ == "__main__":
    # 测试代码
    print("列头检测器模块")