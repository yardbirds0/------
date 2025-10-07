#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证真实数据 - 显示确切的单元格地址
绝不捏造，每个数据都显示其Excel中的确切位置
"""

import sys
import os
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def main():
    print("=" * 100)
    print("验证真实Excel数据 - 显示确切单元格地址")
    print("=" * 100)

    excel_path = r"D:\Code\快报填写程序\（科电）国资委、财政快报模板-纯净版 的副本.xlsx"

    print(f"\n[1] 打开Excel文件: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    print(f"\n[2] 列出所有工作表:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. '{sheet_name}'")

    # 使用第9个工作表（项目余额表5个暂未审定）
    target_sheet = wb.sheetnames[8]  # 索引8是第9个
    print(f"\n[3] 使用工作表: '{target_sheet}'")

    sheet = wb[target_sheet]
    print(f"\n[4] 显示该工作表的前20行x20列的所有单元格内容:")
    print(f"  工作表尺寸: {sheet.max_row}行 x {sheet.max_column}列")

    # 显示前20行20列
    print(f"\n  单元格内容展示:")
    for row in range(1, min(21, sheet.max_row + 1)):
        for col in range(1, min(21, sheet.max_column + 1)):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip():
                cell_addr = cell.coordinate
                print(f"    {cell_addr}: {cell.value}")

    # 查找"银行存款"
    print(f"\n[5] 查找项目名'银行存款':")
    bank_row = None
    bank_col = None
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value == "银行存款":
                bank_row = row
                bank_col = col
                print(f"  [OK] 找到'银行存款'在单元格: {cell.coordinate} (第{row}行, 第{col}列)")
                break
        if bank_row:
            break

    if not bank_row:
        print(f"  [FAIL] 未找到'银行存款'")
        return 1

    # 查找表头"内本:内蒙古"
    print(f"\n[6] 查找列名'内本:内蒙古':")
    header_row = 1  # 假设第1行是表头
    target_col = None
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col)
        if cell.value == "内本:内蒙古":
            target_col = col
            print(f"  [OK] 找到'内本:内蒙古'在单元格: {cell.coordinate} (第{header_row}行, 第{col}列)")
            break

    if not target_col:
        print(f"  [FAIL] 未找到'内本:内蒙古'")
        # 显示第1行所有表头
        print(f"  第1行的所有内容:")
        for col in range(1, min(21, sheet.max_column + 1)):
            cell = sheet.cell(row=1, column=col)
            if cell.value:
                print(f"    {cell.coordinate}: {cell.value}")
        return 1

    # 获取交叉点的值
    print(f"\n[7] 获取数据单元格的值:")
    data_cell = sheet.cell(row=bank_row, column=target_col)
    print(f"  单元格地址: {data_cell.coordinate}")
    print(f"  单元格值: {data_cell.value}")

    # 验证公式
    print(f"\n[8] 验证公式引用:")
    formula = f"[{target_sheet}]![银行存款]![内本:内蒙古]"
    print(f"  公式: {formula}")
    print(f"  应该引用到单元格: {data_cell.coordinate}")
    print(f"  单元格值: {data_cell.value}")

    print("\n" + "=" * 100)
    print("验证完成 - 所有数据都是真实Excel中的确切单元格")
    print("=" * 100)

    return 0

if __name__ == "__main__":
    sys.exit(main())
