#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
独立Excel验证程序 - 使用真实Excel文件测试三段式公式系统
"""

import sys
import os
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.data_models import WorkbookManager
from modules.data_extractor import DataExtractor
from utils.excel_utils import (
    parse_formula_references_three_segment,
    validate_formula_syntax_three_segment,
    evaluate_formula_with_values_three_segment,
    build_formula_reference_three_segment
)

def main():
    print("=" * 80)
    print("独立Excel验证程序 - 三段式公式系统完整测试")
    print("=" * 80)

    # Step 1: 加载真实Excel
    excel_path = r"D:\Code\快报填写程序\（科电）国资委、财政快报模板-纯净版 的副本.xlsx"

    if not os.path.exists(excel_path):
        excel_path = r"D:\Code\快报填写程序\sample.xlsx"

    print(f"\n[步骤1] 加载Excel文件")
    print(f"  文件路径: {excel_path}")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        print(f"  [OK] 成功加载，包含 {len(wb.sheetnames)} 个工作表")
        print(f"  工作表列表: {wb.sheetnames}")
    except Exception as e:
        print(f"  [FAIL] 加载失败: {e}")
        return 1

    # Step 2: 提取数据
    print(f"\n[步骤2] 提取Excel数据")
    try:
        # 先创建WorkbookManager并设置文件路径
        workbook_manager = WorkbookManager()
        workbook_manager.file_path = excel_path

        # 再创建DataExtractor
        extractor = DataExtractor(workbook_manager)

        # 提取数据
        success = extractor.extract_all_data()

        if not success:
            print(f"  [FAIL] extract_all_data返回False")
            return 1

        print(f"  [OK] 提取完成")
        print(f"  工作表数量: {len(workbook_manager.worksheets)}")
        print(f"  来源项数量: {len(workbook_manager.source_items)}")
        print(f"  目标项数量: {len(workbook_manager.target_items)}")
    except Exception as e:
        print(f"  [FAIL] 提取失败: {e}")
        import traceback
        traceback.print_exc()
        return 1

    # Step 3: 验证SourceItem.values填充
    print(f"\n[步骤3] 验证SourceItem.values字典")

    items_with_values = 0
    items_without_values = 0
    sample_items = []

    for item_id, source in workbook_manager.source_items.items():
        if hasattr(source, 'values') and source.values:
            items_with_values += 1
            if len(sample_items) < 3:
                sample_items.append(source)
        else:
            items_without_values += 1

    print(f"  有values的项: {items_with_values}")
    print(f"  无values的项: {items_without_values}")

    if sample_items:
        print(f"\n  示例项目:")
        for item in sample_items:
            print(f"    - {item.sheet_name} / {item.name}")
            print(f"      values: {item.values}")

    # Step 4: 测试用户公式
    print(f"\n[步骤4] 测试用户公式")

    # 尝试找到"利润表"和"一、营业总收入"
    test_formulas = []

    # 先看有没有利润表
    profit_sheet_items = [s for s in workbook_manager.source_items.values()
                          if '利润' in s.sheet_name]

    if profit_sheet_items:
        print(f"\n  找到利润相关工作表:")
        sheets_found = set(s.sheet_name for s in profit_sheet_items)
        for sheet in sheets_found:
            print(f"    - {sheet}")

        # 找营业收入
        revenue_items = [s for s in profit_sheet_items
                        if '营业收入' in s.name or '收入' in s.name]

        if revenue_items:
            print(f"\n  找到收入相关项目:")
            for item in revenue_items[:3]:
                print(f"    - {item.name} (工作表: {item.sheet_name})")
                if item.values:
                    print(f"      可用列: {list(item.values.keys())}")

                    # 构建测试公式
                    for col_name in item.values.keys():
                        formula = build_formula_reference_three_segment(
                            item.sheet_name, item.name, col_name
                        )
                        test_formulas.append({
                            'formula': formula,
                            'item': item,
                            'column': col_name
                        })
                        if len(test_formulas) >= 2:
                            break
                if test_formulas:
                    break

    # 如果没找到，使用第一个有values的项
    if not test_formulas and sample_items:
        item = sample_items[0]
        if item.values:
            col_name = list(item.values.keys())[0]
            formula = build_formula_reference_three_segment(
                item.sheet_name, item.name, col_name
            )
            test_formulas.append({
                'formula': formula,
                'item': item,
                'column': col_name
            })

    # 测试公式
    if test_formulas:
        print(f"\n  测试公式:")
        for i, test in enumerate(test_formulas[:2], 1):
            formula = test['formula']
            print(f"\n  测试{i}: {formula}")

            # 解析
            refs = parse_formula_references_three_segment(formula)
            print(f"    解析: {len(refs)}个引用")
            if refs:
                for ref in refs:
                    print(f"      - 工作表: {ref['sheet_name']}")
                    print(f"        项目: {ref['item_name']}")
                    print(f"        列名: {ref['column_name']}")

            # 验证
            is_valid, error = validate_formula_syntax_three_segment(formula, workbook_manager)
            if is_valid:
                print(f"    验证: [OK] 通过")
            else:
                print(f"    验证: [FAIL] {error}")

            # 计算
            success, result = evaluate_formula_with_values_three_segment(
                formula, workbook_manager.source_items
            )
            if success:
                print(f"    计算: [OK] 结果 = {result}")
            else:
                print(f"    计算: [FAIL] {result}")
    else:
        print(f"  [WARNING] 未找到可测试的公式")

    # Step 5: 测试用户的具体公式
    print(f"\n[步骤5] 测试用户指定的公式")
    user_formula = "[利润表]![一、营业总收入]![本期金额]"
    print(f"  公式: {user_formula}")

    # 解析
    refs = parse_formula_references_three_segment(user_formula)
    print(f"  解析: {len(refs)}个引用")

    # 验证
    is_valid, error = validate_formula_syntax_three_segment(user_formula, workbook_manager)
    if is_valid:
        print(f"  验证: [OK] 通过")
    else:
        print(f"  验证: [FAIL] {error}")

        # 如果失败，分析原因
        if refs:
            ref = refs[0]
            print(f"\n  失败分析:")
            print(f"    期望工作表: {ref['sheet_name']}")
            print(f"    期望项目: {ref['item_name']}")
            print(f"    期望列名: {ref['column_name']}")

            # 检查工作表
            if ref['sheet_name'] in workbook_manager.worksheets:
                print(f"    [OK] 工作表存在")
            else:
                print(f"    [FAIL] 工作表不存在")
                print(f"    可用工作表: {list(workbook_manager.worksheets.keys())}")

            # 检查项目
            found_item = None
            for item in workbook_manager.source_items.values():
                if item.sheet_name == ref['sheet_name'] and item.name == ref['item_name']:
                    found_item = item
                    break

            if found_item:
                print(f"    [OK] 项目存在")
                print(f"    可用列: {list(found_item.values.keys()) if found_item.values else 'None'}")
            else:
                print(f"    [FAIL] 项目不存在")
                # 列出该表的所有项目
                sheet_items = [s.name for s in workbook_manager.source_items.values()
                              if s.sheet_name == ref['sheet_name']]
                print(f"    该表的项目: {sheet_items[:10]}")

    print("\n" + "=" * 80)
    print("验证完成")
    print("=" * 80)

    return 0

if __name__ == "__main__":
    sys.exit(main())
