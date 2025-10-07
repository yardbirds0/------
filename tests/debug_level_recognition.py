#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
调试级别识别逻辑 - 专门测试"3.税金及附加"等项目的级别识别
"""

import sys
import os
import io

# 设置UTF-8输出
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.file_manager import FileManager
from modules.data_extractor import DataExtractor
from models.data_models import WorkbookManager, generate_hierarchical_numbers
import openpyxl

def debug_level_recognition():
    """调试级别识别"""

    # 1. 加载Excel文件
    excel_path = r"d:\Code\快报填写程序\（科电）国资委、财政快报模板-纯净版 的副本.xlsx"

    print("="*80)
    print("级别识别调试程序")
    print("="*80)
    print(f"\n加载文件: {os.path.basename(excel_path)}")

    # 2. 加载工作簿
    fm = FileManager()
    success, error_msg = fm.load_excel_files([excel_path])

    if not success:
        print(f"文件加载失败: {error_msg}")
        return

    workbook_manager = fm.workbook_manager
    if not workbook_manager:
        print("WorkbookManager未初始化")
        return

    # 3. 提取数据
    print("\n提取数据...")
    extractor = DataExtractor(workbook_manager)
    extractor.extract_all_data()

    print(f"提取到 {len(workbook_manager.target_items)} 个目标项")

    # 4. 加载原始Excel用于后续分析
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 5. 查找"税金及附加"项目
    target_item = None
    for tid, item in workbook_manager.target_items.items():
        if "税金及附加" in item.name:
            target_item = item
            break

    if not target_item:
        print("\n未找到'税金及附加'项目")
        return

    print("\n"+"="*80)
    print("找到目标项目")
    print("="*80)
    print(f"项目名称: {target_item.name}")
    print(f"工作表: {target_item.sheet_name}")
    print(f"行号: {target_item.row}")
    print(f"原始编号 (display_index): '{target_item.display_index}'")
    print(f"层级 (hierarchical_level): {target_item.hierarchical_level}")
    print(f"当前级别显示 (hierarchical_number): '{target_item.hierarchical_number}'")

    # 5. 使用原始Excel读取，看看真实数据
    sheet = wb[target_item.sheet_name]

    print("\n"+"="*80)
    print("Excel原始数据（该行）")
    print("="*80)

    for col_idx in range(1, 11):  # 读取前10列
        cell = sheet.cell(target_item.row, col_idx)
        value = cell.value
        if value:
            print(f"列{col_idx}: {value}")

    # 6. 分析缩进
    name_cell = sheet.cell(target_item.row, 2)  # B列是项目名称
    if hasattr(name_cell, 'alignment') and name_cell.alignment and name_cell.alignment.indent:
        indent = name_cell.alignment.indent
        print(f"\n缩进级别: {indent}")

    # 7. 测试新的级别识别逻辑
    print("\n"+"="*80)
    print("测试级别识别逻辑")
    print("="*80)

    # 获取同一工作表的所有项目
    sheet_items = [item for item in workbook_manager.target_items.values()
                   if item.sheet_name == target_item.sheet_name]
    sheet_items.sort(key=lambda x: x.row)

    print(f"\n工作表 '{target_item.sheet_name}' 的所有项目：")
    print(f"{'序号':<5} {'原始编号':<15} {'项目名称':<30} {'层级':<5} {'识别结果':<10}")
    print("-"*85)

    # 手动实现级别识别逻辑
    level_counters = {}

    for idx, item in enumerate(sheet_items, 1):
        level = item.hierarchical_level
        original_num = item.display_index

        # 重置更深层级的计数器
        levels_to_reset = [k for k in level_counters.keys() if k > level]
        for reset_level in levels_to_reset:
            del level_counters[reset_level]

        # 识别逻辑
        recognized_number = "?"

        if original_num and original_num.strip():
            original_clean = original_num.strip()

            # 检查是否是独立数字（如"3"、"3."）
            if original_clean.rstrip('.').isdigit():
                num_value = original_clean.rstrip('.')
                recognized_number = num_value
                level_counters[1] = int(num_value)
                # 重置更深层级
                for k in list(level_counters.keys()):
                    if k > 1:
                        del level_counters[k]

            # 检查是否包含点号（如"2.1"）
            elif '.' in original_clean:
                parts = original_clean.strip('.').split('.')
                if all(p.isdigit() for p in parts if p):
                    recognized_number = '.'.join(parts)
                    for part_idx, part in enumerate(parts, start=1):
                        level_counters[part_idx] = int(part)

            # 非数字编号
            elif level == 1:
                recognized_number = original_clean
                if 1 not in level_counters:
                    level_counters[1] = 0
                level_counters[1] += 1
        else:
            # 无编号，自动生成
            if level == 1:
                if 1 not in level_counters:
                    level_counters[1] = 0
                level_counters[1] += 1
                recognized_number = str(level_counters[1])
            else:
                parent_parts = []
                for i in range(1, level):
                    if i in level_counters:
                        parent_parts.append(str(level_counters[i]))
                    else:
                        parent_parts.append("1")

                if level not in level_counters:
                    level_counters[level] = 0
                level_counters[level] += 1
                number_parts = parent_parts + [str(level_counters[level])]
                recognized_number = ".".join(number_parts)

        # 标记目标项目
        marker = ">>>" if "税金及附加" in item.name else "   "

        print(f"{marker}{idx:<5} {original_num or '(无)':<15} {item.name[:28]:<30} {level:<5} {recognized_number:<10}")

    # 8. 结论
    print("\n"+"="*80)
    print("分析结论")
    print("="*80)

    # 找到测试结果中的目标项
    for item in sheet_items:
        if "税金及附加" in item.name:
            print(f"\n目标项目: {item.name}")
            print(f"原始编号: '{item.display_index}'")
            print(f"Excel中的层级: {item.hierarchical_level}")

            # 分析问题
            if item.display_index and item.display_index.strip():
                clean_num = item.display_index.strip().rstrip('.')
                if clean_num.isdigit():
                    print(f"原始编号'{item.display_index}'是纯数字，应识别为一级项目")
                    print(f"应该显示级别: {clean_num}")
                    if item.hierarchical_number != clean_num:
                        print(f"但实际显示: {item.hierarchical_number}")
                        print(f"问题原因: generate_hierarchical_numbers()未被正确调用或逻辑有误")
                    else:
                        print(f"实际显示: {item.hierarchical_number} (正确!)")
                else:
                    print(f"原始编号'{item.display_index}'不是纯数字")
            else:
                print(f"没有原始编号")

            break

if __name__ == "__main__":
    try:
        debug_level_recognition()
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
