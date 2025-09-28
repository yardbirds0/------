#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查Excel文件中的确切工作表名称
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

def list_exact_sheet_names():
    """列出Excel文件中的确切工作表名称"""

    # 查找Excel文件
    excel_file = None
    current_dir = os.path.dirname(os.path.dirname(__file__))

    for file in os.listdir(current_dir):
        if file.endswith('.xlsx') and '科电' in file:
            excel_file = os.path.join(current_dir, file)
            break

    if not excel_file:
        print("[ERROR] 未找到Excel文件")
        return

    print(f"[INFO] 正在检查Excel文件: {excel_file}")

    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        print(f"\n[SHEETS] 工作表列表 ({len(workbook.sheetnames)} 个):")
        for i, sheet_name in enumerate(workbook.sheetnames):
            print(f"  {i+1:2d}. '{sheet_name}' (长度: {len(sheet_name)})")
            # 显示每个字符的Unicode编码
            char_codes = [f"{ord(c):04x}" for c in sheet_name]
            print(f"      Unicode: {' '.join(char_codes)}")

        print(f"\n[TRIAL_BALANCE] 寻找科目余额表相关的工作表:")
        for sheet_name in workbook.sheetnames:
            if '科目余额' in sheet_name or '余额表' in sheet_name:
                print(f"  找到: '{sheet_name}'")

    except Exception as e:
        print(f"[ERROR] 读取Excel文件失败: {e}")

if __name__ == "__main__":
    list_exact_sheet_names()