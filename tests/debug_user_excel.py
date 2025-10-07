# -*- coding: utf-8 -*-
"""
专门测试: 上年科目余额表.xlsx 的列头显示
"""
import sys
sys.path.insert(0, 'd:\\Code\\快报填写程序-修改UI前(2)')

import openpyxl
from modules.table_schema_analyzer import TableSchemaAnalyzer

# 加载用户提供的Excel文件
workbook_path = 'd:\\Code\\快报填写程序-修改UI前(2)\\上年科目余额表.xlsx'
print(f"分析文件: {workbook_path}\n")

try:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    print(f"工作表列表:")
    for idx, sheet_name in enumerate(workbook.sheetnames):
        print(f"  {idx+1}. {sheet_name}")

    # 分析第一个工作表
    first_sheet = workbook.sheetnames[0]
    print(f"\n分析工作表: {first_sheet}")
    sheet = workbook[first_sheet]

    # 显示前5行原始数据
    print(f"\n原始数据(前5行):")
    for row in range(1, min(6, sheet.max_row + 1)):
        row_data = []
        for col in range(1, min(16, sheet.max_column + 1)):
            cell = sheet.cell(row, col)
            value = str(cell.value)[:20] if cell.value else ''
            col_letter = openpyxl.utils.get_column_letter(col)
            row_data.append(f"{col_letter}:{value}")
        print(f"  第{row}行: {' | '.join(row_data)}")

    # 创建分析器
    analyzer = TableSchemaAnalyzer()

    print(f"\n开始分析表格模式...")
    table_schema = analyzer.analyze_table_schema(sheet)

    print(f"\n分析结果:")
    print(f"  表格类型: {table_schema.table_type.value}")
    print(f"  表头起始行: {table_schema.header_start_row}")
    print(f"  列头行数: {table_schema.header_rows}")
    print(f"  数据开始行: {table_schema.data_start_row}")
    print(f"  名称列: {table_schema.name_columns}")
    print(f"  编码列: {table_schema.code_columns}")

    print(f"\n识别到的数据列 (共 {len(table_schema.data_columns)} 列):")
    for col_info in table_schema.data_columns:
        print(f"  列{col_info.column_letter}:")
        print(f"    主列头: '{col_info.primary_header}'")
        print(f"    次列头: '{col_info.secondary_header}'")
        print(f"    显示名称: '{col_info.display_name}'")
        print(f"    是否占位: {col_info.is_placeholder}")

        # 🔍 关键检查点
        if col_info.is_placeholder:
            print(f"    ⚠️ 警告: 此列被标记为占位列!")
        if not col_info.primary_header and not col_info.secondary_header:
            print(f"    ⚠️ 警告: 此列主次列头都为空!")

    if len(table_schema.data_columns) == 0:
        print("\n⚠️⚠️⚠️ 严重问题: 没有识别到任何数据列!")

    workbook.close()

except Exception as e:
    print(f"\n错误: {e}")
    import traceback
    traceback.print_exc()
