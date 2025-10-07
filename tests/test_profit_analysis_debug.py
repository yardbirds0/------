#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
专门调试"企业财务快报利润因素分析表"导致自动扩宽功能失效的问题
"""

import sys
import os
import json
import traceback
from pathlib import Path
from datetime import datetime

# 添加项目路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout,
    QWidget, QPushButton, QTextEdit, QLabel, QComboBox,
    QTableView, QHeaderView, QSplitter, QGroupBox,
    QMessageBox, QTabWidget
)
from PySide6.QtCore import Qt, QTimer, Signal
from PySide6.QtGui import QStandardItemModel, QStandardItem

from models.data_models import WorkbookManager, TargetItem
from modules.file_manager import FileManager
from modules.data_extractor import DataExtractor
from components.advanced_widgets import (
    ensure_interactive_header,
    schedule_row_resize,
    derive_header_layout_from_metadata
)


class DebugLogger:
    """调试日志记录器"""
    def __init__(self, text_widget):
        self.text_widget = text_widget
        self.logs = []

    def log(self, message, level="INFO"):
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        log_entry = f"[{timestamp}] {level}: {message}"
        self.logs.append(log_entry)

        # 根据级别设置颜色
        if level == "ERROR":
            html = f'<span style="color: red; font-weight: bold;">{log_entry}</span>'
        elif level == "WARNING":
            html = f'<span style="color: orange;">{log_entry}</span>'
        elif level == "SUCCESS":
            html = f'<span style="color: green; font-weight: bold;">{log_entry}</span>'
        elif level == "DEBUG":
            html = f'<span style="color: gray;">{log_entry}</span>'
        else:
            html = log_entry

        self.text_widget.append(html)

    def info(self, message):
        self.log(message, "INFO")

    def debug(self, message):
        self.log(message, "DEBUG")

    def warning(self, message):
        self.log(message, "WARNING")

    def error(self, message):
        self.log(message, "ERROR")

    def success(self, message):
        self.log(message, "SUCCESS")


class ProfitAnalysisDebugWindow(QMainWindow):
    """专门调试利润因素分析表的窗口"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("调试：企业财务快报利润因素分析表")
        self.resize(1600, 900)

        # 初始化组件
        self.workbook_manager = WorkbookManager()
        self.file_manager = FileManager()
        self.data_extractor = DataExtractor(self.workbook_manager)

        # 目标表格名称
        self.target_sheet_name = "企业财务快报利润因素分析表"

        # 数据存储
        self.excel_file = None
        self.workbook = None
        self.sheet_data = {}

        # 设置界面
        self.setup_ui()

        # 初始化日志
        self.logger = DebugLogger(self.log_text)
        self.logger.info("程序启动，准备调试")

    def setup_ui(self):
        """设置界面"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 控制面板
        control_panel = self.create_control_panel()
        main_layout.addWidget(control_panel)

        # 创建标签页
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # 数据显示标签页
        data_tab = self.create_data_tab()
        self.tabs.addTab(data_tab, "数据分析")

        # 模型测试标签页
        model_tab = self.create_model_tab()
        self.tabs.addTab(model_tab, "模型测试")

        # 日志标签页
        log_tab = self.create_log_tab()
        self.tabs.addTab(log_tab, "调试日志")

    def create_control_panel(self):
        """创建控制面板"""
        panel = QGroupBox("控制面板")
        layout = QHBoxLayout()
        panel.setLayout(layout)

        # 加载文件按钮
        self.load_btn = QPushButton("加载Excel文件")
        self.load_btn.clicked.connect(self.load_excel)
        layout.addWidget(self.load_btn)

        # 分析按钮
        self.analyze_btn = QPushButton("分析表格结构")
        self.analyze_btn.clicked.connect(self.analyze_sheet)
        self.analyze_btn.setEnabled(False)
        layout.addWidget(self.analyze_btn)

        # 提取数据按钮
        self.extract_btn = QPushButton("提取目标项")
        self.extract_btn.clicked.connect(self.extract_data)
        self.extract_btn.setEnabled(False)
        layout.addWidget(self.extract_btn)

        # 测试模型按钮
        self.test_model_btn = QPushButton("测试数据模型")
        self.test_model_btn.clicked.connect(self.test_model)
        self.test_model_btn.setEnabled(False)
        layout.addWidget(self.test_model_btn)

        # 模拟切换按钮
        self.simulate_btn = QPushButton("模拟表格切换")
        self.simulate_btn.clicked.connect(self.simulate_sheet_switch)
        self.simulate_btn.setEnabled(False)
        layout.addWidget(self.simulate_btn)

        # 导出日志按钮
        self.export_btn = QPushButton("导出调试日志")
        self.export_btn.clicked.connect(self.export_logs)
        layout.addWidget(self.export_btn)

        layout.addStretch()
        return panel

    def create_data_tab(self):
        """创建数据分析标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # 原始数据显示
        raw_group = QGroupBox("原始Excel数据")
        raw_layout = QVBoxLayout()
        raw_group.setLayout(raw_layout)

        self.raw_data_text = QTextEdit()
        self.raw_data_text.setReadOnly(True)
        raw_layout.addWidget(self.raw_data_text)

        layout.addWidget(raw_group)

        # 元数据显示
        meta_group = QGroupBox("列元数据分析")
        meta_layout = QVBoxLayout()
        meta_group.setLayout(meta_layout)

        self.meta_data_text = QTextEdit()
        self.meta_data_text.setReadOnly(True)
        meta_layout.addWidget(self.meta_data_text)

        layout.addWidget(meta_group)

        return widget

    def create_model_tab(self):
        """创建模型测试标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # 模型信息
        info_group = QGroupBox("模型信息")
        info_layout = QVBoxLayout()
        info_group.setLayout(info_layout)

        self.model_info_text = QTextEdit()
        self.model_info_text.setReadOnly(True)
        info_layout.addWidget(self.model_info_text)

        layout.addWidget(info_group)

        # 表格预览
        preview_group = QGroupBox("表格预览")
        preview_layout = QVBoxLayout()
        preview_group.setLayout(preview_layout)

        self.preview_table = QTableView()
        preview_layout.addWidget(self.preview_table)

        layout.addWidget(preview_group)

        return widget

    def create_log_tab(self):
        """创建日志标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)

        return widget

    def load_excel(self):
        """加载Excel文件"""
        try:
            # 查找Excel文件
            excel_path = None
            for file in Path(r"d:\Code\快报填写程序").glob("*.xls*"):
                if "快报" in file.name and not file.name.startswith("~"):
                    excel_path = file
                    break

            if not excel_path:
                self.logger.error("未找到包含'快报'的Excel文件")
                return

            self.logger.info(f"找到Excel文件: {excel_path.name}")
            self.excel_file = excel_path

            # 加载文件
            from openpyxl import load_workbook
            self.logger.info("正在加载工作簿...")
            self.workbook = load_workbook(str(excel_path), read_only=False, data_only=True)

            # 检查目标表格是否存在
            if self.target_sheet_name not in self.workbook.sheetnames:
                self.logger.error(f"未找到表格: {self.target_sheet_name}")
                self.logger.info(f"可用的表格: {', '.join(self.workbook.sheetnames)}")
                return

            self.logger.success(f"成功加载表格: {self.target_sheet_name}")

            # 启用其他按钮
            self.analyze_btn.setEnabled(True)
            self.extract_btn.setEnabled(True)
            self.test_model_btn.setEnabled(True)
            self.simulate_btn.setEnabled(True)

            # 自动分析
            self.analyze_sheet()

        except Exception as e:
            self.logger.error(f"加载Excel文件失败: {e}")
            self.logger.debug(traceback.format_exc())

    def analyze_sheet(self):
        """分析表格结构"""
        try:
            if not self.workbook:
                self.logger.error("请先加载Excel文件")
                return

            sheet = self.workbook[self.target_sheet_name]
            self.logger.info(f"开始分析表格: {self.target_sheet_name}")

            # 基本信息
            max_row = sheet.max_row
            max_col = sheet.max_column
            self.logger.info(f"表格大小: {max_row} 行 x {max_col} 列")

            # 显示前10行数据
            raw_data = []
            raw_data.append(f"表格: {self.target_sheet_name}")
            raw_data.append(f"大小: {max_row} 行 x {max_col} 列")
            raw_data.append("-" * 80)

            for row_idx in range(1, min(11, max_row + 1)):
                row_data = []
                for col_idx in range(1, min(6, max_col + 1)):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is None:
                        value = "[空]"
                    row_data.append(str(value))
                raw_data.append(f"第{row_idx}行: {' | '.join(row_data)}")

            self.raw_data_text.setPlainText("\n".join(raw_data))

            # 分析列结构
            self.logger.info("分析列结构...")
            headers = []
            for col_idx in range(1, max_col + 1):
                # 查找第一个非空单元格作为列头
                header_value = None
                for row_idx in range(1, min(5, max_row + 1)):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        header_value = str(cell.value)
                        break
                headers.append(header_value or f"列{col_idx}")

            self.logger.info(f"检测到的列头: {headers}")

            # 保存数据
            self.sheet_data = {
                "max_row": max_row,
                "max_col": max_col,
                "headers": headers,
                "sheet": sheet
            }

            # 使用FileManager加载
            self.logger.info("使用FileManager加载文件...")
            success = self.file_manager.load_file(str(self.excel_file))
            if success:
                self.logger.success("FileManager加载成功")
                self.workbook_manager = self.file_manager.workbook_manager
                self.data_extractor = DataExtractor(self.workbook_manager)
            else:
                self.logger.error("FileManager加载失败")

        except Exception as e:
            self.logger.error(f"分析表格失败: {e}")
            self.logger.debug(traceback.format_exc())

    def extract_data(self):
        """提取数据"""
        try:
            if not self.workbook_manager:
                self.logger.error("WorkbookManager未初始化")
                return

            self.logger.info("开始提取数据...")

            # 设置表格类型
            for wb_name in self.workbook_manager.workbooks:
                for sheet_name in self.workbook_manager.workbooks[wb_name].sheetnames:
                    if "快报" in sheet_name:
                        self.workbook_manager.set_sheet_type(wb_name, sheet_name, "flash_report")
                        self.logger.debug(f"设置为快报表: {sheet_name}")
                    else:
                        self.workbook_manager.set_sheet_type(wb_name, sheet_name, "data_source")

            # 提取所有数据
            self.logger.info("调用extract_all_data...")
            self.data_extractor.extract_all_data()

            # 检查目标表的元数据
            if self.target_sheet_name in self.workbook_manager.target_sheet_columns:
                columns = self.workbook_manager.target_sheet_columns[self.target_sheet_name]
                self.logger.success(f"成功提取 {len(columns)} 列元数据")

                # 显示元数据
                meta_info = []
                meta_info.append(f"表格: {self.target_sheet_name}")
                meta_info.append(f"列数: {len(columns)}")
                meta_info.append("-" * 80)

                for i, col in enumerate(columns):
                    meta_info.append(f"列 {i}:")
                    meta_info.append(f"  key: {col.get('key', 'N/A')}")
                    meta_info.append(f"  display_name: {col.get('display_name', 'N/A')}")
                    meta_info.append(f"  is_data_column: {col.get('is_data_column', False)}")
                    meta_info.append(f"  column_index: {col.get('column_index', 'N/A')}")
                    meta_info.append("")

                self.meta_data_text.setPlainText("\n".join(meta_info))
            else:
                self.logger.warning(f"未找到'{self.target_sheet_name}'的列元数据")
                self.logger.info(f"可用的目标表: {list(self.workbook_manager.target_sheet_columns.keys())}")

            # 检查目标项
            target_count = 0
            for target_id, target in self.workbook_manager.target_items.items():
                if target.sheet_name == self.target_sheet_name:
                    target_count += 1

            self.logger.info(f"找到 {target_count} 个目标项")

        except Exception as e:
            self.logger.error(f"提取数据失败: {e}")
            self.logger.debug(traceback.format_exc())

    def test_model(self):
        """测试数据模型"""
        try:
            from main import TargetItemModel

            self.logger.info("创建TargetItemModel...")
            model = TargetItemModel(self.workbook_manager)

            # 设置活动表格
            model.set_active_sheet(self.target_sheet_name)

            # 检查模型信息
            col_count = model.columnCount()
            row_count = model.rowCount()

            self.logger.info(f"模型列数: {col_count}")
            self.logger.info(f"模型行数: {row_count}")

            # 检查headers
            if hasattr(model, "headers"):
                self.logger.info(f"headers: {model.headers}")
                self.logger.info(f"headers长度: {len(model.headers)}")
            else:
                self.logger.error("模型缺少headers属性")

            # 显示模型信息
            info = []
            info.append(f"模型类型: TargetItemModel")
            info.append(f"活动表格: {self.target_sheet_name}")
            info.append(f"列数: {col_count}")
            info.append(f"行数: {row_count}")

            if hasattr(model, "headers"):
                info.append(f"\nHeaders ({len(model.headers)}):")
                for i, header in enumerate(model.headers):
                    info.append(f"  [{i}] {header}")

            if hasattr(model, "static_headers"):
                info.append(f"\n静态列: {model.static_headers}")

            if hasattr(model, "dynamic_columns"):
                info.append(f"\n动态列数: {len(model.dynamic_columns)}")
                for col in model.dynamic_columns:
                    info.append(f"  - {col.get('name', 'N/A')}")

            self.model_info_text.setPlainText("\n".join(info))

            # 设置到表格视图
            self.preview_table.setModel(model)

            # 测试列宽调整
            self.test_column_resize(model)

        except Exception as e:
            self.logger.error(f"测试模型失败: {e}")
            self.logger.debug(traceback.format_exc())

    def test_column_resize(self, model):
        """测试列宽调整逻辑"""
        try:
            self.logger.info("测试列宽调整逻辑...")

            header = self.preview_table.horizontalHeader()

            # 模拟adjust_main_table_columns的逻辑
            for column in range(model.columnCount()):
                # 安全获取列名
                column_name = ""
                try:
                    if hasattr(model, "headers") and column < len(model.headers):
                        column_name = model.headers[column] or ""
                        self.logger.debug(f"列{column}: {column_name}")
                    else:
                        self.logger.warning(f"列{column}超出headers范围")
                except Exception as e:
                    self.logger.error(f"获取列{column}名称失败: {e}")

            self.logger.success("列宽调整测试完成")

        except Exception as e:
            self.logger.error(f"列宽调整测试失败: {e}")

    def simulate_sheet_switch(self):
        """模拟表格切换"""
        try:
            self.logger.info("=" * 80)
            self.logger.info("模拟表格切换测试")
            self.logger.info("=" * 80)

            # 获取所有快报表
            flash_sheets = []
            for sheet_id, sheet_type in self.workbook_manager.sheet_types.items():
                if sheet_type == "flash_report":
                    _, sheet_name = sheet_id.rsplit("::", 1)
                    flash_sheets.append(sheet_name)

            self.logger.info(f"找到 {len(flash_sheets)} 个快报表")

            from main import TargetItemModel

            # 测试切换到每个表格
            for sheet_name in flash_sheets[:3]:  # 只测试前3个
                self.logger.info(f"\n切换到: {sheet_name}")

                try:
                    model = TargetItemModel(self.workbook_manager)
                    model.set_active_sheet(sheet_name)

                    # 检查headers
                    if hasattr(model, "headers"):
                        self.logger.success(f"  headers正常: {len(model.headers)} 列")
                    else:
                        self.logger.error(f"  缺少headers属性")

                except Exception as e:
                    self.logger.error(f"  切换失败: {e}")

            # 最后切换回目标表格
            self.logger.info(f"\n切换回: {self.target_sheet_name}")
            model = TargetItemModel(self.workbook_manager)
            model.set_active_sheet(self.target_sheet_name)

            if hasattr(model, "headers"):
                self.logger.success(f"最终headers: {model.headers}")
            else:
                self.logger.error("最终缺少headers")

        except Exception as e:
            self.logger.error(f"模拟切换失败: {e}")
            self.logger.debug(traceback.format_exc())

    def export_logs(self):
        """导出日志"""
        try:
            log_file = Path("profit_analysis_debug_log.txt")
            with open(log_file, "w", encoding="utf-8") as f:
                f.write("企业财务快报利润因素分析表 调试日志\n")
                f.write("=" * 80 + "\n")
                f.write(f"生成时间: {datetime.now()}\n")
                f.write("=" * 80 + "\n\n")

                for log in self.logger.logs:
                    f.write(log + "\n")

            self.logger.success(f"日志已导出到: {log_file}")
            QMessageBox.information(self, "导出成功", f"日志已保存到:\n{log_file}")

        except Exception as e:
            self.logger.error(f"导出日志失败: {e}")


def main():
    app = QApplication(sys.argv)
    window = ProfitAnalysisDebugWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()