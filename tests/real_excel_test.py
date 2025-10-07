#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
真实Excel数据测试 - 绝不捏造数据
直接从Excel读取真实数据，测试三段式公式系统
"""

import sys
import os
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.data_models import SourceItem, WorkbookManager
from utils.excel_utils import (
    parse_formula_references_three_segment,
    validate_formula_syntax_three_segment,
    evaluate_formula_with_values_three_segment,
    build_formula_reference_three_segment
)

def main():
    print("=" * 100)
    print("真实Excel数据测试 - 绝不捏造数据")
    print("=" * 100)

    # 加载真实Excel
    excel_path = r"D:\Code\快报填写程序\（科电）国资委、财政快报模板-纯净版 的副本.xlsx"

    print(f"\n[1] 加载Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"   包含 {len(wb.sheetnames)} 个工作表:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"     {i}. {name}")

    # 手动从每个sheet提取真实数据
    print(f"\n[2] 提取每个sheet的真实数据（每个sheet前10行）:")

    all_source_items = {}
    workbook_manager = WorkbookManager()

    for sheet_name in wb.sheetnames:
        print(f"\n  === Sheet: {sheet_name} ===")
        sheet = wb[sheet_name]

        workbook_manager.worksheets[sheet_name] = sheet

        # 读取前10行数据
        items_count = 0
        for row_idx in range(1, min(20, sheet.max_row + 1)):  # 前20行
            # 尝试从B列读取项目名（常见格式）
            name_cell = sheet.cell(row=row_idx, column=2)  # B列
            name = name_cell.value

            if name and isinstance(name, str) and name.strip():
                name = name.strip()

                # 读取同行的所有数据列（C到J列）
                row_data = {}
                for col_idx in range(3, min(11, sheet.max_column + 1)):  # C到J列
                    header = sheet.cell(row=1, column=col_idx).value  # 第1行作为表头
                    value = sheet.cell(row=row_idx, column=col_idx).value

                    if header and isinstance(header, str):
                        header = header.strip()
                        if header and value is not None:
                            row_data[header] = value

                # 创建SourceItem
                item_id = f"{sheet_name}_{row_idx}"
                source = SourceItem(
                    id=item_id,
                    sheet_name=sheet_name,
                    name=name,
                    cell_address=f"B{row_idx}",
                    row=row_idx,
                    column="B",
                    value=list(row_data.values())[0] if row_data else None
                )
                source.values = row_data

                all_source_items[item_id] = source
                workbook_manager.source_items[item_id] = source

                # 显示前10个
                if items_count < 10:
                    print(f"    [{items_count+1}] 项目: {name}")
                    print(f"        数据列: {list(row_data.keys())}")
                    print(f"        数据值: {row_data}")
                    items_count += 1

                if items_count >= 10:
                    break

    print(f"\n[3] 数据提取完成")
    print(f"   总计提取: {len(all_source_items)} 个数据项")
    print(f"   工作表数: {len(workbook_manager.worksheets)}")

    # 测试公式
    print(f"\n[4] 列出所有有数字值的数据项（用于调试）")

    valid_items = []
    for item in all_source_items.values():
        if item.values:
            for col_name, col_value in item.values.items():
                if isinstance(col_value, (int, float)):
                    valid_items.append({
                        'sheet': item.sheet_name,
                        'name': item.name,
                        'column': col_name,
                        'value': col_value
                    })
                    if len(valid_items) <= 5:  # 显示前5个
                        print(f"   [{len(valid_items)}] [{item.sheet_name}]![{item.name}]![{col_name}] = {col_value}")

    print(f"\n[5] 调试工作表名称")
    print(f"   workbook_manager.worksheets的keys:")
    for i, key in enumerate(workbook_manager.worksheets.keys(), 1):
        print(f"     {i}. '{key}'")

    print(f"\n[6] 测试真实公式")

    # 用第一个有效项测试
    first_item = valid_items[0]
    formula = build_formula_reference_three_segment(
        first_item['sheet'],
        first_item['name'],
        first_item['column']
    )

    test_formulas = [{
        'formula': formula,
        'expected_value': first_item['value']
    }]

    print(f"   使用真实数据:")
    print(f"     公式: {formula}")
    print(f"     期望值: {first_item['value']}")

    if not test_formulas:
        print("   [FAIL] 未能构建任何测试公式")
        return 1

    print(f"\n   构建了 {len(test_formulas)} 个真实公式:")

    for i, test in enumerate(test_formulas, 1):
        print(f"\n   测试 {i}:")
        print(f"     公式: {test['formula']}")
        print(f"     期望值: {test['expected_value']}")

        # 解析并显示详细信息
        refs = parse_formula_references_three_segment(test['formula'])
        print(f"     解析: {len(refs)}个引用", end="")
        if len(refs) > 0:
            print(" [OK]")
            for ref in refs:
                print(f"       - sheet_name: '{ref['sheet_name']}'")
                print(f"       - item_name: '{ref['item_name']}'")
                print(f"       - column_name: '{ref['column_name']}'")
        else:
            print(" [FAIL]")
            continue

        # 验证
        is_valid, error = validate_formula_syntax_three_segment(test['formula'], workbook_manager)
        print(f"     验证: ", end="")
        if is_valid:
            print("[OK]")
        else:
            print(f"[FAIL] {error}")
            # 调试：检查工作表是否存在
            ref = refs[0]
            print(f"       调试: '{ref['sheet_name']}' in worksheets? {ref['sheet_name'] in workbook_manager.worksheets}")
            continue

        # 计算
        success, result = evaluate_formula_with_values_three_segment(test['formula'], all_source_items)
        print(f"     计算: ", end="")
        if success:
            print(f"[OK] 结果={result}")
            if result == test['expected_value']:
                print(f"     匹配: [OK] 结果与期望值一致")
            else:
                print(f"     匹配: [WARNING] 结果{result} != 期望{test['expected_value']}")
        else:
            print(f"[FAIL] {result}")

    print("\n" + "=" * 100)
    print("测试完成 - 所有数据来自真实Excel，绝无捏造")
    print("=" * 100)

    return 0

if __name__ == "__main__":
    sys.exit(main())
