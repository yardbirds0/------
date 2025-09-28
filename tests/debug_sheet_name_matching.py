#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
调试工作表名称匹配问题
检查为什么data_extractor找不到科目余额表
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from modules.data_extractor import DataExtractor
from models.data_models import WorkbookManager

class SheetNameMatchingDebugger:
    """工作表名称匹配调试器"""

    def __init__(self, excel_file_path: str):
        self.excel_file_path = excel_file_path

    def debug_sheet_name_matching(self):
        """调试工作表名称匹配问题"""
        print(f"[DEBUG] 开始调试工作表名称匹配问题...")
        print(f"[DEBUG] Excel文件: {self.excel_file_path}")

        # 1. 直接用openpyxl读取工作表名称
        print(f"\n[OPENPYXL] 直接读取的工作表名称:")
        workbook = openpyxl.load_workbook(self.excel_file_path, data_only=True)
        for i, sheet_name in enumerate(workbook.sheetnames):
            print(f"  {i+1:2d}. '{sheet_name}' (类型: {type(sheet_name)}, 长度: {len(sheet_name)})")
            char_codes = [f"U+{ord(c):04X}" for c in sheet_name]
            print(f"      字符编码: {' '.join(char_codes)}")

        # 2. 从WorkbookManager获取的工作表名称
        print(f"\n[WORKBOOK_MANAGER] WorkbookManager中的data_source_sheets:")
        workbook_manager = WorkbookManager(self.excel_file_path)

        # 模拟FileManager的加载过程
        from modules.file_manager import FileManager
        file_manager = FileManager()
        success, message = file_manager.load_excel_files([self.excel_file_path])

        if success:
            workbook_manager = file_manager.workbook_manager
            for i, sheet_name in enumerate(workbook_manager.data_source_sheets):
                print(f"  {i+1:2d}. '{sheet_name}' (类型: {type(sheet_name)}, 长度: {len(sheet_name)})")
                char_codes = [f"U+{ord(c):04X}" for c in sheet_name]
                print(f"      字符编码: {' '.join(char_codes)}")

        # 3. 检查data_extractor的工作表名称处理
        print(f"\n[EXTRACTOR] DataExtractor的工作表处理:")
        extractor = DataExtractor(workbook_manager)

        # 检查extractor加载的workbook
        if extractor._load_workbook():
            print(f"  Extractor工作簿加载成功")
            print(f"  Extractor中的工作表名称:")
            for i, sheet_name in enumerate(extractor.workbook.sheetnames):
                print(f"    {i+1:2d}. '{sheet_name}' (类型: {type(sheet_name)}, 长度: {len(sheet_name)})")
                char_codes = [f"U+{ord(c):04X}" for c in sheet_name]
                print(f"        字符编码: {' '.join(char_codes)}")

        # 4. 逐一检查每个data_source_sheet是否在extractor的workbook中
        print(f"\n[MATCHING] 检查data_source_sheets与workbook的匹配:")
        for sheet_name in workbook_manager.data_source_sheets:
            exists = sheet_name in extractor.workbook.sheetnames
            print(f"  '{sheet_name}' -> 存在: {exists}")

            if not exists:
                print(f"    [ERROR] 工作表不存在!")
                print(f"    可能的匹配项:")
                for wb_sheet in extractor.workbook.sheetnames:
                    similarity = self._calculate_similarity(sheet_name, wb_sheet)
                    if similarity > 0.8:
                        print(f"      '{wb_sheet}' (相似度: {similarity:.2f})")

        # 5. 特别检查科目余额表
        print(f"\n[TRIAL_BALANCE] 科目余额表特别检查:")
        trial_balance_sheets = [s for s in workbook_manager.data_source_sheets if '科目余额' in s]
        for sheet_name in trial_balance_sheets:
            print(f"  检查: '{sheet_name}'")
            exists = sheet_name in extractor.workbook.sheetnames
            print(f"    在workbook中存在: {exists}")

            if not exists:
                # 尝试找到相似的工作表
                print(f"    寻找相似的工作表:")
                for wb_sheet in extractor.workbook.sheetnames:
                    if '科目余额' in wb_sheet or '余额表' in wb_sheet:
                        print(f"      候选: '{wb_sheet}'")
                        print(f"        字符编码差异:")
                        self._compare_char_encoding(sheet_name, wb_sheet)

    def _calculate_similarity(self, str1: str, str2: str) -> float:
        """计算字符串相似度"""
        if str1 == str2:
            return 1.0

        # 简单的字符匹配相似度
        common_chars = sum(1 for c1, c2 in zip(str1, str2) if c1 == c2)
        max_len = max(len(str1), len(str2))
        return common_chars / max_len if max_len > 0 else 0.0

    def _compare_char_encoding(self, str1: str, str2: str):
        """比较两个字符串的字符编码差异"""
        print(f"          '{str1}' vs '{str2}'")
        max_len = max(len(str1), len(str2))
        for i in range(max_len):
            c1 = str1[i] if i < len(str1) else '(无)'
            c2 = str2[i] if i < len(str2) else '(无)'
            code1 = f"U+{ord(c1):04X}" if c1 != '(无)' else '(无)'
            code2 = f"U+{ord(c2):04X}" if c2 != '(无)' else '(无)'
            match = "✓" if c1 == c2 else "✗"
            print(f"          {i:2d}: '{c1}' ({code1}) vs '{c2}' ({code2}) {match}")

def main():
    """主函数"""
    # 查找Excel文件
    excel_file = None
    current_dir = os.path.dirname(os.path.dirname(__file__))

    for file in os.listdir(current_dir):
        if file.endswith('.xlsx') and '科电' in file:
            excel_file = os.path.join(current_dir, file)
            break

    if not excel_file:
        print("[ERROR] 未找到Excel文件")
        return

    # 创建调试器并运行
    debugger = SheetNameMatchingDebugger(excel_file)
    debugger.debug_sheet_name_matching()

    print(f"\n[DONE] 工作表名称匹配调试完成!")

if __name__ == "__main__":
    main()