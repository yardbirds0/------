#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel列头分析脚本 - 检查利润表的实际列头结构
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

def analyze_excel_headers():
    """分析Excel表格的列头结构"""
    print("=== Excel列头分析 ===")

    # 1. 找到Excel文件
    current_dir = os.path.dirname(os.path.dirname(__file__))
    excel_file = None
    for file in os.listdir(current_dir):
        if file.endswith('.xlsx') and '科电' in file:
            excel_file = os.path.join(current_dir, file)
            break

    if not excel_file:
        print("[错误] 未找到Excel文件")
        return

    print(f"Excel文件: {excel_file}")

    # 2. 直接用openpyxl打开文件
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        print(f"工作表列表: {workbook.sheetnames}")

        # 3. 重点分析利润表
        if "利润表" in workbook.sheetnames:
            print(f"\n=== 利润表详细分析 ===")
            sheet = workbook["利润表"]

            print(f"表格尺寸: {sheet.max_row} 行 x {sheet.max_column} 列")

            # 分析前几行的内容，找到列头
            print(f"\n前5行内容:")
            for row in range(1, min(6, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(8, sheet.max_column + 1)):  # 只看前8列
                    cell = sheet.cell(row=row, column=col)
                    value = cell.value if cell.value is not None else ""
                    row_data.append(str(value)[:15])  # 截断长文本
                print(f"  行{row}: {row_data}")

            # 分析列头（假设在第1-3行）
            print(f"\n列头分析:")
            for header_row in range(1, 4):
                print(f"  第{header_row}行作为列头:")
                headers = []
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=header_row, column=col)
                    if cell.value:
                        headers.append((col, str(cell.value)))

                for col, header in headers:
                    print(f"    列{col}: '{header}'")

            # 分析数据行（从第4行开始）
            print(f"\n数据行示例 (第4-6行):")
            for row in range(4, min(7, sheet.max_row + 1)):
                print(f"  行{row}:")
                for col in range(1, min(8, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        print(f"    列{col}: {cell.value}")

        else:
            print(f"\n[警告] 没有找到'利润表'工作表")

    except Exception as e:
        print(f"[错误] 分析Excel文件失败: {e}")

if __name__ == "__main__":
    analyze_excel_headers()