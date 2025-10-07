#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
全面测试 - 5个来源表 x 100个单元格 = 500个测试案例
确保三段式公式系统100%可靠
"""

import sys
import os
import openpyxl
from collections import defaultdict
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.data_models import SourceItem, WorkbookManager
from utils.excel_utils import (
    parse_formula_references_three_segment,
    validate_formula_syntax_three_segment,
    evaluate_formula_with_values_three_segment,
    build_formula_reference_three_segment
)

def main():
    print("=" * 100)
    print("全面测试 - 5个来源表 x 100个单元格 = 500个测试")
    print("=" * 100)

    excel_path = r"D:\Code\快报填写程序\（科电）国资委、财政快报模板-纯净版 的副本.xlsx"

    print(f"\n[1] 加载Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 识别数据来源表（排除快报表）
    print(f"\n[2] 识别数据来源表（排除快报表）:")
    source_sheets = []
    for sheet_name in wb.sheetnames:
        # 快报表通常包含"快报"关键字，排除它们
        if "快报" not in sheet_name:
            source_sheets.append(sheet_name)

    print(f"  找到 {len(source_sheets)} 个数据来源表:")
    for i, name in enumerate(source_sheets, 1):
        print(f"    {i}. {name}")

    # 选择前6个数据来源表（补足500个测试）
    selected_sheets = source_sheets[:6]
    print(f"\n[3] 选择前6个数据来源表进行测试（目标500个）:")
    for i, name in enumerate(selected_sheets, 1):
        print(f"    {i}. {name}")

    # 从每个表提取数据单元格
    print(f"\n[4] 从每个表提取数据单元格（目标500个）:")

    all_test_cases = []  # 所有测试案例
    wm = WorkbookManager()
    target_total = 500  # 目标总数

    for sheet_name in selected_sheets:
        print(f"\n  === {sheet_name} ===")
        sheet = wb[sheet_name]
        wm.worksheets[sheet_name] = sheet

        # 查找所有有数值的单元格
        data_cells = []

        # 假设第1行是表头，从第2行开始
        # 假设B列是项目名
        for row in range(2, min(sheet.max_row + 1, 2000)):  # 扩大扫描范围到2000行
            item_name_cell = sheet.cell(row=row, column=2)  # B列
            item_name = item_name_cell.value

            if not item_name or not isinstance(item_name, str) or not item_name.strip():
                continue

            # 扫描该行的所有数据列（从C列开始）
            for col in range(3, min(sheet.max_column + 1, 100)):  # 扩大扫描范围到100列
                # 获取表头（第1行）
                header_cell = sheet.cell(row=1, column=col)
                column_name = header_cell.value

                # 获取数据单元格
                data_cell = sheet.cell(row=row, column=col)
                data_value = data_cell.value

                # 只收集数值单元格
                if isinstance(data_value, (int, float)) and column_name and isinstance(column_name, str):
                    data_cells.append({
                        'sheet_name': sheet_name,
                        'item_name': item_name,
                        'column_name': column_name,
                        'cell_address': data_cell.coordinate,
                        'value': data_value,
                        'row': row,
                        'col': col
                    })

        print(f"    找到 {len(data_cells)} 个数据单元格")

        # 创建SourceItem
        created_items = {}
        for cell_info in data_cells:
            item_key = f"{sheet_name}_{cell_info['item_name']}"

            if item_key not in created_items:
                # 创建新的SourceItem
                source = SourceItem(
                    id=item_key,
                    sheet_name=sheet_name,
                    name=cell_info['item_name'],  # 保留原始值
                    cell_address=f"B{cell_info['row']}",
                    row=cell_info['row'],
                    column="B",
                    value=None
                )
                source.values = {}
                created_items[item_key] = source
                wm.source_items[item_key] = source

            # 添加列数据
            created_items[item_key].values[cell_info['column_name']] = cell_info['value']

        # 将所有数据单元格加入测试案例
        all_test_cases.extend(data_cells)

    # 如果测试案例不足500个，从现有的重复采样
    if len(all_test_cases) < target_total:
        print(f"\n  [注意] 实际只找到 {len(all_test_cases)} 个数据单元格")
        print(f"         已全部加入测试")
    else:
        # 如果超过500个，只取前500个
        all_test_cases = all_test_cases[:target_total]
        print(f"\n  [注意] 找到 {len(all_test_cases)} 个数据单元格，取前500个进行测试")

    print(f"\n[5] 测试案例统计:")
    print(f"  总计: {len(all_test_cases)} 个测试案例")

    # 按工作表分组统计
    by_sheet = defaultdict(int)
    for case in all_test_cases:
        by_sheet[case['sheet_name']] += 1

    for sheet_name, count in by_sheet.items():
        print(f"    {sheet_name}: {count} 个")

    # 开始测试
    print(f"\n[6] 开始全面测试（解析 → 验证 → 计算）:")
    print(f"  进度: ", end="", flush=True)

    passed = 0
    failed = 0
    failed_cases = []

    for i, case in enumerate(all_test_cases):
        # 构建公式
        formula = build_formula_reference_three_segment(
            case['sheet_name'],
            case['item_name'],
            case['column_name']
        )

        # 解析
        refs = parse_formula_references_three_segment(formula)
        if not refs:
            failed += 1
            failed_cases.append({
                'case': case,
                'formula': formula,
                'stage': 'parse',
                'error': '解析失败，引用数为0'
            })
            print("x", end="", flush=True)
            continue

        # 验证
        is_valid, error = validate_formula_syntax_three_segment(formula, wm)
        if not is_valid:
            failed += 1
            failed_cases.append({
                'case': case,
                'formula': formula,
                'stage': 'validate',
                'error': error
            })
            print("x", end="", flush=True)
            continue

        # 计算
        success, result = evaluate_formula_with_values_three_segment(formula, wm.source_items)
        if not success:
            failed += 1
            failed_cases.append({
                'case': case,
                'formula': formula,
                'stage': 'calculate',
                'error': result
            })
            print("x", end="", flush=True)
            continue

        # 验证结果
        if result != case['value']:
            failed += 1
            failed_cases.append({
                'case': case,
                'formula': formula,
                'stage': 'verify',
                'error': f'结果不匹配: 计算值{result} != 期望值{case["value"]}'
            })
            print("x", end="", flush=True)
            continue

        # 通过
        passed += 1
        print(".", end="", flush=True)

        # 每50个测试显示一次进度
        if (i + 1) % 50 == 0:
            print(f" {i+1}/{len(all_test_cases)}", end="", flush=True)

    print()  # 换行

    # 测试结果
    print(f"\n[7] 测试结果:")
    print(f"  通过: {passed}/{len(all_test_cases)} ({passed*100//len(all_test_cases) if all_test_cases else 0}%)")
    print(f"  失败: {failed}/{len(all_test_cases)} ({failed*100//len(all_test_cases) if all_test_cases else 0}%)")

    if failed_cases:
        print(f"\n[8] 失败案例详情（前20个）:")
        for i, fail in enumerate(failed_cases[:20], 1):
            print(f"\n  失败 {i}:")
            print(f"    工作表: {fail['case']['sheet_name']}")
            print(f"    单元格: {fail['case']['cell_address']}")
            print(f"    项目名: '{fail['case']['item_name']}'")
            print(f"    列名: '{fail['case']['column_name']}'")
            print(f"    公式: {fail['formula']}")
            print(f"    失败阶段: {fail['stage']}")
            print(f"    错误: {fail['error']}")

        if len(failed_cases) > 20:
            print(f"\n  ... 还有 {len(failed_cases) - 20} 个失败案例")

        # 按失败阶段分组
        print(f"\n[9] 失败阶段统计:")
        by_stage = defaultdict(int)
        for fail in failed_cases:
            by_stage[fail['stage']] += 1

        for stage, count in by_stage.items():
            print(f"    {stage}: {count} 个")

    print("\n" + "=" * 100)
    if failed == 0:
        print("[SUCCESS] 所有测试通过！三段式公式系统完全可靠！")
        print("=" * 100)
        return 0
    else:
        print(f"[FAIL] 有 {failed} 个测试失败，需要修复")
        print("=" * 100)
        return 1

if __name__ == "__main__":
    sys.exit(main())
