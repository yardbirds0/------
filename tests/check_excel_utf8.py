# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from pathlib import Path

def main():
    excel_file = Path("d:/Code/快报填写程序/（科电）国资委、财政快报模板-纯净版 的副本.xlsx")

    print("="  * 80)
    print("Checking Excel column headers - UTF8 output")
    print("=" * 80)

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb["财政局-快报"]

    print(f"\nSheet name: {sheet.title}")
    print(f"Dimensions: {sheet.dimensions}")

    print("\n--- Row 1 (Header) ---")
    for col_idx in range(1, 6):
        cell = sheet.cell(row=1, column=col_idx)
        print(f"Col {col_idx} ({cell.column_letter}): {repr(cell.value)}")

    print("\n--- Searching for row number column ---")
    for col_idx in range(1, 10):
        cell = sheet.cell(row=1, column=col_idx)
        if cell.value and ('行次' in str(cell.value) or 'rownum' in str(cell.value).lower()):
            print(f">>> FOUND! Column {col_idx} ({cell.column_letter}): {repr(cell.value)}")

    wb.close()

if __name__ == "__main__":
    main()
