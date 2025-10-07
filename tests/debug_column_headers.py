"""
调试脚本：检查科目余额表的列头识别情况
"""
import sys
sys.path.insert(0, 'd:\\Code\\快报填写程序-修改UI前(2)')

import openpyxl
from modules.table_schema_analyzer import TableSchemaAnalyzer

# 加载Excel文件
workbook_path = 'd:\\Code\\快报填写程序-修改UI前(2)\\sample.xlsx'
print(f"加载工作簿: {workbook_path}")
workbook = openpyxl.load_workbook(workbook_path, data_only=True)

# 列出所有工作表
print(f"\n工作表列表:")
for idx, sheet_name in enumerate(workbook.sheetnames):
    print(f"  {idx+1}. {sheet_name}")

# 查找科目余额表
target_sheet = None
for sheet_name in workbook.sheetnames:
    if '科目余额' in sheet_name or '余额' in sheet_name:
        target_sheet = sheet_name
        break

if not target_sheet:
    print("\n未找到科目余额表")
    sys.exit(1)

print(f"\n分析工作表: {target_sheet}")
sheet = workbook[target_sheet]

# 创建分析器
analyzer = TableSchemaAnalyzer()

# 分析表头
print("\n开始分析表格模式...")
table_schema = analyzer.analyze_table_schema(sheet)

print(f"\n表格类型: {table_schema.table_type.value}")
print(f"表头起始行: {table_schema.header_start_row}")
print(f"列头行数: {table_schema.header_rows}")
print(f"数据开始行: {table_schema.data_start_row}")
print(f"名称列: {table_schema.name_columns}")
print(f"编码列: {table_schema.code_columns}")

print(f"\n识别到的数据列 (共 {len(table_schema.data_columns)} 列):")
for col_info in table_schema.data_columns:
    print(f"  列{col_info.column_letter} ({col_info.column_index}):")
    print(f"    主列头: '{col_info.primary_header}'")
    print(f"    次列头: '{col_info.secondary_header}'")
    print(f"    显示名称: '{col_info.display_name}'")
    print(f"    normalized_key: '{col_info.normalized_key}'")
    print(f"    数据类型: {col_info.data_type}")
    print(f"    是否数值: {col_info.is_numeric}")
    print(f"    是否占位: {col_info.is_placeholder}")
    print()

if len(table_schema.data_columns) == 0:
    print("⚠️ 警告: 没有识别到任何数据列！")
    print("\n让我们检查前几行的原始数据:")
    for row in range(1, min(6, sheet.max_row + 1)):
        row_data = []
        for col in range(1, min(11, sheet.max_column + 1)):
            cell = sheet.cell(row, col)
            value = cell.value
            row_data.append(f"{openpyxl.utils.get_column_letter(col)}: {value}")
        print(f"  第{row}行: {', '.join(row_data)}")

workbook.close()
