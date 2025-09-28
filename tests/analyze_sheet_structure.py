#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析数据来源表sheet结构的测试脚本
用于诊断数据显示不全和空数据的问题
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from typing import Dict, List, Any
import json
from datetime import datetime

class SheetStructureAnalyzer:
    """Sheet结构分析器"""

    def __init__(self, excel_file_path: str):
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.analysis_results = {}

    def load_workbook(self):
        """加载工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file_path, data_only=True)
            print(f"[OK] 成功加载工作簿: {self.excel_file_path}")
            print(f"[INFO] 工作表数量: {len(self.workbook.sheetnames)}")
            print(f"[INFO] 工作表列表: {self.workbook.sheetnames}")
            return True
        except Exception as e:
            print(f"[ERROR] 加载工作簿失败: {e}")
            return False

    def analyze_sheet_structure(self, sheet_name: str) -> Dict[str, Any]:
        """分析单个sheet的结构"""
        if not self.workbook:
            return {}

        try:
            sheet = self.workbook[sheet_name]
            print(f"\n" + "="*60)
            print(f"[ANALYZE] 分析工作表: {sheet_name}")
            print("="*60)

            analysis = {
                "sheet_name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "dimension": sheet.dimensions,
                "data_rows": 0,
                "empty_rows": 0,
                "header_analysis": {},
                "data_analysis": {},
                "sample_data": []
            }

            # 基本信息
            print(f"[INFO] 最大行数: {sheet.max_row}")
            print(f"[INFO] 最大列数: {sheet.max_column}")
            print(f"[INFO] 数据范围: {sheet.dimensions}")

            # 分析列头结构
            print(f"\n[HEADER] 列头分析:")
            header_rows = []
            for row_num in range(1, min(6, sheet.max_row + 1)):  # 检查前5行
                row_data = []
                has_data = False
                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    row_data.append(value)
                    if value is not None and str(value).strip():
                        has_data = True

                if has_data:
                    header_rows.append({
                        "row_num": row_num,
                        "data": row_data,
                        "non_empty_count": sum(1 for v in row_data if v is not None and str(v).strip())
                    })
                    print(f"  第{row_num}行 ({len([v for v in row_data if v is not None and str(v).strip()])}个非空): {row_data[:10]}{'...' if len(row_data) > 10 else ''}")

            analysis["header_analysis"] = header_rows

            # 分析数据行
            print(f"\n[DATA] 数据行分析:")
            data_rows = 0
            empty_rows = 0
            sample_data_rows = []

            # 从第6行开始检查数据（假设前5行是表头）
            for row_num in range(6, min(sheet.max_row + 1, 106)):  # 检查前100行数据
                row_data = []
                has_data = False

                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    row_data.append(value)
                    if value is not None and str(value).strip():
                        has_data = True

                if has_data:
                    data_rows += 1
                    if len(sample_data_rows) < 10:  # 保存前10行作为样本
                        sample_data_rows.append({
                            "row_num": row_num,
                            "data": row_data,
                            "non_empty_count": sum(1 for v in row_data if v is not None and str(v).strip())
                        })
                else:
                    empty_rows += 1

            analysis["data_rows"] = data_rows
            analysis["empty_rows"] = empty_rows
            analysis["sample_data"] = sample_data_rows

            print(f"  [INFO] 数据行数: {data_rows}")
            print(f"  [INFO] 空行数: {empty_rows}")
            print(f"  [SAMPLE] 样本数据（前10行）:")

            for sample in sample_data_rows[:5]:  # 只显示前5行样本
                print(f"    第{sample['row_num']}行 ({sample['non_empty_count']}个非空): {sample['data'][:8]}{'...' if len(sample['data']) > 8 else ''}")

            # 分析第一列的数据特征（可能是科目代码）
            print(f"\n[FIRST_COL] 第一列数据特征分析:")
            first_col_data = []
            for row_num in range(1, min(sheet.max_row + 1, 51)):  # 检查前50行
                cell = sheet.cell(row=row_num, column=1)
                if cell.value is not None and str(cell.value).strip():
                    first_col_data.append({
                        "row": row_num,
                        "value": cell.value,
                        "type": type(cell.value).__name__
                    })

            print(f"  [INFO] 第一列非空数据数量: {len(first_col_data)}")
            if first_col_data:
                print(f"  [SAMPLE] 第一列样本数据:")
                for i, data in enumerate(first_col_data[:10]):
                    print(f"    第{data['row']}行: {data['value']} ({data['type']})")

            analysis["first_column_analysis"] = first_col_data

            # 检查合并单元格
            merged_ranges = list(sheet.merged_cells.ranges)
            if merged_ranges:
                print(f"\n[MERGED] 合并单元格: {len(merged_ranges)}个")
                for i, merged_range in enumerate(merged_ranges[:5]):  # 只显示前5个
                    print(f"  {i+1}. {merged_range}")

            analysis["merged_cells"] = [str(r) for r in merged_ranges]

            return analysis

        except Exception as e:
            print(f"[ERROR] 分析工作表 {sheet_name} 失败: {e}")
            return {}

    def analyze_all_sheets(self):
        """分析所有工作表"""
        if not self.workbook:
            print("[ERROR] 工作簿未加载")
            return

        print(f"\n[START] 开始分析所有工作表...")
        print(f"[TIME] 分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        for sheet_name in self.workbook.sheetnames:
            analysis = self.analyze_sheet_structure(sheet_name)
            if analysis:
                self.analysis_results[sheet_name] = analysis

        # 生成总结报告
        self.generate_summary_report()

    def generate_summary_report(self):
        """生成总结报告"""
        print(f"\n" + "="*80)
        print(f"[SUMMARY] 总结报告")
        print("="*80)

        total_sheets = len(self.analysis_results)
        sheets_with_data = 0
        sheets_empty = 0
        total_data_rows = 0

        print(f"[STATS] 工作表统计:")

        for sheet_name, analysis in self.analysis_results.items():
            data_rows = analysis.get("data_rows", 0)
            max_row = analysis.get("max_row", 0)

            if data_rows > 0:
                sheets_with_data += 1
                total_data_rows += data_rows
                print(f"  [OK] {sheet_name}: {data_rows}行数据 (总{max_row}行)")
            else:
                sheets_empty += 1
                print(f"  [EMPTY] {sheet_name}: 无数据 (总{max_row}行)")

        print(f"\n[TOTAL] 汇总:")
        print(f"  [INFO] 总工作表数: {total_sheets}")
        print(f"  [OK] 有数据的工作表: {sheets_with_data}")
        print(f"  [EMPTY] 空数据的工作表: {sheets_empty}")
        print(f"  [INFO] 总数据行数: {total_data_rows}")

        # 识别可能的问题
        print(f"\n[ISSUES] 潜在问题:")

        for sheet_name, analysis in self.analysis_results.items():
            issues = []

            if analysis.get("data_rows", 0) == 0:
                issues.append("无数据行")

            if analysis.get("max_row", 0) > 1000:
                issues.append("行数过多")

            header_analysis = analysis.get("header_analysis", [])
            if len(header_analysis) > 3:
                issues.append("表头行数过多")

            if issues:
                print(f"  [WARNING] {sheet_name}: {', '.join(issues)}")

    def save_analysis_to_file(self, output_file: str = "sheet_analysis_report.json"):
        """保存分析结果到文件"""
        try:
            output_path = os.path.join(os.path.dirname(__file__), output_file)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.analysis_results, f, ensure_ascii=False, indent=2, default=str)
            print(f"\n[SAVE] 分析报告已保存到: {output_path}")
        except Exception as e:
            print(f"[ERROR] 保存分析报告失败: {e}")

def main():
    """主函数"""
    # 查找Excel文件
    excel_file = None
    current_dir = os.path.dirname(os.path.dirname(__file__))

    # 寻找Excel文件
    for file in os.listdir(current_dir):
        if file.endswith('.xlsx') and '科电' in file:
            excel_file = os.path.join(current_dir, file)
            break

    if not excel_file:
        print("[ERROR] 未找到Excel文件")
        return

    print(f"[TARGET] 分析Excel文件: {excel_file}")

    # 创建分析器并分析
    analyzer = SheetStructureAnalyzer(excel_file)

    if analyzer.load_workbook():
        analyzer.analyze_all_sheets()
        analyzer.save_analysis_to_file()

    print(f"\n[DONE] 分析完成!")

if __name__ == "__main__":
    main()